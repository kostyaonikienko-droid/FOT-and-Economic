import sys
import json
import datetime
import calendar
import os
import re

from typing import List, Dict, Tuple, Optional, Any, Set
from enum import Enum

import openpyxl
from openpyxl.utils import get_column_letter

# Настройка Qt окружения для matplotlib
os.environ['QT_API'] = 'PySide6'

import matplotlib
matplotlib.use('QtAgg')  # или 'Qt5Agg' (оба подходят)
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import numpy as np

import os

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTableWidget, QTableWidgetItem, QMenuBar, QMenu, QDialog,
    QFormLayout, QLineEdit, QDialogButtonBox, QComboBox, QPushButton,
    QHeaderView, QStyledItemDelegate, QSpinBox, QLabel, QMessageBox,
    QFileDialog, QToolBar, QTabWidget, QDoubleSpinBox, QGroupBox,
    QGridLayout, QCompleter, QRadioButton, QButtonGroup, QDateEdit,
    QAbstractItemView, QInputDialog, QTextEdit, QLabel
)
from PySide6.QtGui import QAction, QKeySequence, QColor, QFont, QBrush, QIcon
from PySide6.QtCore import QTimer, QDate
from openpyxl.chart import LineChart, Reference, PieChart

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

import pandas as pd
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QComboBox, QDoubleSpinBox,
    QSpinBox, QCheckBox, QPushButton, QTableWidget, QTableWidgetItem,
    QHeaderView, QMessageBox, QFileDialog
)
from PySide6.QtCore import Qt

class ActionLog:
    def __init__(self, filename="actions.json"):
        self.filename = filename
        self.actions = []
        self.load()

    def load(self):
        if os.path.exists(self.filename):
            try:
                with open(self.filename, "r", encoding="utf-8") as f:
                    self.actions = json.load(f)
            except:
                self.actions = []
        else:
            self.actions = []

    def save(self):
        try:
            with open(self.filename, "w", encoding="utf-8") as f:
                json.dump(self.actions[-1000:], f, ensure_ascii=False, indent=2)  # храним последние 1000 записей
        except Exception as e:
            print(f"Ошибка сохранения лога: {e}")

    def add(self, action_type, description, user="Пользователь"):
        entry = {
            "timestamp": datetime.datetime.now().isoformat(),
            "user": user,
            "type": action_type,
            "description": description
        }
        self.actions.append(entry)
        self.save()

    def clear(self):
        self.actions = []
        self.save()

    def get_last(self, n=10):
        return self.actions[-n:]

class LogDialog(QDialog):
    def __init__(self, log: ActionLog, parent=None):
        super().__init__(parent)

        self.resize(800, 400)

        layout = QVBoxLayout(self)

        self.table = QTableWidget(len(log.actions), 3)
        self.table.setHorizontalHeaderLabels(["Время", "Тип", "Описание"])
        for i, entry in enumerate(log.actions):
            self.table.setItem(i, 0, QTableWidgetItem(entry["timestamp"][:19]))
            self.table.setItem(i, 1, QTableWidgetItem(entry["type"]))
            self.table.setItem(i, 2, QTableWidgetItem(entry["description"]))
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.table)

        btn_close = QPushButton("Закрыть")
        btn_close.clicked.connect(self.accept)
        layout.addWidget(btn_close)

class BonusDialog(QDialog):
    def __init__(self, project, parent=None):
        super().__init__(parent)
        self.project = project

        self.resize(1000, 650)

        layout = QVBoxLayout(self)

        # Информационная строка с месяцем (не редактируется)
        month_info_layout = QHBoxLayout()
        month_info_layout.addWidget(QLabel("Месяц расчёта:"))
        self.month_label = QLabel()
        month_info_layout.addWidget(self.month_label)
        month_info_layout.addStretch()
        layout.addLayout(month_info_layout)

        # Панель управления
        control_layout = QHBoxLayout()
        control_layout.addWidget(QLabel("Фонд премии (руб.):"))
        self.bonus_fund_spin = QDoubleSpinBox()
        self.bonus_fund_spin.setRange(0, 1e9)
        self.bonus_fund_spin.setValue(0)
        control_layout.addWidget(self.bonus_fund_spin)

        control_layout.addWidget(QLabel("Процент от оклада:"))
        self.percent_spin = QSpinBox()
        self.percent_spin.setRange(0, 100)
        self.percent_spin.setValue(35)
        self.percent_spin.setSuffix("%")
        control_layout.addWidget(self.percent_spin)

        self.round_check = QCheckBox("Округлять итог до десятков")
        control_layout.addWidget(self.round_check)

        layout.addLayout(control_layout)

        calc_btn = QPushButton("Рассчитать")
        calc_btn.clicked.connect(self.calculate_bonus)
        layout.addWidget(calc_btn)

        # Таблица
        self.table = QTableWidget(0, 8)
        headers = ["ФИО", "Таб.№", "Отработано дней", "Тариф, руб.",
                   "Сумма за отраб. время, руб.", "Доплата, руб.", "Вычет, руб.", "Итого, руб."]
        self.table.setHorizontalHeaderLabels(headers)
        # Настройка ширины столбцов
        self.table.setColumnWidth(0, 250)  # ФИО
        self.table.setColumnWidth(1, 100)  # Таб.№
        self.table.setColumnWidth(2, 120)  # Отработано дней
        self.table.setColumnWidth(3, 120)  # Тариф, руб.
        self.table.setColumnWidth(4, 180)  # Сумма за отраб. время, руб.
        self.table.setColumnWidth(5, 120)  # Доплата, руб.
        self.table.setColumnWidth(6, 120)  # Вычет, руб.
        self.table.setColumnWidth(7, 120)  # Итого, руб.
        # Разрешить пользователю изменять ширину
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked)
        self.table.cellChanged.connect(self.on_cell_changed)
        layout.addWidget(self.table)

        # Кнопки экспорта и закрытия
        btn_layout = QHBoxLayout()
        export_btn = QPushButton("Экспорт в Excel")
        export_btn.clicked.connect(self.export_to_excel)
        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(self.accept)
        btn_layout.addStretch()
        btn_layout.addWidget(export_btn)
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)

        # Статусная строка
        self.status_label = QLabel()
        self.status_label.setStyleSheet("padding: 5px; color: gray;")
        layout.addWidget(self.status_label)

        # Инициализация данных
        self.update_data()

    def is_driver(self, emp):
        return "водитель" in emp.position.lower()

    def is_excluded(self, emp):
        """Проверяет, исключён ли сотрудник из расчёта премии."""
        pos_lower = emp.position.lower().strip()
        return "и.о. начальника" in pos_lower or "начальник отдела" in pos_lower

    def update_data(self):
        year, month = self.project.current_month
        self.month_label.setText(f"{month:02d}.{year}")
        norm = self.project.get_norm(year, month)
        data = self.project.get_current_data()

        # Фильтруем сотрудников: исключаем и.о. начальника
        self.filtered_indices = []
        self.filtered_employees = []
        for idx, emp in enumerate(self.project.employees):
            if self.is_excluded(emp):
                continue
            self.filtered_indices.append(idx)
            self.filtered_employees.append(emp)

        # Отработанные дни
        emp_fact_days = []
        for emp_idx in self.filtered_indices:
            fact = 0
            for day in range(1, self.project.days_in_month(year, month)+1):
                code, _, _ = data.get((emp_idx, day), ("", "", ""))
                if code == 'Ф':
                    fact += 1
            emp_fact_days.append(fact)

        # Заполняем таблицу
        self.table.blockSignals(True)
        self.table.setRowCount(len(self.filtered_employees))
        for i, emp in enumerate(self.filtered_employees):
            self.table.setItem(i, 0, QTableWidgetItem(emp.fio))
            self.table.setItem(i, 1, QTableWidgetItem(emp.tab_num))
            self.table.setItem(i, 2, QTableWidgetItem(str(emp_fact_days[i])))
            for col in range(3, 8):
                self.table.setItem(i, col, QTableWidgetItem(""))
            item_ded = QTableWidgetItem("0")
            item_ded.setFlags(item_ded.flags() | Qt.ItemIsEditable)
            self.table.setItem(i, 6, item_ded)
        self.table.blockSignals(False)

        # Статус
        if len(self.filtered_employees) < len(self.project.employees):
            excluded = len(self.project.employees) - len(self.filtered_employees)
            self.status_label.setText(f"Исключено сотрудников: {excluded}")
            self.status_label.setStyleSheet("padding: 5px; color: blue;")
        else:
            self.status_label.clear()

        self.calculate_bonus()

    def calculate_bonus(self):
        year, month = self.project.current_month
        norm = self.project.get_norm(year, month)
        percent = self.percent_spin.value() / 100
        bonus_fund = self.bonus_fund_spin.value()

        data = self.project.get_current_data()
        days_in_month = self.project.days_in_month(year, month)

        emp_data = []
        total_fact_all = 0
        total_salary_part_all = 0
        total_fact_non_driver = 0

        for i, emp in enumerate(self.filtered_employees):
            emp_idx = self.filtered_indices[i]
            fact = 0
            for day in range(1, days_in_month+1):
                code, _, _ = data.get((emp_idx, day), ("", "", ""))
                if code == 'Ф':
                    fact += 1
            tariff = round(emp.salary * percent)
            salary_part = round(tariff / norm * fact) if fact > 0 and norm > 0 else 0
            emp_data.append((i, emp.fio, emp.tab_num, fact, tariff, salary_part, self.is_driver(emp)))
            total_fact_all += fact
            total_salary_part_all += salary_part
            if not self.is_driver(emp):
                total_fact_non_driver += fact

        remainder = bonus_fund - total_salary_part_all
        if remainder < 0:
            self.status_label.setText("ВНИМАНИЕ! Фонд премии меньше общей суммы по тарифу. Доплата будет отрицательной.")
            self.status_label.setStyleSheet("padding: 5px; color: red; font-weight: bold;")
        else:
            self.status_label.clear()

        coeff = round(remainder / total_fact_non_driver, 2) if total_fact_non_driver > 0 else 0

        self.table.blockSignals(True)
        for i, (idx, fio, tab, fact, tariff, salary_part, is_driver) in enumerate(emp_data):
            self.table.setItem(i, 0, QTableWidgetItem(fio))
            self.table.setItem(i, 1, QTableWidgetItem(tab))
            self.table.setItem(i, 2, QTableWidgetItem(str(fact)))
            self.table.setItem(i, 3, QTableWidgetItem(f"{tariff:,.0f}"))
            self.table.setItem(i, 4, QTableWidgetItem(f"{salary_part:,.0f}"))

            extra = 0 if is_driver else round(coeff * fact)
            self.table.setItem(i, 5, QTableWidgetItem(f"{extra:,.0f}"))

            ded_item = self.table.item(i, 6)
            try:
                deduction = float(ded_item.text().replace(',', '')) if ded_item and ded_item.text() else 0
            except:
                deduction = 0

            total = salary_part + extra - deduction
            if self.round_check.isChecked():
                total = round(total / 10) * 10
            self.table.setItem(i, 7, QTableWidgetItem(f"{total:,.0f}"))
        self.table.blockSignals(False)

    def on_cell_changed(self, row, col):
        if col == 6:
            self.table.blockSignals(True)
            try:
                salary_part = float(self.table.item(row, 4).text().replace(',', ''))
                extra = float(self.table.item(row, 5).text().replace(',', ''))
                deduction = float(self.table.item(row, 6).text().replace(',', ''))
            except:
                self.table.blockSignals(False)
                return
            total = salary_part + extra - deduction
            if self.round_check.isChecked():
                total = round(total / 10) * 10
            self.table.setItem(row, 7, QTableWidgetItem(f"{total:,.0f}"))
            self.table.blockSignals(False)

    def export_to_excel(self):
        fname, _ = QFileDialog.getSaveFileName(self, "Сохранить расчёт премии",
                                               "", "Excel Files (*.xlsx)")
        if not fname:
            return
        if not fname.endswith('.xlsx'):
            fname += '.xlsx'

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, NamedStyle
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Премия ППП"

        # Заголовки с единицами измерения
        headers = [
            "ФИО", "Таб.№", "Отработано дней (дн.)", "Тариф (руб.)",
            "Сумма за отраб. время (руб.)", "Доплата (руб.)",
            "Вычет (руб.)", "Итого (руб.)"
        ]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Данные из таблицы
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item:
                    # Для числовых колонок (3-7) пытаемся получить число
                    if col in (3,4,5,6,7):
                        try:
                            val = float(item.text().replace(',', ''))
                            cell = ws.cell(row=row+2, column=col+1, value=val)
                            cell.number_format = '#,##0.00'  # два знака после запятой
                        except:
                            ws.cell(row=row+2, column=col+1, value=item.text())
                    else:
                        ws.cell(row=row+2, column=col+1, value=item.text())
                else:
                    ws.cell(row=row+2, column=col+1, value="")

        # Итоговая строка (суммируем из таблицы, но можно и посчитать заново)
        total_row = self.table.rowCount() + 2
        ws.cell(row=total_row, column=1, value="ИТОГО")
        # Суммируем по колонкам
        for col in range(2, 8):  # колонки B-H (2..7)
            col_letter = get_column_letter(col+1)
            formula = f"=SUM({col_letter}2:{col_letter}{self.table.rowCount()+1})"
            cell = ws.cell(row=total_row, column=col+1, value=formula)
            cell.number_format = '#,##0.00'

        # Информация о фонде премии
        fund_row = total_row + 2
        ws.cell(row=fund_row, column=1, value="Фонд премии (руб.):")
        cell = ws.cell(row=fund_row, column=2, value=self.bonus_fund_spin.value())
        cell.number_format = '#,##0.00'

        # Настройка ширины колонок
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15

        wb.save(fname)
        QMessageBox.information(self, "Экспорт", f"Расчёт сохранён в {fname}")


# Базовый класс для всех диалогов
class BaseDialog(QDialog):
    def __init__(self, title: str, parent=None):
        super().__init__(parent)
        #self.setWindowTitle(title)
        self.setModal(True)
        self.main_layout = QVBoxLayout(self)
        self.form_layout = QFormLayout()
        self.main_layout.addLayout(self.form_layout)
        self.button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        self.main_layout.addWidget(self.button_box)


# ------------------------------------------------------------
# Перечисление типов доплат
class AllowanceType(Enum):
    FIXED = "Фиксированная"
    PROPORTIONAL = "Пропорционально отработанному времени"

    def to_str(self):
        return self.value

    @staticmethod
    def from_str(s: str):
        for t in AllowanceType:
            if t.value == s:
                return t
        return AllowanceType.FIXED


# ------------------------------------------------------------
# Класс для хранения одной доплаты
class AllowanceItem:
    def __init__(self, name: str = "", amount: float = 0.0,
                 type: AllowanceType = AllowanceType.FIXED):
        self.name = name
        self.amount = amount
        self.type = type

    def to_dict(self):
        return {
            "name": self.name,
            "amount": self.amount,
            "type": self.type.value
        }

    @classmethod
    def from_dict(cls, data):
        return cls(
            data.get("name", ""),
            float(data.get("amount", 0.0)),
            AllowanceType.from_str(data.get("type", "Фиксированная"))
        )


# ------------------------------------------------------------
# Класс для хранения записи графика отпусков


class VacationRecord:
    def __init__(self, employee_fio: str = "", employee_tab_num: str = "",
                 start_date: datetime.date = None, end_date: datetime.date = None,
                 code: str = "О", description: str = ""):
        self.employee_fio = employee_fio
        self.employee_tab_num = employee_tab_num
        self.start_date = start_date
        self.end_date = end_date
        self.code = code
        self.description = description

    def to_dict(self):
        return {
            "employee_fio": self.employee_fio,
            "employee_tab_num": self.employee_tab_num,
            "start_date": self.start_date.isoformat() if self.start_date else None,
            "end_date": self.end_date.isoformat() if self.end_date else None,
            "code": self.code,
            "description": self.description
        }

    @classmethod
    def from_dict(cls, data):
        start = datetime.date.fromisoformat(data["start_date"]) if data.get("start_date") else None
        end = datetime.date.fromisoformat(data["end_date"]) if data.get("end_date") else None
        return cls(
            data.get("employee_fio", ""),
            data.get("employee_tab_num", ""),
            start,
            end,
            data.get("code", "О"),
            data.get("description", "")
        )

# ------------------------------------------------------------
# Класс сотрудника с данными для расчёта зарплаты
class Employee:
    def __init__(self, fio: str = "", tab_num: str = "", position: str = "",
                 salary: float = 0.0, allowances: List[AllowanceItem] = None,
                 avg_salary_by_month: Dict[Tuple[int, int], float] = None,
                 avg_salary_with_coeff_by_month: Dict[Tuple[int, int], float] = None,
                 avg_sick_leave: float = 0.0,
                 experience_years: int = 0):
        self.fio = fio
        self.tab_num = tab_num
        self.position = position
        self.salary = salary
        self.allowances = allowances if allowances is not None else []
        self.avg_salary_by_month = avg_salary_by_month if avg_salary_by_month is not None else {}
        self.avg_salary_with_coeff_by_month = avg_salary_with_coeff_by_month if avg_salary_with_coeff_by_month is not None else {}
        self.avg_sick_leave = avg_sick_leave
        self.experience_years = experience_years

    def get_avg_salary_for_month(self, year: int, month: int) -> Tuple[float, Optional[Tuple[int, int]]]:
        available = [(y, m) for (y, m), val in self.avg_salary_by_month.items() if val != 0]
        available.sort(reverse=True)
        for y, m in available:
            if (y < year) or (y == year and m <= month):
                return self.avg_salary_by_month[(y, m)], (y, m)
        return 0.0, None

    def get_sick_pay_coefficient(self) -> float:
        if self.experience_years < 5:
            return 0.6
        elif self.experience_years < 8:
            return 0.8
        else:
            return 1.0

    def total_allowance(self, norm_days: int, fact_days: int) -> float:
        total = 0.0
        for a in self.allowances:
            if a.type == AllowanceType.FIXED:
                total += a.amount
            else:
                if norm_days > 0:
                    total += a.amount / norm_days * fact_days
                else:
                    total += a.amount
        return total

    def allowance_details(self, norm_days: int, fact_days: int) -> List[Tuple[str, float, str, float]]:
        details = []
        for a in self.allowances:
            if a.type == AllowanceType.FIXED:
                accrued = a.amount
            else:
                accrued = a.amount / norm_days * fact_days if norm_days > 0 else a.amount
            details.append((a.name, a.amount, a.type.value, accrued))
        return details

    def to_dict(self):
        return {
            "fio": self.fio,
            "tab_num": self.tab_num,
            "position": self.position,
            "salary": self.salary,
            "allowances": [a.to_dict() for a in self.allowances],
            "avg_salary_by_month": {f"{y},{m}": v for (y, m), v in self.avg_salary_by_month.items()},
            "avg_salary_with_coeff_by_month": {f"{y},{m}": v for (y, m), v in self.avg_salary_with_coeff_by_month.items()},
            "avg_sick_leave": self.avg_sick_leave,
            "experience_years": self.experience_years
        }

    @classmethod
    def from_dict(cls, data):
        allowances_data = data.get("allowances", [])
        allowances = [AllowanceItem.from_dict(a) for a in allowances_data]
        avg_salary_dict = {}
        for key, val in data.get("avg_salary_by_month", {}).items():
            y, m = key.split(",")
            avg_salary_dict[(int(y), int(m))] = float(val)
        avg_salary_coeff_dict = {}
        for key, val in data.get("avg_salary_with_coeff_by_month", {}).items():
            y, m = key.split(",")
            avg_salary_coeff_dict[(int(y), int(m))] = float(val)
        avg_sick_leave = float(data.get("avg_sick_leave", 0.0))
        experience = data.get("experience_years", 0)
        return cls(
            data["fio"],
            data["tab_num"],
            data["position"],
            float(data.get("salary", 0.0)),
            allowances,
            avg_salary_dict,
            avg_salary_coeff_dict,
            avg_sick_leave,
            experience
        )

# ------------------------------------------------------------
# Класс базы сотрудников (для хранения постоянного списка)
class EmployeeDatabase:
    def __init__(self, filename="employees.json"):
        self.filename = filename
        self.employees: List[Employee] = []
        self.load()

    def load(self):
        if os.path.exists(self.filename):
            try:
                with open(self.filename, "r", encoding="utf-8") as f:
                    data = json.load(f)
                self.employees = [Employee.from_dict(e) for e in data]
            except:
                self.employees = []

    def save(self):
        try:
            with open(self.filename, "w", encoding="utf-8") as f:
                json.dump([emp.to_dict() for emp in self.employees], f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Ошибка сохранения базы сотрудников: {e}")

    def add_employee(self, emp: Employee):
        if emp.tab_num:
            for e in self.employees:
                if e.tab_num == emp.tab_num:
                    self.employees.remove(e)
                    break
        self.employees.append(emp)

    def remove_employee(self, emp: Employee):
        if emp in self.employees:
            self.employees.remove(emp)

    def clear(self):
        self.employees.clear()

    def find_by_fio(self, fio: str) -> Optional[Employee]:
        for e in self.employees:
            if e.fio == fio:
                return e
        return None

    def find_by_tab_num(self, tab_num: str) -> Optional[Employee]:
        for e in self.employees:
            if e.tab_num == tab_num:
                return e
        return None

class VacationDatabase:
    def __init__(self, filename="vacations.json"):
        self.filename = filename
        self.vacations: List[VacationRecord] = []
        self.load()

    def load(self):
        if os.path.exists(self.filename):
            try:
                with open(self.filename, "r", encoding="utf-8") as f:
                    data = json.load(f)
                self.vacations = [VacationRecord.from_dict(d) for d in data]
            except:
                self.vacations = []

    def save(self):
        try:
            with open(self.filename, "w", encoding="utf-8") as f:
                json.dump([v.to_dict() for v in self.vacations], f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Ошибка сохранения базы отпусков: {e}")

    def add_vacation(self, rec: VacationRecord):
        # Можно добавить проверку на дубликат, но пока просто добавляем
        self.vacations.append(rec)

    def clear(self):
        self.vacations.clear()

# ------------------------------------------------------------
# Класс для хранения праздничных дней (с переносами)
class Holidays:
    def __init__(self, filename="holidays.json"):
        self.filename = filename
        self.data = {}
        self.load()

    def load(self):
        if os.path.exists(self.filename):
            try:
                with open(self.filename, "r", encoding="utf-8") as f:
                    self.data = json.load(f)
            except:
                self.create_default()
        else:
            self.create_default()
            self.save()

    def save(self):
        with open(self.filename, "w", encoding="utf-8") as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)

    def _add_holiday_transfers(self, year):
        """Добавляет переносы праздников, выпавших на выходные."""
        year_str = str(year)
        holidays = set(self.data.get(year_str, []))
        new_holidays = set()
        for date_str in holidays:
            date = datetime.date.fromisoformat(date_str)
            if date.weekday() >= 5:  # суббота или воскресенье
                next_day = date + datetime.timedelta(days=1)
                while next_day.weekday() >= 5 or next_day.isoformat() in holidays:
                    next_day += datetime.timedelta(days=1)
                new_holidays.add(next_day.isoformat())
        if new_holidays:
            self.data[year_str].extend(list(new_holidays))
            self.data[year_str] = list(set(self.data[year_str]))

    def create_default(self):
        self.data = {}
        fixed = [
            (1, 1), (1, 2), (1, 3), (1, 4), (1, 5), (1, 6), (1, 7), (1, 8),
            (2, 23), (3, 8), (5, 1), (5, 9), (6, 12), (11, 4)
        ]
        for year in range(2026, 2036):
            year_str = str(year)
            self.data[year_str] = []
            for month, day in fixed:
                try:
                    date = datetime.date(year, month, day)
                    self.data[year_str].append(date.isoformat())
                except ValueError:
                    pass
            self._add_holiday_transfers(year)

    def is_holiday(self, date: datetime.date) -> bool:
        year_str = str(date.year)
        if year_str in self.data:
            return date.isoformat() in self.data[year_str]
        return False


# ------------------------------------------------------------
# Класс для маппинга кодов табеля
class CodeMapping:
    def __init__(self):
        self.vacation = {'О', 'ОД', 'ОУ'}
        self.business_trip = {'К'}
        self.sick_leave = {'Б'}
        self.unpaid = {'А', 'НН', 'ДО'}
        self.special_codes = {
            'С': {'type': 'overtime', 'payment_multiplier': 1.5, 'name': 'Сверхурочные'},
            'РП': {'type': 'weekend_work', 'payment_multiplier': 2.0, 'name': 'Работа в выходной'},
            'КРВ': {'type': 'business_weekend', 'payment_multiplier': 2.0, 'name': 'Командировка в выходной'},
        }
        self.special_payment_option = {}  # код -> 'double' or 'day_off' (глобальная настройка)

    def get_category(self, code):
        code = str(code).strip()
        if code in self.vacation:
            return 'vacation'
        if code in self.business_trip:
            return 'business'
        if code in self.sick_leave:
            return 'sick'
        if code in self.unpaid:
            return 'unpaid'
        return None

    def is_special(self, code):
        return code in self.special_codes

    def get_special_info(self, code):
        return self.special_codes.get(code, None)

    def to_dict(self):
        return {
            "vacation": list(self.vacation),
            "business_trip": list(self.business_trip),
            "sick_leave": list(self.sick_leave),
            "unpaid": list(self.unpaid),
            "special_payment_option": self.special_payment_option
        }

    @classmethod
    def from_dict(cls, data):
        mapping = cls()
        mapping.vacation = set(data.get("vacation", []))
        mapping.business_trip = set(data.get("business_trip", []))
        mapping.sick_leave = set(data.get("sick_leave", []))
        mapping.unpaid = set(data.get("unpaid", []))
        mapping.special_payment_option = data.get("special_payment_option", {})
        return mapping


# ------------------------------------------------------------
# Класс проекта (табель + данные для ФОТ + график отпусков + несколько месяцев)
class Project:
    def __init__(self):
        self.organization: str = "ФБУ \"Пермскй ЦСМ\""
        self.department: str = "отдел физико-химических измерений"
        self.okpo: str = "02567679"
        self.resp_dolgnost: str = "И.о. начальника отдела"
        self.resp_fio: str = "Оникиенко К.С."
        self.resp_signature: str = ""
        self.executor_dolgnost: str = "Бухгалтер 2 категории"
        self.executor_fio: str = "Метусалло Е.В."
        self.executor_signature: str = ""
        self.employees: List[Employee] = []
        self.months_data: Dict[Tuple[int, int], Dict[Tuple[int, int], Tuple[str, str, str]]] = {}
        self.current_month: Tuple[int, int] = (datetime.date.today().year, datetime.date.today().month)
        self.norm_by_month: Dict[Tuple[int, int], int] = {}
        if self.current_month not in self.months_data:
            self.months_data[self.current_month] = {}
            self.norm_by_month[self.current_month] = self.days_in_month(*self.current_month)
        self.code_mapping: CodeMapping = CodeMapping()
        self.vacations: List[VacationRecord] = []
        self.earnings_history: Dict[Tuple[int, int, int], Dict] = {}  # (year, month, emp_idx) -> начисления
        # --- ДОБАВЛЕНО: хранение планов по кварталам ---
        self.plans: Dict[Tuple[int, int], Dict] = {}  # ключ (year, quarter) -> данные плана

    def ensure_current_month(self):
        """Убеждается, что текущий месяц существует в данных. Если нет – устанавливает первый доступный."""
        if self.current_month not in self.months_data:
            if self.months_data:
                # выбираем первый месяц в порядке возрастания
                self.current_month = sorted(self.months_data.keys())[0]
            else:
                # если нет ни одного месяца, создаём текущий
                today = datetime.date.today()
                self.current_month = (today.year, today.month)
                self.months_data[self.current_month] = {}

    def days_in_month(self, year: int = None, month: int = None) -> int:
        if year is None:
            year, month = self.current_month
        return calendar.monthrange(year, month)[1]

    def get_norm(self, year: int = None, month: int = None) -> int:
        if year is None:
            year, month = self.current_month
        return self.norm_by_month.get((year, month), self.days_in_month(year, month))

    def set_norm(self, value: int, year: int = None, month: int = None):
        if year is None:
            year, month = self.current_month
        self.norm_by_month[(year, month)] = value

    def get_current_data(self) -> Dict[Tuple[int, int], Tuple[str, str, str]]:
        return self.months_data.setdefault(self.current_month, {})

    def set_current_data(self, data: Dict[Tuple[int, int], Tuple[str, str, str]]):
        self.months_data[self.current_month] = data

    def apply_weekends_and_holidays(self, holidays: Holidays, year: int = None, month: int = None):
        if year is None:
            year, month = self.current_month
        data = self.months_data.setdefault((year, month), {})
        days = self.days_in_month(year, month)
        for emp_idx in range(len(self.employees)):
            for day in range(1, days + 1):
                if (emp_idx, day) in data:
                    code, hours, _ = data[(emp_idx, day)]
                    if code:
                        continue
                try:
                    date = datetime.date(year, month, day)
                except ValueError:
                    continue
                if date.weekday() >= 5 or holidays.is_holiday(date):
                    data[(emp_idx, day)] = ("В", "", "double")

    def apply_vacation_schedule(self, holidays: Holidays):
        emp_by_tab = {e.tab_num: idx for idx, e in enumerate(self.employees) if e.tab_num}
        emp_by_fio = {e.fio: idx for idx, e in enumerate(self.employees)}
        for vac in self.vacations:
            if not vac.start_date or not vac.end_date:
                continue
            emp_idx = None
            if vac.employee_tab_num and vac.employee_tab_num in emp_by_tab:
                emp_idx = emp_by_tab[vac.employee_tab_num]
            elif vac.employee_fio and vac.employee_fio in emp_by_fio:
                emp_idx = emp_by_fio[vac.employee_fio]
            else:
                continue
            current = vac.start_date
            while current <= vac.end_date:
                for (year, month), data in self.months_data.items():
                    if current.year == year and current.month == month:
                        day = current.day
                        data[(emp_idx, day)] = (vac.code, "", "double")
                current += datetime.timedelta(days=1)

    def to_dict(self):
        months_serialized = {}
        for (year, month), data in self.months_data.items():
            key = f"{year},{month}"
            months_serialized[key] = {f"{emp_idx},{day}": [code, hours, option] for (emp_idx, day), (code, hours, option) in data.items()}
        norms_serialized = {f"{y},{m}": norm for (y, m), norm in self.norm_by_month.items()}
        # --- ДОБАВЛЕНО: сериализация планов ---
        plans_serialized = {f"{y},{q}": plan for (y, q), plan in self.plans.items()}
        return {
            "organization": self.organization,
            "department": self.department,
            "okpo": self.okpo,
            "resp_dolgnost": self.resp_dolgnost,
            "resp_fio": self.resp_fio,
            "resp_signature": self.resp_signature,
            "executor_dolgnost": self.executor_dolgnost,
            "executor_fio": self.executor_fio,
            "executor_signature": self.executor_signature,
            "employees": [emp.to_dict() for emp in self.employees],
            "months_data": months_serialized,
            "norm_by_month": norms_serialized,
            "current_month": f"{self.current_month[0]},{self.current_month[1]}",
            "code_mapping": self.code_mapping.to_dict(),
            "vacations": [v.to_dict() for v in self.vacations],
            "earnings_history": {f"{y},{m},{e}": v for (y, m, e), v in self.earnings_history.items()},
            "plans": plans_serialized,   # <-- добавлено
        }

    @classmethod
    def from_dict(cls, data):
        proj = cls()
        proj.organization = data.get("organization", "")
        proj.department = data.get("department", "")
        proj.okpo = data.get("okpo", "")
        proj.resp_dolgnost = data.get("resp_dolgnost", "")
        proj.resp_fio = data.get("resp_fio", "")
        proj.resp_signature = data.get("resp_signature", "")
        proj.executor_dolgnost = data.get("executor_dolgnost", "")
        proj.executor_fio = data.get("executor_fio", "")
        proj.executor_signature = data.get("executor_signature", "")

        # Загрузка сотрудников с защитой
        emp_list = data.get("employees", [])
        if isinstance(emp_list, list):
            proj.employees = [Employee.from_dict(e) for e in emp_list if isinstance(e, dict)]

        # Загрузка данных по месяцам
        months_serialized = data.get("months_data", {})
        if isinstance(months_serialized, dict):
            for key, val in months_serialized.items():
                if not isinstance(key, str) or ',' not in key:
                    continue
                try:
                    year_str, month_str = key.split(",")
                    year = int(year_str)
                    month = int(month_str)
                except:
                    continue
                if not isinstance(val, dict):
                    continue
                data_dict = {}
                for cell_key, cell_val in val.items():
                    if not isinstance(cell_key, str) or ',' not in cell_key:
                        continue
                    try:
                        emp_idx_str, day_str = cell_key.split(",")
                        emp_idx = int(emp_idx_str)
                        day = int(day_str)
                    except:
                        continue
                    # Обработка cell_val – должно быть списком
                    if not isinstance(cell_val, (list, tuple)):
                        continue
                    if len(cell_val) == 2:
                        code, hours = cell_val
                        option = "double"
                    elif len(cell_val) >= 3:
                        code, hours, option = cell_val[:3]
                    else:
                        continue
                    data_dict[(emp_idx, day)] = (str(code) if code is not None else "",
                                                 str(hours) if hours is not None else "",
                                                 str(option) if option is not None else "double")
                proj.months_data[(year, month)] = data_dict

        # Загрузка норм
        norms_serialized = data.get("norm_by_month", {})
        if isinstance(norms_serialized, dict):
            for key, norm in norms_serialized.items():
                if not isinstance(key, str) or ',' not in key:
                    continue
                try:
                    y, m = key.split(",")
                    proj.norm_by_month[(int(y), int(m))] = int(norm) if isinstance(norm, (int, float)) else 0
                except:
                    continue

        # Текущий месяц
        cur = data.get("current_month", "")
        if cur and isinstance(cur, str) and ',' in cur:
            try:
                y, m = cur.split(",")
                proj.current_month = (int(y), int(m))
            except:
                proj.current_month = (datetime.date.today().year, datetime.date.today().month)
        else:
            proj.current_month = (datetime.date.today().year, datetime.date.today().month)

        # Загрузка маппинга
        mapping_data = data.get("code_mapping", {})
        if isinstance(mapping_data, dict):
            proj.code_mapping = CodeMapping.from_dict(mapping_data)

        # Загрузка отпусков
        vacations_list = data.get("vacations", [])
        if isinstance(vacations_list, list):
            proj.vacations = [VacationRecord.from_dict(v) for v in vacations_list if isinstance(v, dict)]

        # Загрузка истории
        earnings = data.get("earnings_history", {})
        if isinstance(earnings, dict):
            for key, val in earnings.items():
                if not isinstance(key, str) or ',' not in key:
                    continue
                try:
                    y, m, e = key.split(",")
                    proj.earnings_history[(int(y), int(m), int(e))] = val
                except:
                    continue

        # --- ДОБАВЛЕНО: загрузка планов ---
        plans_serialized = data.get("plans", {})
        if isinstance(plans_serialized, dict):
            for key, plan in plans_serialized.items():
                if not isinstance(key, str) or ',' not in key:
                    continue
                try:
                    y, q = key.split(",")
                    proj.plans[(int(y), int(q))] = plan
                except:
                    continue

        # Убедимся, что текущий месяц существует
        proj.ensure_current_month()
        return proj

# ------------------------------------------------------------
# Коды и делегаты
CODES = [
    "В", "Ф", "Б", "О", "К", "С", "РП", "ОР", "КРВ",  # приоритетные
    "Н", "Г", "П", "НН", "А", "ВУ", "ОУ",      # остальные
    "ЗН", "ЗП", "ЗС", "ОВ", "Д", "НОД", "ВВ", "ПД",
    "КВ", "РВ"
]
HOURS_CODES = ["Ф", "С", "Н", "К", "РП", "КРВ"]

class CodeDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        combo = QComboBox(parent)
        combo.addItems(CODES)
        combo.setEditable(True)
        combo.setMaxVisibleItems(15)
        combo.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        combo.lineEdit().setPlaceholderText("Код")
        combo.view().setMinimumWidth(150)
        return combo

    def setEditorData(self, editor, index):
        value = index.data(Qt.ItemDataRole.DisplayRole)
        if value:
            editor.setCurrentText(value)

    def setModelData(self, editor, model, index):
        model.setData(index, editor.currentText())


# ------------------------------------------------------------
# Диалог редактирования ячейки (с выбором опции для специальных кодов)
class CellEditDialog(BaseDialog):
    def __init__(self, mapping: CodeMapping, current_code="", current_hours="", current_option="double", parent=None):
        super().__init__("Редактирование ячейки", parent)
        self.mapping = mapping

        self.code_combo = QComboBox()
        self.code_combo.addItems(CODES)
        self.code_combo.setCurrentText(current_code)
        self.code_combo.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        self.code_combo.currentTextChanged.connect(self.on_code_changed)
        self.form_layout.addRow("Код:", self.code_combo)

        self.hours_edit = QLineEdit(current_hours)
        self.form_layout.addRow("Часы:", self.hours_edit)

        # Для специальных кодов
        self.option_combo = QComboBox()
        self.option_combo.addItems(["Двойная оплата", "Отгул"])
        self.option_combo.setVisible(False)
        self.form_layout.addRow("Вариант оплаты:", self.option_combo)

        # Чекбокс для больничного по уходу за ребенком
        self.child_care_check = QCheckBox("По уходу за ребенком")
        self.child_care_check.setVisible(False)
        self.child_care_check.toggled.connect(self.on_child_care_toggled)
        self.form_layout.addRow(self.child_care_check)

        # Чекбокс для принудительного начала нового периода
        self.new_period_check = QCheckBox("Начало нового периода")
        self.new_period_check.setVisible(False)
        self.new_period_check.toggled.connect(self.on_new_period_toggled)
        self.form_layout.addRow(self.new_period_check)

        # Установка текущих значений
        if current_code == 'Б':
            if current_option == 'child_care':
                self.child_care_check.setChecked(True)
                self.new_period_check.setEnabled(False)
            elif current_option == 'new_period':
                self.new_period_check.setChecked(True)
                self.child_care_check.setEnabled(False)
        elif current_code in mapping.special_codes:
            self.option_combo.setCurrentText("Отгул" if current_option == 'day_off' else "Двойная оплата")
            self.option_combo.setVisible(True)

        self.on_code_changed(current_code)

    def on_code_changed(self, code):
        is_special = self.mapping.is_special(code)
        self.option_combo.setVisible(is_special)
        is_b = (code == 'Б')
        self.child_care_check.setVisible(is_b)
        self.new_period_check.setVisible(is_b)
        if not is_special and not is_b:
            self.option_combo.setVisible(False)
            self.child_care_check.setVisible(False)
            self.new_period_check.setVisible(False)

    def on_child_care_toggled(self, checked):
        if checked:
            self.new_period_check.setEnabled(False)
            self.new_period_check.setChecked(False)
        else:
            self.new_period_check.setEnabled(True)

    def on_new_period_toggled(self, checked):
        if checked:
            self.child_care_check.setEnabled(False)
            self.child_care_check.setChecked(False)
        else:
            self.child_care_check.setEnabled(True)

    def get_values(self):
        code = self.code_combo.currentText()
        hours = self.hours_edit.text()
        if code in self.mapping.special_codes:
            option_text = self.option_combo.currentText()
            option = "day_off" if option_text == "Отгул" else "double"
        elif code == 'Б':
            if self.child_care_check.isChecked():
                option = "child_care"
            elif self.new_period_check.isChecked():
                option = "new_period"
            else:
                option = "normal"
        else:
            option = "normal"
        return code, hours, option

# ------------------------------------------------------------
# Диалог массового заполнения
class MassFillDialog(BaseDialog):
    def __init__(self, mapping: CodeMapping, parent=None):
        super().__init__("Редактирование", parent)
        self.mapping = mapping

        self.code_combo = QComboBox()
        self.code_combo.addItems(CODES)
        self.code_combo.setEditable(True)
        self.code_combo.setMaxVisibleItems(10)
        self.code_combo.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        self.code_combo.view().setMinimumWidth(120)
        self.code_combo.currentTextChanged.connect(self.on_code_changed)
        self.form_layout.addRow("Выберите код:", self.code_combo)

        self.hours_edit = QLineEdit()
        self.hours_edit.setPlaceholderText("Оставьте пустым, если часы не нужны")
        self.form_layout.addRow("Часы (опционально):", self.hours_edit)

        # Для специальных кодов
        self.option_combo = QComboBox()
        self.option_combo.addItems(["Двойная оплата", "Отгул"])
        self.option_combo.setVisible(False)
        self.form_layout.addRow("Вариант оплаты:", self.option_combo)

        # Чекбокс для больничного по уходу за ребенком
        self.child_care_check = QCheckBox("По уходу за ребенком")
        self.child_care_check.setVisible(False)
        self.child_care_check.toggled.connect(self.on_child_care_toggled)
        self.form_layout.addRow(self.child_care_check)

        # Чекбокс для принудительного начала нового периода
        self.new_period_check = QCheckBox("Начало нового периода")
        self.new_period_check.setVisible(False)
        self.new_period_check.toggled.connect(self.on_new_period_toggled)
        self.form_layout.addRow(self.new_period_check)

        self.on_code_changed(self.code_combo.currentText())

    def on_code_changed(self, code):
        is_special = self.mapping.is_special(code)
        self.option_combo.setVisible(is_special)
        is_b = (code == 'Б')
        self.child_care_check.setVisible(is_b)
        self.new_period_check.setVisible(is_b)
        if not is_special and not is_b:
            self.option_combo.setVisible(False)
            self.child_care_check.setVisible(False)
            self.new_period_check.setVisible(False)

    def on_child_care_toggled(self, checked):
        if checked:
            self.new_period_check.setEnabled(False)
            self.new_period_check.setChecked(False)
        else:
            self.new_period_check.setEnabled(True)

    def on_new_period_toggled(self, checked):
        if checked:
            self.child_care_check.setEnabled(False)
            self.child_care_check.setChecked(False)
        else:
            self.child_care_check.setEnabled(True)

    def get_values(self):
        code = self.code_combo.currentText()
        hours = self.hours_edit.text()
        if code in self.mapping.special_codes:
            option_text = self.option_combo.currentText()
            option = "day_off" if option_text == "Отгул" else "double"
        elif code == 'Б':
            if self.child_care_check.isChecked():
                option = "child_care"
            elif self.new_period_check.isChecked():
                option = "new_period"
            else:
                option = "normal"
        else:
            option = "normal"
        return code, hours, option

# ------------------------------------------------------------
# Диалог настроек проекта
class SettingsDialog(BaseDialog):
    def __init__(self, project: Project, holidays: Holidays, parent=None):
        super().__init__("Настройки проекта", parent)
        self.project = project
        self.holidays = holidays

        self.organization_edit = QLineEdit(project.organization)
        self.form_layout.addRow("Учреждение:", self.organization_edit)

        self.department_edit = QLineEdit(project.department)
        self.form_layout.addRow("Структурное подразделение:", self.department_edit)

        self.okpo_edit = QLineEdit(project.okpo)
        self.form_layout.addRow("Код по ОКПО:", self.okpo_edit)

        self.resp_dolgnost_edit = QLineEdit(project.resp_dolgnost)
        self.form_layout.addRow("Должность ответственного:", self.resp_dolgnost_edit)
        self.resp_fio_edit = QLineEdit(project.resp_fio)
        self.form_layout.addRow("ФИО ответственного:", self.resp_fio_edit)

        self.executor_dolgnost_edit = QLineEdit(project.executor_dolgnost)
        self.form_layout.addRow("Должность исполнителя:", self.executor_dolgnost_edit)
        self.executor_fio_edit = QLineEdit(project.executor_fio)
        self.form_layout.addRow("ФИО исполнителя:", self.executor_fio_edit)

    def accept(self):
        self.project.organization = self.organization_edit.text()
        self.project.department = self.department_edit.text()
        self.project.okpo = self.okpo_edit.text()
        self.project.resp_dolgnost = self.resp_dolgnost_edit.text()
        self.project.resp_fio = self.resp_fio_edit.text()
        self.project.executor_dolgnost = self.executor_dolgnost_edit.text()
        self.project.executor_fio = self.executor_fio_edit.text()
        super().accept()


# ------------------------------------------------------------
# Диалог редактирования одной доплаты
class AllowanceEditDialog(BaseDialog):
    def __init__(self, allowance: AllowanceItem = None, history: Set[str] = None, parent=None):
        super().__init__("Редактирование доплаты", parent)
        self.name_edit = QLineEdit(allowance.name if allowance else "")
        if history:
            completer = QCompleter(sorted(history))
            completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
            completer.setFilterMode(Qt.MatchFlag.MatchContains)
            self.name_edit.setCompleter(completer)
        self.form_layout.addRow("Название:", self.name_edit)

        self.amount_edit = QDoubleSpinBox()
        self.amount_edit.setRange(0, 1000000)
        self.amount_edit.setValue(allowance.amount if allowance else 0.0)
        self.form_layout.addRow("Сумма/ставка:", self.amount_edit)

        self.type_combo = QComboBox()
        self.type_combo.addItems([t.value for t in AllowanceType])
        if allowance:
            self.type_combo.setCurrentText(allowance.type.value)
        self.form_layout.addRow("Тип:", self.type_combo)

    def get_allowance(self):
        return AllowanceItem(
            self.name_edit.text(),
            self.amount_edit.value(),
            AllowanceType.from_str(self.type_combo.currentText())
        )


# ------------------------------------------------------------
# Таблица доплат (для переиспользования)
class AllowanceTableWidget(QTableWidget):
    def __init__(self, parent=None):
        super().__init__(0, 3, parent)
        self.setHorizontalHeaderLabels(["Название", "Сумма/ставка", "Тип"])
        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.allowances: List[AllowanceItem] = []
        self.history: Set[str] = set()

    def set_allowances(self, allowances: List[AllowanceItem], history: Set[str]):
        self.allowances = allowances
        self.history = history
        self.refresh()

    def refresh(self):
        self.setRowCount(len(self.allowances))
        for i, a in enumerate(self.allowances):
            self.setItem(i, 0, QTableWidgetItem(a.name))
            self.setItem(i, 1, QTableWidgetItem(f"{a.amount:.2f}"))
            self.setItem(i, 2, QTableWidgetItem(a.type.value))

    def add_allowance(self):
        dlg = AllowanceEditDialog(history=self.history)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            new_a = dlg.get_allowance()
            self.allowances.append(new_a)
            if new_a.name:
                self.history.add(new_a.name)
            self.refresh()

    def edit_allowance(self):
        row = self.currentRow()
        if row < 0:
            return
        a = self.allowances[row]
        dlg = AllowanceEditDialog(a, self.history)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            new_a = dlg.get_allowance()
            self.allowances[row] = new_a
            if new_a.name:
                self.history.add(new_a.name)
            self.refresh()

    def delete_allowance(self):
        row = self.currentRow()
        if row >= 0:
            del self.allowances[row]
            self.refresh()

    def refresh_avg_table(self):
        items = sorted(self.avg_salaries.items())  # self.avg_salaries нужно будет определить
        self.avg_table.setRowCount(len(items))
        for i, ((y, m), val) in enumerate(items):
            self.avg_table.setItem(i, 0, QTableWidgetItem(f"{y}-{m:02d}"))
            self.avg_table.setItem(i, 1, QTableWidgetItem(f"{val:.2f}"))
            coeff = self.avg_salaries_coeff.get((y, m), 0.0)
            self.avg_table.setItem(i, 2, QTableWidgetItem(f"{coeff:.2f}"))

    def add_avg_salary(self):
        dlg = AvgSalaryMonthDialog(parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            y, m, val, coeff = dlg.get_values()
            self.avg_salaries[(y, m)] = val
            self.avg_salaries_coeff[(y, m)] = coeff
            self.refresh_avg_table()

    def edit_avg_salary(self):
        row = self.avg_table.currentRow()
        if row < 0:
            return
        key_str = self.avg_table.item(row, 0).text()
        y, m = map(int, key_str.split('-'))
        val = self.avg_salaries.get((y, m), 0.0)
        coeff = self.avg_salaries_coeff.get((y, m), 0.0)
        dlg = AvgSalaryMonthDialog(y, m, val, coeff, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            new_y, new_m, new_val, new_coeff = dlg.get_values()
            if (new_y, new_m) != (y, m):
                del self.avg_salaries[(y, m)]
                del self.avg_salaries_coeff[(y, m)]
            self.avg_salaries[(new_y, new_m)] = new_val
            self.avg_salaries_coeff[(new_y, new_m)] = new_coeff
            self.refresh_avg_table()

    def delete_avg_salary(self):
        row = self.avg_table.currentRow()
        if row < 0:
            return
        key_str = self.avg_table.item(row, 0).text()
        y, m = map(int, key_str.split('-'))
        del self.avg_salaries[(y, m)]
        del self.avg_salaries_coeff[(y, m)]
        self.refresh_avg_table()

class AvgSalaryMonthDialog(BaseDialog):
    def __init__(self, year=None, month=None, value=0.0, coeff=0.0, parent=None):
        super().__init__("Средняя зарплата за месяц", parent)
        self.year_spin = QSpinBox()
        self.year_spin.setRange(2000, 2100)
        self.year_spin.setValue(year if year else datetime.date.today().year)
        self.month_spin = QSpinBox()
        self.month_spin.setRange(1, 12)
        self.month_spin.setValue(month if month else datetime.date.today().month)

        # Поле для значения без коэффициента
        self.value_edit = QDoubleSpinBox()
        self.value_edit.setRange(0, 1000000)
        self.value_edit.setValue(value)
        self.value_edit.valueChanged.connect(self.on_value_changed)

        # Поле для значения с коэффициентом
        self.coeff_edit = QDoubleSpinBox()
        self.coeff_edit.setRange(0, 1000000)
        self.coeff_edit.setValue(coeff)
        self.coeff_edit.valueChanged.connect(self.on_coeff_changed)

        self.form_layout.addRow("Год:", self.year_spin)
        self.form_layout.addRow("Месяц:", self.month_spin)
        self.form_layout.addRow("Средняя (без коэф.):", self.value_edit)
        self.form_layout.addRow("Средняя (с коэф.):", self.coeff_edit)

    def on_value_changed(self):
        self.coeff_edit.blockSignals(True)
        self.coeff_edit.setValue(self.value_edit.value() * 1.15)
        self.coeff_edit.blockSignals(False)

    def on_coeff_changed(self):
        self.value_edit.blockSignals(True)
        self.value_edit.setValue(self.coeff_edit.value() / 1.15)
        self.value_edit.blockSignals(False)

    def get_values(self):
        return self.year_spin.value(), self.month_spin.value(), self.value_edit.value(), self.coeff_edit.value()

# ------------------------------------------------------------
# Диалог редактирования сотрудника
class EmployeeEditDialog(BaseDialog):
    def __init__(self, employee: Employee = None, history: Set[str] = None, parent=None):
        super().__init__("Редактирование сотрудника", parent)
        self.resize(700, 600)

        if employee:
            self.employee = employee
            self.avg_salaries = employee.avg_salary_by_month.copy()
            self.avg_salaries_coeff = employee.avg_salary_with_coeff_by_month.copy()
        else:
            self.employee = None
            self.avg_salaries = {}
            self.avg_salaries_coeff = {}

        # Основные поля
        self.fio_edit = QLineEdit(employee.fio if employee else "")
        self.form_layout.addRow("ФИО:", self.fio_edit)

        self.tab_edit = QLineEdit(employee.tab_num if employee else "")
        self.form_layout.addRow("Табельный номер:", self.tab_edit)

        self.pos_edit = QLineEdit(employee.position if employee else "")
        self.form_layout.addRow("Должность:", self.pos_edit)

        self.salary_edit = QDoubleSpinBox()
        self.salary_edit.setRange(0, 1000000)
        self.salary_edit.setValue(employee.salary if employee else 0.0)
        self.form_layout.addRow("Оклад (руб.):", self.salary_edit)

        self.experience_spin = QSpinBox()
        self.experience_spin.setRange(0, 70)
        self.experience_spin.setValue(employee.experience_years if employee else 0)
        self.experience_spin.setSuffix(" лет")
        self.form_layout.addRow("Стаж (полных лет):", self.experience_spin)

        self.avg_sick_leave_edit = QDoubleSpinBox()
        self.avg_sick_leave_edit.setRange(0, 1000000)
        display_value = employee.avg_sick_leave * 1.15 if employee else 0.0
        self.avg_sick_leave_edit.setValue(display_value)
        self.avg_sick_leave_edit.valueChanged.connect(self.update_sick_leave_labels)
        self.form_layout.addRow("Средняя для больничных (с коэф.(Уральский)):", self.avg_sick_leave_edit)

        # Метка для отображения чистого значения
        self.sick_leave_pure_label = QLabel("Без коэф.(Уральский): 0.00")
        self.sick_leave_pure_label.setStyleSheet("font-style: italic; color: gray;")
        self.form_layout.addRow("", self.sick_leave_pure_label)  # пустая метка слева для выравнивания

        # Первоначальное обновление
        self.update_sick_leave_labels()

        # Таблица доплат
        self.allowance_table = AllowanceTableWidget()
        self.allowance_table.set_allowances(employee.allowances.copy() if employee else [], history or set())
        self.form_layout.addRow(QLabel("Доплаты:"))
        self.form_layout.addRow(self.allowance_table)

        # Кнопки управления доплатами
        btn_layout = QHBoxLayout()
        btn_add = QPushButton("Добавить")
        btn_edit = QPushButton("Изменить")
        btn_delete = QPushButton("Удалить")
        btn_layout.addWidget(btn_add)
        btn_layout.addWidget(btn_edit)
        btn_layout.addWidget(btn_delete)
        btn_layout.addStretch()
        self.form_layout.addRow(btn_layout)

        btn_add.clicked.connect(self.allowance_table.add_allowance)
        btn_edit.clicked.connect(self.allowance_table.edit_allowance)
        btn_delete.clicked.connect(self.allowance_table.delete_allowance)

        # Таблица средних зарплат по месяцам
        self.avg_table = QTableWidget(0, 3)
        self.avg_table.setHorizontalHeaderLabels(["Год-Месяц", "Без коэф.(Уральский)", "С коэф.(Уральский)"])
        self.avg_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.form_layout.addRow(QLabel("Средние зарплаты по месяцам:"))
        self.form_layout.addRow(self.avg_table)



        avg_btn_layout = QHBoxLayout()
        btn_avg_add = QPushButton("Добавить")
        btn_avg_edit = QPushButton("Изменить")
        btn_avg_delete = QPushButton("Удалить")
        avg_btn_layout.addWidget(btn_avg_add)
        avg_btn_layout.addWidget(btn_avg_edit)
        avg_btn_layout.addWidget(btn_avg_delete)
        avg_btn_layout.addStretch()
        self.form_layout.addRow(avg_btn_layout)

        btn_avg_add.clicked.connect(self.add_avg_salary)
        btn_avg_edit.clicked.connect(self.edit_avg_salary)
        btn_avg_delete.clicked.connect(self.delete_avg_salary)

        self.refresh_avg_table()

    def update_sick_leave_labels(self):
        val = self.avg_sick_leave_edit.value()
        pure = val / 1.15
        self.sick_leave_pure_label.setText(f"Без коэф.(Уральский): {pure:.2f}")

    def refresh_avg_table(self):
        items = sorted(self.avg_salaries.items())
        self.avg_table.setRowCount(len(items))
        for i, ((y, m), val) in enumerate(items):
            self.avg_table.setItem(i, 0, QTableWidgetItem(f"{y}-{m:02d}"))
            self.avg_table.setItem(i, 1, QTableWidgetItem(f"{val:.2f}"))
            coeff = self.avg_salaries_coeff.get((y, m), 0.0)
            self.avg_table.setItem(i, 2, QTableWidgetItem(f"{coeff:.2f}"))

    def add_avg_salary(self):
        dlg = AvgSalaryMonthDialog(parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            y, m, val, coeff = dlg.get_values()
            self.avg_salaries[(y, m)] = val
            self.avg_salaries_coeff[(y, m)] = coeff
            self.refresh_avg_table()

    def edit_avg_salary(self):
        row = self.avg_table.currentRow()
        if row < 0:
            return
        key_str = self.avg_table.item(row, 0).text()
        y, m = map(int, key_str.split('-'))
        val = self.avg_salaries.get((y, m), 0.0)
        coeff = self.avg_salaries_coeff.get((y, m), 0.0)
        dlg = AvgSalaryMonthDialog(y, m, val, coeff, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            new_y, new_m, new_val, new_coeff = dlg.get_values()
            if (new_y, new_m) != (y, m):
                del self.avg_salaries[(y, m)]
                del self.avg_salaries_coeff[(y, m)]
            self.avg_salaries[(new_y, new_m)] = new_val
            self.avg_salaries_coeff[(new_y, new_m)] = new_coeff
            self.refresh_avg_table()

    def delete_avg_salary(self):
        row = self.avg_table.currentRow()
        if row < 0:
            return
        key_str = self.avg_table.item(row, 0).text()
        y, m = map(int, key_str.split('-'))
        del self.avg_salaries[(y, m)]
        del self.avg_salaries_coeff[(y, m)]
        self.refresh_avg_table()

    def get_employee(self):
        allowances = self.allowance_table.allowances
        history = self.allowance_table.history
        avg_sick_leave_value = self.avg_sick_leave_edit.value() / 1.15  # делим на 1.15
        emp = Employee(
            self.fio_edit.text(),
            self.tab_edit.text(),
            self.pos_edit.text(),
            self.salary_edit.value(),
            allowances,
            self.avg_salaries,
            self.avg_salaries_coeff,
            avg_sick_leave_value,  # передаём уже поделённое значение
            self.experience_spin.value()
        )
        return emp, history


# ------------------------------------------------------------
# Диалог списка сотрудников
class EmployeeDialog(QDialog):
    def __init__(self, employees: List[Employee], history: Set[str], database: EmployeeDatabase, project: Project, parent=None):
        super().__init__(parent)
        self.employees = employees
        self.history = history
        self.database = database
        self.project = project
        #self.setWindowTitle("Сотрудники")
        self.setModal(True)
        self.resize(900, 500)

        layout = QVBoxLayout(self)

        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(["ФИО", "Таб. номер", "Должность", "Оклад (руб.)", "Средняя (без коэф.) (руб.)"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.table)

        btn_layout = QHBoxLayout()
        self.btn_add = QPushButton("Добавить")
        self.btn_edit = QPushButton("Изменить")
        self.btn_delete = QPushButton("Удалить")
        self.btn_import = QPushButton("Импорт из Excel")
        self.btn_export = QPushButton("Экспорт в Excel")
        self.btn_save_to_db = QPushButton("Сохранить в базу")
        self.btn_load_from_db = QPushButton("Загрузить из базы (с обновлением)")
        self.btn_close = QPushButton("Закрыть")
        btn_layout.addWidget(self.btn_add)
        btn_layout.addWidget(self.btn_edit)
        btn_layout.addWidget(self.btn_delete)
        btn_layout.addWidget(self.btn_import)
        btn_layout.addWidget(self.btn_export)
        btn_layout.addWidget(self.btn_save_to_db)
        btn_layout.addWidget(self.btn_load_from_db)
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_close)
        layout.addLayout(btn_layout)

        self.btn_add.clicked.connect(self.add_employee)
        self.btn_edit.clicked.connect(self.edit_employee)
        self.btn_delete.clicked.connect(self.delete_employee)
        self.btn_import.clicked.connect(self.import_employees)
        self.btn_export.clicked.connect(self.export_employees)
        self.btn_save_to_db.clicked.connect(self.save_to_database)
        self.btn_load_from_db.clicked.connect(self.load_from_database)
        self.btn_close.clicked.connect(self.accept)

        self.refresh_table()

    def refresh_table(self):
        self.table.setRowCount(len(self.employees))
        current = self.project.current_month
        for i, emp in enumerate(self.employees):
            self.table.setItem(i, 0, QTableWidgetItem(emp.fio))
            self.table.setItem(i, 1, QTableWidgetItem(emp.tab_num))
            self.table.setItem(i, 2, QTableWidgetItem(emp.position))
            self.table.setItem(i, 3, QTableWidgetItem(f"{emp.salary:.2f}"))
            avg, source = emp.get_avg_salary_for_month(current[0], current[1])
            if source:
                text = f"{avg:.2f} (за {source[1]:02d}.{source[0]})"
            else:
                text = f"{avg:.2f} (нет данных)"
            self.table.setItem(i, 4, QTableWidgetItem(text))

    def add_employee(self):
        dlg = EmployeeEditDialog(history=self.history)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            emp, new_history = dlg.get_employee()
            self.employees.append(emp)
            self.history.update(new_history)
            self.refresh_table()

    def edit_employee(self):
        row = self.table.currentRow()
        if row < 0:
            return
        emp = self.employees[row]
        dlg = EmployeeEditDialog(emp, self.history)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            new_emp, new_history = dlg.get_employee()
            self.employees[row] = new_emp
            self.history.update(new_history)
            self.refresh_table()

    def delete_employee(self):
        row = self.table.currentRow()
        if row >= 0:
            del self.employees[row]
            self.refresh_table()

    def import_employees(self):
        if not PANDAS_AVAILABLE:
            QMessageBox.warning(self, "Ошибка", "Библиотека pandas не установлена.")
            return
        fname, _ = QFileDialog.getOpenFileName(self, "Выберите файл Excel", "", "Excel Files (*.xlsx)")
        if fname:
            try:
                df = pd.read_excel(fname, header=None)
                added = 0
                for idx, row in df.iterrows():
                    try:
                        fio = str(row[0]) if pd.notna(row[0]) else ""
                        if not fio:
                            continue
                        tab_num = str(row[1]) if pd.notna(row[1]) else ""
                        position = str(row[2]) if pd.notna(row[2]) else ""
                        salary = float(row[3]) if pd.notna(row[3]) else 0.0
                        avg_salary = float(row[4]) if len(row) > 4 and pd.notna(row[4]) else 0.0
                    except ValueError as e:
                        QMessageBox.warning(self, "Ошибка",
                                            f"Строка {idx + 1} содержит некорректные данные: {e}. Строка пропущена.")
                        continue
                    current = self.project.current_month
                    avg_dict = {current: avg_salary}
                    emp = Employee(fio, tab_num, position, salary, [], avg_dict, {})
                    self.employees.append(emp)
                    added += 1
                if added > 0:
                    QMessageBox.information(self, "Импорт", f"Добавлено {added} сотрудников.")
                    self.refresh_table()
                else:
                    QMessageBox.warning(self, "Импорт",
                                        "Не удалось добавить ни одного сотрудника. Проверьте формат файла.")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка импорта", str(e))

    def export_employees(self):
        fname, _ = QFileDialog.getSaveFileName(self, "Сохранить список сотрудников", "", "Excel Files (*.xlsx)")
        if fname and PANDAS_AVAILABLE:
            try:
                data = []
                current = self.project.current_month
                for emp in self.employees:
                    avg = emp.avg_salary_by_month.get(current, 0.0)
                    data.append([emp.fio, emp.tab_num, emp.position, emp.salary, avg])
                df = pd.DataFrame(data)
                df.to_excel(fname, index=False, header=False)
                QMessageBox.information(self, "Экспорт", "Список сотрудников сохранён.")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка экспорта", str(e))

    def save_to_database(self):
        for emp in self.employees:
            self.database.add_employee(emp)
        self.database.save()
        QMessageBox.information(self, "База сотрудников", f"{len(self.employees)} сотрудников сохранено в базу.")

    def load_from_database(self):
        if not self.database.employees:
            QMessageBox.information(self, "База сотрудников", "База сотрудников пуста.")
            return
        updated = 0
        added = 0
        existing_by_tab = {e.tab_num: e for e in self.employees if e.tab_num}
        for db_emp in self.database.employees:
            if db_emp.tab_num and db_emp.tab_num in existing_by_tab:
                idx = self.employees.index(existing_by_tab[db_emp.tab_num])
                self.employees[idx] = db_emp
                updated += 1
            else:
                self.employees.append(db_emp)
                added += 1
        self.refresh_table()
        QMessageBox.information(self, "База сотрудников", f"Обновлено {updated} сотрудников, добавлено {added}.")


# ------------------------------------------------------------
# Диалог редактирования записи отпуска
class VacationEditDialog(BaseDialog):
    def __init__(self, employees: List[Employee], holidays: Holidays, record: VacationRecord = None, parent=None):
        super().__init__("Редактирование отпуска", parent)
        self.employees = employees
        self.holidays = holidays
        self.record = record

        # Сотрудник
        self.emp_combo = QComboBox()
        self.emp_combo.setEditable(True)
        self.emp_combo.setMaxVisibleItems(15)
        employee_list = [f"{e.fio} ({e.tab_num})" if e.tab_num else e.fio for e in employees]
        self.emp_combo.addItems(employee_list)
        if record:
            for i, e in enumerate(employees):
                if (e.tab_num and e.tab_num == record.employee_tab_num) or e.fio == record.employee_fio:
                    self.emp_combo.setCurrentIndex(i)
                    break
        self.form_layout.addRow("Сотрудник:", self.emp_combo)

        # Дата начала
        self.start_edit = QDateEdit()
        self.start_edit.setCalendarPopup(True)
        self.start_edit.setDisplayFormat("dd.MM.yyyy")
        if record and record.start_date:
            self.start_edit.setDate(QDate(record.start_date.year, record.start_date.month, record.start_date.day))
        else:
            self.start_edit.setDate(QDate.currentDate())
        self.start_edit.dateChanged.connect(self.on_start_or_days_changed)
        self.form_layout.addRow("Дата начала:", self.start_edit)

        # Количество дней
        self.days_spin = QSpinBox()
        self.days_spin.setRange(1, 365)
        if record and record.start_date and record.end_date:
            days = (record.end_date - record.start_date).days + 1
            self.days_spin.setValue(days)
        else:
            self.days_spin.setValue(1)
        self.days_spin.valueChanged.connect(self.on_start_or_days_changed)
        self.form_layout.addRow("Количество дней:", self.days_spin)

        # Дата окончания
        self.end_edit = QDateEdit()
        self.end_edit.setCalendarPopup(True)
        self.end_edit.setDisplayFormat("dd.MM.yyyy")
        if record and record.end_date:
            self.end_edit.setDate(QDate(record.end_date.year, record.end_date.month, record.end_date.day))
        else:
            self.end_edit.setDate(QDate.currentDate())
        self.end_edit.dateChanged.connect(self.on_end_changed)
        self.form_layout.addRow("Дата окончания:", self.end_edit)

        # Код отпуска
        self.code_combo = QComboBox()
        self.code_combo.addItems(["О", "ОД", "ОУ", "ОР"])
        if record and record.code:
            self.code_combo.setCurrentText(record.code)
        self.form_layout.addRow("Код:", self.code_combo)

        # Описание
        self.desc_edit = QLineEdit(record.description if record else "")
        self.form_layout.addRow("Описание:", self.desc_edit)

        # Инициализация
        self.on_start_or_days_changed()

    def on_start_or_days_changed(self):
        start = self.start_edit.date().toPython()
        days = self.days_spin.value()
        # Вычисляем дату окончания, пропуская праздники
        current = start
        count = 0
        while count < days:
            # Если день не праздничный, засчитываем его
            if not self.holidays.is_holiday(current):
                count += 1
            # Переходим к следующему дню
            current += datetime.timedelta(days=1)
        # current теперь указывает на день после последнего дня отпуска, поэтому вычитаем 1
        end = current - datetime.timedelta(days=1)
        self.end_edit.blockSignals(True)
        self.end_edit.setDate(QDate(end.year, end.month, end.day))
        self.end_edit.blockSignals(False)

    def on_end_changed(self):
        start = self.start_edit.date().toPython()
        end = self.end_edit.date().toPython()
        if end < start:
            return
        # Подсчитываем количество не-праздничных дней между start и end включительно
        days = 0
        current = start
        while current <= end:
            if not self.holidays.is_holiday(current):
                days += 1
            current += datetime.timedelta(days=1)
        self.days_spin.blockSignals(True)
        self.days_spin.setValue(days)
        self.days_spin.blockSignals(False)

    def get_record(self) -> VacationRecord:
        idx = self.emp_combo.currentIndex()
        if idx >= 0 and idx < len(self.employees):
            emp = self.employees[idx]
            fio = emp.fio
            tab_num = emp.tab_num
        else:
            text = self.emp_combo.currentText()
            match = re.search(r'\(([^)]+)\)', text)
            if match:
                tab_num = match.group(1)
                fio = text[:text.find('(')].strip()
            else:
                fio = text
                tab_num = ""
        start = self.start_edit.date().toPython()
        end = self.end_edit.date().toPython()
        code = self.code_combo.currentText()
        desc = self.desc_edit.text()
        return VacationRecord(fio, tab_num, start, end, code, desc)


# ------------------------------------------------------------
# Диалог графика отпусков
class VacationDialog(QDialog):
    def __init__(self, project: Project, holidays: Holidays, parent=None):
        super().__init__(parent)
        self.project = project
        self.holidays = holidays
        self.setWindowTitle("График отпусков")  # добавлен заголовок
        self.setModal(True)
        self.resize(900, 500)

        layout = QVBoxLayout(self)

        # Панель выбора года и поиска
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Год:"))
        self.year_combo = QComboBox()
        self.year_combo.currentIndexChanged.connect(self.refresh_table)
        filter_layout.addWidget(self.year_combo)

        filter_layout.addWidget(QLabel("Поиск по ФИО:"))
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Введите фамилию...")
        self.search_edit.textChanged.connect(self.refresh_table)
        filter_layout.addWidget(self.search_edit)

        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        self.table = QTableWidget(0, 6)
        self.table.setHorizontalHeaderLabels(["Сотрудник", "Таб. номер", "Дата начала", "Дата окончания", "Код", "Описание"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.table)

        btn_layout = QHBoxLayout()
        self.btn_add = QPushButton("Добавить")
        self.btn_edit = QPushButton("Изменить")
        self.btn_delete = QPushButton("Удалить")
        self.btn_import = QPushButton("Импорт из Excel")
        self.btn_export = QPushButton("Экспорт в Excel")
        self.btn_save_file = QPushButton("Сохранить в JSON")
        self.btn_load_file = QPushButton("Загрузить из JSON")
        self.btn_close = QPushButton("Закрыть")
        btn_layout.addWidget(self.btn_add)
        btn_layout.addWidget(self.btn_edit)
        btn_layout.addWidget(self.btn_delete)
        btn_layout.addWidget(self.btn_import)
        btn_layout.addWidget(self.btn_export)
        btn_layout.addWidget(self.btn_save_file)
        btn_layout.addWidget(self.btn_load_file)
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_close)
        layout.addLayout(btn_layout)

        self.btn_add.clicked.connect(self.add_vacation)
        self.btn_edit.clicked.connect(self.edit_vacation)
        self.btn_delete.clicked.connect(self.delete_vacation)
        self.btn_import.clicked.connect(self.import_vacations)
        self.btn_export.clicked.connect(self.export_vacations)
        self.btn_save_file.clicked.connect(self.save_to_file)
        self.btn_load_file.clicked.connect(self.load_from_file)
        self.btn_close.clicked.connect(self.accept)

        # Список отфильтрованных отпусков для текущего года и поиска
        self.filtered_vacations = []

        self.update_year_combo()
        self.refresh_table()

    def update_year_combo(self):
        """Обновляет выпадающий список годов на основе всех отпусков в проекте."""
        current_data = self.year_combo.currentData()
        self.year_combo.blockSignals(True)
        self.year_combo.clear()
        self.year_combo.addItem("Все года", None)
        years = set()
        for v in self.project.vacations:
            if v.start_date:
                years.add(v.start_date.year)
            if v.end_date:
                years.add(v.end_date.year)
        for y in sorted(years):
            self.year_combo.addItem(str(y), y)
        if current_data is not None and current_data in years:
            index = self.year_combo.findData(current_data)
            if index >= 0:
                self.year_combo.setCurrentIndex(index)
        else:
            self.year_combo.setCurrentIndex(0)
        self.year_combo.blockSignals(False)

    def refresh_table(self):
        """Обновляет таблицу, отображая отпуска в соответствии с выбранным годом и поиском по ФИО."""
        selected_year = self.year_combo.currentData()
        search_text = self.search_edit.text().strip().lower()

        self.filtered_vacations.clear()
        for vac in self.project.vacations:
            # Фильтр по году
            if selected_year is not None:
                # Проверяем, попадает ли отпуск в выбранный год
                if not (vac.start_date and vac.start_date.year == selected_year or
                        vac.end_date and vac.end_date.year == selected_year or
                        (vac.start_date and vac.end_date and
                         vac.start_date.year <= selected_year <= vac.end_date.year)):
                    continue
            # Фильтр по ФИО (поиск по подстроке)
            if search_text and search_text not in vac.employee_fio.lower():
                continue
            self.filtered_vacations.append(vac)

        self.table.setRowCount(len(self.filtered_vacations))
        for i, vac in enumerate(self.filtered_vacations):
            self.table.setItem(i, 0, QTableWidgetItem(vac.employee_fio))
            self.table.setItem(i, 1, QTableWidgetItem(vac.employee_tab_num))
            if vac.start_date:
                self.table.setItem(i, 2, QTableWidgetItem(vac.start_date.strftime("%d.%m.%Y")))
            else:
                self.table.setItem(i, 2, QTableWidgetItem(""))
            if vac.end_date:
                self.table.setItem(i, 3, QTableWidgetItem(vac.end_date.strftime("%d.%m.%Y")))
            else:
                self.table.setItem(i, 3, QTableWidgetItem(""))
            self.table.setItem(i, 4, QTableWidgetItem(vac.code))
            self.table.setItem(i, 5, QTableWidgetItem(vac.description))

    # Остальные методы (add_vacation, edit_vacation, delete_vacation, импорт/экспорт) остаются без изменений
    # ...

    def add_vacation(self):
        dlg = VacationEditDialog(self.project.employees, self.holidays)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            rec = dlg.get_record()
            self.project.vacations.append(rec)
            self.update_year_combo()
            self.refresh_table()

    def edit_vacation(self):
        row = self.table.currentRow()
        if row < 0:
            return
        vac = self.filtered_vacations[row]
        dlg = VacationEditDialog(self.project.employees, self.holidays, vac)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            new_rec = dlg.get_record()
            for i, v in enumerate(self.project.vacations):
                if v is vac:
                    self.project.vacations[i] = new_rec
                    break
            self.update_year_combo()
            self.refresh_table()

    def delete_vacation(self):
        row = self.table.currentRow()
        if row < 0:
            return
        vac = self.filtered_vacations[row]
        self.project.vacations.remove(vac)
        self.update_year_combo()
        self.refresh_table()

    def import_vacations(self):
        if not PANDAS_AVAILABLE:
            QMessageBox.warning(self, "Ошибка", "Библиотека pandas не установлена.")
            return
        fname, _ = QFileDialog.getOpenFileName(self, "Выберите файл Excel", "", "Excel Files (*.xlsx)")
        if fname:
            try:
                df = pd.read_excel(fname, header=None)
                added = 0
                for idx, row in df.iterrows():
                    try:
                        fio = str(row[0]) if pd.notna(row[0]) else ""
                        tab_num = str(row[1]) if pd.notna(row[1]) else ""
                        start_str = str(row[2]) if pd.notna(row[2]) else ""
                        end_str = str(row[3]) if pd.notna(row[3]) else ""
                        code = str(row[4]) if pd.notna(row[4]) else "О"
                        desc = str(row[5]) if len(row) > 5 and pd.notna(row[5]) else ""
                        start_date = None
                        end_date = None
                        if start_str:
                            try:
                                start_date = datetime.datetime.strptime(start_str, "%d.%m.%Y").date()
                            except:
                                start_date = datetime.date.fromisoformat(start_str)
                        if end_str:
                            try:
                                end_date = datetime.datetime.strptime(end_str, "%d.%m.%Y").date()
                            except:
                                end_date = datetime.date.fromisoformat(end_str)
                    except Exception as e:
                        QMessageBox.warning(self, "Ошибка", f"Строка {idx+1} содержит некорректные данные: {e}. Строка пропущена.")
                        continue
                    rec = VacationRecord(fio, tab_num, start_date, end_date, code, desc)
                    self.project.vacations.append(rec)
                    added += 1
                self.update_year_combo()
                self.refresh_table()
                QMessageBox.information(self, "Импорт", f"Добавлено {added} записей.")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка импорта", str(e))

    def export_vacations(self):
        fname, _ = QFileDialog.getSaveFileName(self, "Сохранить график отпусков", "", "Excel Files (*.xlsx)")
        if fname and PANDAS_AVAILABLE:
            try:
                data = []
                for vac in self.project.vacations:
                    row = [
                        vac.employee_fio,
                        vac.employee_tab_num,
                        vac.start_date.strftime("%d.%m.%Y") if vac.start_date else "",
                        vac.end_date.strftime("%d.%m.%Y") if vac.end_date else "",
                        vac.code,
                        vac.description
                    ]
                    data.append(row)
                df = pd.DataFrame(data)
                df.to_excel(fname, index=False, header=False)
                QMessageBox.information(self, "Экспорт", "График отпусков сохранён.")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка экспорта", str(e))

    def save_to_file(self):
        fname, _ = QFileDialog.getSaveFileName(self, "Сохранить график отпусков", "", "JSON Files (*.json)")
        if fname:
            if not fname.endswith(".json"):
                fname += ".json"
            try:
                data = [v.to_dict() for v in self.project.vacations]
                with open(fname, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                QMessageBox.information(self, "Сохранение", "График отпусков сохранён.")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))

    def load_from_file(self):
        fname, _ = QFileDialog.getOpenFileName(self, "Загрузить график отпусков", "", "JSON Files (*.json)")
        if fname:
            try:
                with open(fname, "r", encoding="utf-8") as f:
                    data = json.load(f)
                msg_box = QMessageBox(self)
                msg_box.setWindowTitle("Загрузка")
                msg_box.setText("Заменить текущий график или добавить записи?")
                btn_replace = msg_box.addButton("Заменить", QMessageBox.ButtonRole.YesRole)
                btn_add = msg_box.addButton("Добавить", QMessageBox.ButtonRole.NoRole)
                btn_cancel = msg_box.addButton("Отмена", QMessageBox.ButtonRole.RejectRole)
                msg_box.exec()
                clicked = msg_box.clickedButton()
                if clicked == btn_cancel:
                    return
                new_vacations = [VacationRecord.from_dict(d) for d in data]
                if clicked == btn_replace:
                    self.project.vacations = new_vacations
                else:
                    self.project.vacations.extend(new_vacations)
                self.update_year_combo()
                self.refresh_table()
                QMessageBox.information(self, "Загрузка", f"Загружено {len(new_vacations)} записей.")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))


# ------------------------------------------------------------
# Объединённый диалог настроек кодов
class CodeSettingsDialog(QDialog):
    def __init__(self, mapping: CodeMapping, parent=None):
        super().__init__(parent)
        self.mapping = mapping
        #self.setWindowTitle("Настройка кодов табеля")
        self.setModal(True)
        self.resize(500, 400)

        layout = QVBoxLayout(self)
        tabs = QTabWidget()
        layout.addWidget(tabs)

        main_tab = QWidget()
        main_layout = QFormLayout(main_tab)
        self.vacation_edit = QLineEdit(", ".join(sorted(mapping.vacation)))
        self.business_edit = QLineEdit(", ".join(sorted(mapping.business_trip)))
        self.sick_edit = QLineEdit(", ".join(sorted(mapping.sick_leave)))
        self.unpaid_edit = QLineEdit(", ".join(sorted(mapping.unpaid)))
        main_layout.addRow("Отпуск (коды):", self.vacation_edit)
        main_layout.addRow("Командировка:", self.business_edit)
        main_layout.addRow("Больничный:", self.sick_edit)
        main_layout.addRow("Неявки без содержания:", self.unpaid_edit)
        tabs.addTab(main_tab, "Основные")

        special_tab = QWidget()
        special_layout = QVBoxLayout(special_tab)
        self.code_widgets = {}
        for code, info in mapping.special_codes.items():
            group_box = QGroupBox(f"Код {code} – {info['name']}")
            group_layout = QVBoxLayout()
            rb_double = QRadioButton("Двойная оплата")
            rb_day_off = QRadioButton("Отгул")
            if mapping.special_payment_option.get(code) == 'day_off':
                rb_day_off.setChecked(True)
            else:
                rb_double.setChecked(True)
            group_layout.addWidget(rb_double)
            group_layout.addWidget(rb_day_off)
            group_box.setLayout(group_layout)
            special_layout.addWidget(group_box)
            self.code_widgets[code] = (rb_double, rb_day_off)
        special_layout.addStretch()
        tabs.addTab(special_tab, "Специальные коды")

        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

    def accept(self):
        def parse(s):
            return {x.strip() for x in s.split(",") if x.strip()}
        self.mapping.vacation = parse(self.vacation_edit.text())
        self.mapping.business_trip = parse(self.business_edit.text())
        self.mapping.sick_leave = parse(self.sick_edit.text())
        self.mapping.unpaid = parse(self.unpaid_edit.text())
        for code, (rb_double, rb_day_off) in self.code_widgets.items():
            if rb_double.isChecked():
                self.mapping.special_payment_option[code] = 'double'
            else:
                self.mapping.special_payment_option[code] = 'day_off'
        super().accept()


# ------------------------------------------------------------
# Класс для расчёта зарплаты (вынесен из диалога)
class PayrollCalculator:
    def __init__(self, project: Project, holidays: Holidays):
        self.project = project
        self.holidays = holidays

    def calculate_for_month(self, year: int, month: int) -> Tuple[List[Tuple], List[Tuple]]:
        data = self.project.months_data.get((year, month), {})
        norm = self.project.get_norm(year, month)
        if norm <= 0:
            norm = 1
        days_in_month = calendar.monthrange(year, month)[1]
        mapping = self.project.code_mapping

        summary_rows = []
        detail_rows = []

        for emp_idx, emp in enumerate(self.project.employees):
            vac_days = 0
            bus_days = 0
            fact_days = 0
            special_payments = {}
            sick_periods = []
            current_period = 0
            in_sick = False

            for day in range(1, days_in_month + 1):
                code, hours, option = data.get((emp_idx, day), ("", "", "double"))
                if not code:
                    continue

                date = datetime.date(year, month, day)
                is_holiday = self.holidays.is_holiday(date)

                cat = mapping.get_category(code)

                if cat == 'vacation' and code != 'ОР':
                    if not is_holiday:
                        vac_days += 1
                elif cat == 'business':
                    if not is_holiday:
                        bus_days += 1
                elif cat == 'sick':
                    if option == 'child_care':
                        # больничный по уходу – не оплачивается, разрывает период
                        if in_sick:
                            sick_periods.append(current_period)
                            in_sick = False
                    elif option == 'new_period':
                        # принудительное начало нового периода
                        if in_sick:
                            sick_periods.append(current_period)
                            in_sick = False
                        # начинаем новый период
                        if not in_sick:
                            in_sick = True
                            current_period = 1
                    else:
                        # обычный больничный, продолжает текущий период или начинает новый
                        if not in_sick:
                            in_sick = True
                            current_period = 1
                        else:
                            current_period += 1
                else:
                    if code == 'Ф':
                        fact_days += 1
                    elif mapping.is_special(code):
                        if option == 'double':
                            special_payments[code] = special_payments.get(code, 0) + 1
                        else:
                            fact_days += 1

            if in_sick:
                sick_periods.append(current_period)

            sick_days_paid = sum(min(p, 3) for p in sick_periods)

            salary = float(emp.salary) if emp.salary else 0.0
            avg_salary, source = emp.get_avg_salary_for_month(year, month)

            salary_part = salary / norm * fact_days
            allowance_total = emp.total_allowance(norm, fact_days)

            otp_pay = avg_salary * vac_days
            kom_pay = avg_salary * bus_days

            # Расчёт больничных (только за счёт работодателя – первые 3 дня) с учётом стажа
            sick_coeff = emp.get_sick_pay_coefficient()
            sick_avg = emp.avg_sick_leave if emp.avg_sick_leave != 0 else avg_salary
            bol_pay = sick_avg * sick_coeff * sick_days_paid

            special_total = 0.0
            for code, count in special_payments.items():
                info = mapping.get_special_info(code)
                if info:
                    multiplier = info.get('payment_multiplier', 1.0)
                    special_total += avg_salary * multiplier * count

            total = salary_part + allowance_total + otp_pay + kom_pay + bol_pay + special_total

            summary_rows.append((
                emp.fio, fact_days, salary_part, allowance_total,
                otp_pay, kom_pay, bol_pay, total
            ))

            # Обычные доплаты
            for name, rate, typ, accrued in emp.allowance_details(norm, fact_days):
                detail_rows.append((emp.fio, name, rate, typ, accrued))

            # Специальные коды
            for code, count in special_payments.items():
                info = mapping.get_special_info(code)
                if info:
                    multiplier = info.get('payment_multiplier', 1.0)
                    accrued = avg_salary * multiplier * count
                    detail_rows.append((emp.fio, f"Код {code}: {info['name']}", multiplier, "Спец.", accrued))

            if vac_days > 0:
                source_text = f"{source[0]}-{source[1]:02d}" if source else "нет данных"
                detail_rows.append((emp.fio, f"Отпускные (средняя за {source_text})", avg_salary, "Средняя", otp_pay))

            if bus_days > 0:
                source_text = f"{source[0]}-{source[1]:02d}" if source else "нет данных"
                detail_rows.append(
                    (emp.fio, f"Командировочные (средняя за {source_text})", avg_salary, "Средняя", kom_pay))

            if sick_days_paid > 0:
                source_text = f"{source[0]}-{source[1]:02d}" if source else "нет данных"
                sick_source = f"ср.больн. {emp.avg_sick_leave:.0f}" if emp.avg_sick_leave else f"ср.зарплата за {source_text}"
                detail_rows.append((emp.fio,
                                    f"Больничные (за счёт работодателя, {sick_days_paid} дн., стаж {emp.experience_years} лет, {sick_source})",
                                    sick_avg, "Средняя", bol_pay))

        return summary_rows, detail_rows

    def calculate_for_month_with_data(self, year: int, month: int, data: Dict, norm: int) -> Tuple[
        List[Tuple], List[Tuple]]:
        # (код аналогичен calculate_for_month, но использует переданные data и norm)
        if norm <= 0:
            norm = 1
        days_in_month = calendar.monthrange(year, month)[1]
        mapping = self.project.code_mapping

        summary_rows = []
        detail_rows = []

        for emp_idx, emp in enumerate(self.project.employees):
            vac_days = 0
            bus_days = 0
            fact_days = 0
            special_payments = {}
            sick_periods = []
            current_period = 0
            in_sick = False

            for day in range(1, days_in_month + 1):
                code, hours, option = data.get((emp_idx, day), ("", "", "double"))
                if not code:
                    continue

                date = datetime.date(year, month, day)
                is_holiday = self.holidays.is_holiday(date)
                cat = mapping.get_category(code)

                if cat == 'vacation' and code != 'ОР':
                    if not is_holiday:
                        vac_days += 1
                elif cat == 'business':
                    if not is_holiday:
                        bus_days += 1
                elif cat == 'sick':
                    if option == 'child_care':
                        if in_sick:
                            sick_periods.append(current_period)
                            in_sick = False
                    elif option == 'new_period':
                        if in_sick:
                            sick_periods.append(current_period)
                            in_sick = False
                        if not in_sick:
                            in_sick = True
                            current_period = 1
                    else:
                        if not in_sick:
                            in_sick = True
                            current_period = 1
                        else:
                            current_period += 1
                else:
                    if code == 'Ф':
                        fact_days += 1
                    elif mapping.is_special(code):
                        if option == 'double':
                            special_payments[code] = special_payments.get(code, 0) + 1
                        else:
                            fact_days += 1

            if in_sick:
                sick_periods.append(current_period)
            sick_days_paid = sum(min(p, 3) for p in sick_periods)

            salary = float(emp.salary) if emp.salary else 0.0
            avg_salary, source = emp.get_avg_salary_for_month(year, month)

            salary_part = salary / norm * fact_days
            allowance_total = emp.total_allowance(norm, fact_days)

            otp_pay = avg_salary * vac_days
            kom_pay = avg_salary * bus_days

            sick_avg = emp.avg_sick_leave if emp.avg_sick_leave != 0 else avg_salary
            sick_coeff = emp.get_sick_pay_coefficient()
            bol_pay = sick_avg * sick_coeff * sick_days_paid

            special_total = 0.0
            for code, count in special_payments.items():
                info = mapping.get_special_info(code)
                if info:
                    multiplier = info.get('payment_multiplier', 1.0)
                    special_total += avg_salary * multiplier * count

            total = salary_part + allowance_total + otp_pay + kom_pay + bol_pay + special_total

            summary_rows.append((
                emp.fio, fact_days, salary_part, allowance_total,
                otp_pay, kom_pay, bol_pay, total
            ))

            # (здесь можно добавить детализацию, если нужно)

        return summary_rows, detail_rows




class QuarterPlanningDialog(QDialog):
    def __init__(self, project: Project, holidays: Holidays, parent=None):
        super().__init__(parent)
        self.project = project
        self.holidays = holidays
        #self.setWindowTitle("Планирование на квартал")
        self.resize(1300, 700)

        # Словарь для хранения опций ячеек
        self.cell_options = {}

        layout = QVBoxLayout(self)

        # Панель выбора года и квартала
        top_layout = QHBoxLayout()
        top_layout.addWidget(QLabel("Год:"))
        self.year_combo = QComboBox()
        current_year = self.project.current_month[0]
        for y in range(current_year - 2, current_year + 3):
            self.year_combo.addItem(str(y), y)
        self.year_combo.setCurrentIndex(2)
        self.year_combo.currentIndexChanged.connect(self.on_period_changed)
        top_layout.addWidget(self.year_combo)

        top_layout.addWidget(QLabel("Квартал:"))
        self.quarter_combo = QComboBox()
        self.quarter_combo.addItems(["I (янв-март)", "II (апр-июнь)", "III (июль-сент)", "IV (окт-дек)"])
        self.quarter_combo.currentIndexChanged.connect(self.on_period_changed)
        top_layout.addWidget(self.quarter_combo)

        top_layout.addStretch()
        layout.addLayout(top_layout)

        # Основной таб-виджет с тремя разделами
        self.main_tabs = QTabWidget()
        layout.addWidget(self.main_tabs)

        # --- Раздел "Планирование" ---
        self.plan_tabs = QTabWidget()
        self.main_tabs.addTab(self.plan_tabs, "Планирование")

        # --- Раздел "Итоги по месяцам" ---
        self.result_month_tabs = QTabWidget()
        self.main_tabs.addTab(self.result_month_tabs, "Итоги по месяцам")

        # --- Раздел "Итог квартала" ---
        self.quarter_tab = QWidget()
        self.main_tabs.addTab(self.quarter_tab, "Итог квартала")
        self.setup_quarter_tab()

        # Создадим вкладки для месяцев (будут добавлены позже)
        self.month_tabs = []          # список виджетов вкладок планирования
        self.month_result_tabs = []   # список виджетов вкладок итогов по месяцам

        # Кнопки действий внизу
        btn_layout = QHBoxLayout()

        # Кнопки управления планом


        self.save_to_project_btn = QPushButton("Сохранить в проект")
        self.save_to_project_btn.clicked.connect(self.save_to_project)
        btn_layout.addWidget(self.save_to_project_btn)

        self.load_from_project_btn = QPushButton("Загрузить из проекта")
        self.load_from_project_btn.clicked.connect(self.load_from_project)
        btn_layout.addWidget(self.load_from_project_btn)

        self.clear_plan_btn = QPushButton("Очистить всё")
        self.clear_plan_btn.clicked.connect(self.clear_plan)
        btn_layout.addWidget(self.clear_plan_btn)

        btn_layout.addStretch()

        self.save_plan_btn = QPushButton("Сохранить план")
        self.save_plan_btn.clicked.connect(self.save_plan)
        btn_layout.addWidget(self.save_plan_btn)

        self.load_plan_btn = QPushButton("Загрузить план")
        self.load_plan_btn.clicked.connect(self.load_plan)
        btn_layout.addWidget(self.load_plan_btn)

        self.calc_btn = QPushButton("Рассчитать квартал")
        self.calc_btn.clicked.connect(self.calculate_quarter)
        btn_layout.addWidget(self.calc_btn)

        self.export_btn = QPushButton("Экспорт в Excel")
        self.export_btn.clicked.connect(self.export_to_excel)
        btn_layout.addWidget(self.export_btn)

        self.close_btn = QPushButton("Закрыть")
        self.close_btn.clicked.connect(self.accept)
        btn_layout.addWidget(self.close_btn)

        layout.addLayout(btn_layout)

        # Заполняем вкладки (после создания всех виджетов)
        self.create_tabs()

    # ---------- Вспомогательные методы ----------
    def get_color_for_code(self, code):
        colors = {
            'Ф': QColor(200, 255, 200),
            'В': QColor(220, 220, 220),
            'О': QColor(200, 220, 255),
            'ОД': QColor(200, 220, 255),
            'ОУ': QColor(200, 220, 255),
            'ОР': QColor(230, 200, 255),
            'К': QColor(255, 255, 200),
            'Б': QColor(255, 200, 200),
            'А': QColor(255, 220, 180),
            'НН': QColor(255, 220, 180),
            'ДО': QColor(255, 220, 180),
            'С': QColor(200, 230, 255),
            'РП': QColor(255, 200, 220),
            'КРВ': QColor(255, 255, 130),
        }
        return colors.get(code, QColor(230, 230, 230))

    def get_months_for_quarter(self):
        year = self.year_combo.currentData()
        quarter = self.quarter_combo.currentIndex()
        months = []
        for i in range(3):
            month = quarter * 3 + i + 1
            months.append((year, month))
        return months

    def on_period_changed(self):
        self.create_tabs()

    def create_tabs(self):
        # Очищаем старые вкладки
        self.plan_tabs.clear()
        self.result_month_tabs.clear()
        self.month_tabs = []
        self.month_result_tabs = []
        self.cell_options.clear()

        months = self.get_months_for_quarter()
        # Создаём вкладки планирования для каждого месяца
        for i, (y, m) in enumerate(months):
            tab = QWidget()
            self.plan_tabs.addTab(tab, f"{m:02d}.{y}")
            self.setup_month_tab(tab, y, m, i)
            self.month_tabs.append(tab)

        # Создаём вкладки итогов для каждого месяца
        for i, (y, m) in enumerate(months):
            tab = QWidget()
            self.result_month_tabs.addTab(tab, f"{m:02d}.{y}")
            self.setup_result_month_tab(tab, i)
            self.month_result_tabs.append(tab)

    # ---------- Настройка вкладок планирования ----------
    def setup_month_tab(self, tab, year, month, tab_index):
        layout = QVBoxLayout(tab)

        info_label = QLabel(f"Планирование на {month:02d}.{year} (норма: {self.project.get_norm(year, month)} дней)")
        layout.addWidget(info_label)

        btn_row = QHBoxLayout()
        fill_weekends_btn = QPushButton("Заполнить выходные")
        fill_weekends_btn.clicked.connect(lambda: self.apply_weekends_to_tab(tab))
        btn_row.addWidget(fill_weekends_btn)

        fill_vacations_btn = QPushButton("Заполнить отпуска из графика")
        fill_vacations_btn.clicked.connect(lambda: self.fill_vacations(tab))
        btn_row.addWidget(fill_vacations_btn)

        btn_row.addStretch()
        layout.addLayout(btn_row)

        table = QTableWidget()
        emp_count = len(self.project.employees)
        days = self.project.days_in_month(year, month)
        table.setRowCount(emp_count)
        table.setColumnCount(days)

        headers = [str(i+1) for i in range(days)]
        table.setHorizontalHeaderLabels(headers)

        table.setAlternatingRowColors(True)
        table.setShowGrid(True)
        table.setGridStyle(Qt.PenStyle.SolidLine)
        table.setStyleSheet("background-color: #e6e6e6; gridline-color: black;")
        table.horizontalHeader().setStyleSheet(
            "QHeaderView::section { background-color: #e6e6e6; color: black; font-weight: bold; border: 1px solid black; }"
        )
        table.verticalHeader().setStyleSheet(
            "QHeaderView::section { background-color: #e6e6e6; color: black; font-weight: bold; border: 1px solid black; }"
        )
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        for i, emp in enumerate(self.project.employees):
            item = QTableWidgetItem(emp.fio)
            item.setFlags(Qt.ItemFlag.ItemIsEnabled)
            table.setVerticalHeaderItem(i, item)

        delegate = CodeDelegate(table)
        for r in range(emp_count):
            table.setItemDelegateForRow(r, delegate)

        self.default_font = QFont()
        self.default_font.setBold(True)

        def on_item_changed(item):
            if item is None:
                return
            row = item.row()
            col = item.column()
            code = item.text()
            color = self.get_color_for_code(code)
            item.setBackground(color)
            item.setFont(self.default_font)
            item.setForeground(QBrush(QColor(0, 0, 0)))

        table.itemChanged.connect(on_item_changed)

        table.cellDoubleClicked.connect(lambda r, c: self.on_cell_double_clicked(tab, r, c))

        table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        table.customContextMenuRequested.connect(lambda pos: self.show_context_menu(tab, pos))

        layout.addWidget(table)

        tab.table = table
        tab.year = year
        tab.month = month
        tab.index = tab_index

    # ---------- Настройка вкладок итогов по месяцам ----------
    def setup_result_month_tab(self, tab, month_index):
        layout = QVBoxLayout(tab)
        table = QTableWidget(0, 7)
        table.setHorizontalHeaderLabels(["ФИО", "Окладная часть", "Доплаты", "Отпускные", "Командировки", "Больничные", "ИТОГО"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(table)
        tab.result_table = table

    # ---------- Настройка вкладки итога квартала ----------
    def setup_quarter_tab(self):
        layout = QVBoxLayout(self.quarter_tab)
        self.quarter_table = QTableWidget(0, 7)
        self.quarter_table.setHorizontalHeaderLabels(["ФИО", "Окладная часть", "Доплаты", "Отпускные", "Командировки", "Больничные", "ИТОГО"])
        self.quarter_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.quarter_table)

    # ---------- Функции заполнения ----------
    def apply_weekends_to_tab(self, tab):
        table = tab.table
        year = tab.year
        month = tab.month
        days = self.project.days_in_month(year, month)
        emp_count = len(self.project.employees)

        for emp_idx in range(emp_count):
            for day in range(1, days + 1):
                code_item = table.item(emp_idx, day - 1)
                if code_item and code_item.text():
                    continue
                try:
                    date = datetime.date(year, month, day)
                except ValueError:
                    continue
                if date.weekday() >= 5 or self.holidays.is_holiday(date):
                    table.setItem(emp_idx, day - 1, QTableWidgetItem('В'))
                    self.cell_options[(tab.index, emp_idx, day)] = "normal"

    def fill_vacations(self, tab):
        year = tab.year
        month = tab.month
        table = tab.table

        for emp_idx, emp in enumerate(self.project.employees):
            for vac in self.project.vacations:
                if vac.employee_fio != emp.fio and vac.employee_tab_num != emp.tab_num:
                    continue
                if not vac.start_date or not vac.end_date:
                    continue
                first = datetime.date(year, month, 1)
                last = datetime.date(year, month, calendar.monthrange(year, month)[1])
                start = max(vac.start_date, first)
                end = min(vac.end_date, last)
                if start <= end:
                    current = start
                    while current <= end:
                        day = current.day
                        table.setItem(emp_idx, day - 1, QTableWidgetItem(vac.code))
                        key = (tab.index, emp_idx, day)
                        self.cell_options[key] = "normal"
                        current += datetime.timedelta(days=1)

    # ---------- Редактирование ячеек ----------
    def on_cell_double_clicked(self, tab, row, col):
        emp_idx = row
        day = col + 1
        key = (tab.index, emp_idx, day)
        current_code = ""
        current_option = self.cell_options.get(key, "normal")

        code_item = tab.table.item(row, col)
        if code_item:
            current_code = code_item.text()

        dlg = CellEditDialog(self.project.code_mapping, current_code, "", current_option, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            new_code, _, new_option = dlg.get_values()
            tab.table.setItem(row, col, QTableWidgetItem(new_code))
            self.cell_options[key] = new_option

    def show_context_menu(self, tab, pos):
        selected = tab.table.selectedIndexes()
        if not selected:
            return

        menu = QMenu()
        fill_action = menu.addAction("Редактировать...")
        action = menu.exec(tab.table.viewport().mapToGlobal(pos))
        if action == fill_action:
            dlg = MassFillDialog(self.project.code_mapping, self)
            if dlg.exec() == QDialog.DialogCode.Accepted:
                new_code, _, new_option = dlg.get_values()
                if not new_code:
                    return
                for idx in selected:
                    row, col = idx.row(), idx.column()
                    tab.table.setItem(row, col, QTableWidgetItem(new_code))
                    emp_idx = row
                    day = col + 1
                    key = (tab.index, emp_idx, day)
                    self.cell_options[key] = new_option

    # ---------- Обработка клавиш (Ctrl+M) ----------
    def keyPressEvent(self, event):
        if event.modifiers() & Qt.ControlModifier and event.key() == Qt.Key_M:
            # Определяем, какая вкладка сейчас активна в разделе планирования
            current_plan_tab = self.plan_tabs.currentWidget()
            if current_plan_tab in self.month_tabs:
                self.mass_fill_current_tab(current_plan_tab)
        super().keyPressEvent(event)

    def mass_fill_current_tab(self, tab):
        selected = tab.table.selectedIndexes()
        if not selected:
            QMessageBox.information(self, "Редактирование", "Сначала выделите ячейки для заполнения.")
            return
        dlg = MassFillDialog(self.project.code_mapping, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            new_code, _, new_option = dlg.get_values()
            if not new_code:
                return
            for idx in selected:
                row, col = idx.row(), idx.column()
                tab.table.setItem(row, col, QTableWidgetItem(new_code))
                emp_idx = row
                day = col + 1
                key = (tab.index, emp_idx, day)
                self.cell_options[key] = new_option

    # ---------- Сбор данных и расчёт ----------
    def collect_month_data(self, tab):
        year = tab.year
        month = tab.month
        table = tab.table
        days = self.project.days_in_month(year, month)
        data = {}
        for emp_idx in range(len(self.project.employees)):
            for day in range(1, days + 1):
                code_item = table.item(emp_idx, day - 1)
                code = code_item.text() if code_item else ""
                option = self.cell_options.get((tab.index, emp_idx, day), "normal")
                if code:
                    data[(emp_idx, day)] = (code, "", option)
        return data

    def calculate_quarter(self):
        # Очищаем все итоговые таблицы
        for tab in self.month_result_tabs:
            tab.result_table.setRowCount(0)
        self.quarter_table.setRowCount(0)

        months = self.get_months_for_quarter()
        all_results = []  # список кортежей (month, fio, sal, allow, otp, kom, bol, total)
        calculator = PayrollCalculator(self.project, self.holidays)

        for tab in self.month_tabs:
            year = tab.year
            month = tab.month
            data = self.collect_month_data(tab)
            norm = self.project.get_norm(year, month)
            summary, _ = calculator.calculate_for_month_with_data(year, month, data, norm)
            for row in summary:
                all_results.append((month, row[0], row[2], row[3], row[4], row[5], row[6], row[7]))

        # Заполнение итогов по месяцам
        for (month, fio, sal, allow, otp, kom, bol, total) in all_results:
            month_idx = None
            for i, (y, m) in enumerate(months):
                if m == month:
                    month_idx = i
                    break
            if month_idx is None:
                continue

            table = self.month_result_tabs[month_idx].result_table
            row = table.rowCount()
            table.insertRow(row)
            table.setItem(row, 0, QTableWidgetItem(fio))
            table.setItem(row, 1, QTableWidgetItem(f"{sal:.2f}"))
            table.setItem(row, 2, QTableWidgetItem(f"{allow:.2f}"))
            table.setItem(row, 3, QTableWidgetItem(f"{otp:.2f}"))
            table.setItem(row, 4, QTableWidgetItem(f"{kom:.2f}"))
            table.setItem(row, 5, QTableWidgetItem(f"{bol:.2f}"))
            table.setItem(row, 6, QTableWidgetItem(f"{total:.2f}"))

        # Итоги по кварталу
        totals_by_fio = {}
        for (month, fio, sal, allow, otp, kom, bol, total) in all_results:
            if fio not in totals_by_fio:
                totals_by_fio[fio] = [0.0] * 6
            totals_by_fio[fio][0] += sal
            totals_by_fio[fio][1] += allow
            totals_by_fio[fio][2] += otp
            totals_by_fio[fio][3] += kom
            totals_by_fio[fio][4] += bol
            totals_by_fio[fio][5] += total

        for fio, vals in totals_by_fio.items():
            row = self.quarter_table.rowCount()
            self.quarter_table.insertRow(row)
            self.quarter_table.setItem(row, 0, QTableWidgetItem(fio))
            for i, v in enumerate(vals):
                self.quarter_table.setItem(row, i+1, QTableWidgetItem(f"{v:.2f}"))

        if totals_by_fio:
            total_vals = [0.0] * 6
            for vals in totals_by_fio.values():
                for i in range(6):
                    total_vals[i] += vals[i]
            row = self.quarter_table.rowCount()
            self.quarter_table.insertRow(row)
            self.quarter_table.setItem(row, 0, QTableWidgetItem("ИТОГО"))
            for i, v in enumerate(total_vals):
                self.quarter_table.setItem(row, i+1, QTableWidgetItem(f"{v:.2f}"))

    # ---------- Сохранение и загрузка плана (внешние файлы) ----------
    def save_plan(self):
        fname, _ = QFileDialog.getSaveFileName(self, "Сохранить план", "", "Plan Files (*.plan);;JSON Files (*.json)")
        if not fname:
            return
        if not fname.endswith('.plan') and not fname.endswith('.json'):
            fname += '.plan'

        data = {
            'year': self.year_combo.currentData(),
            'quarter': self.quarter_combo.currentIndex(),
            'months': []
        }

        for tab in self.month_tabs:
            month_data = {
                'year': tab.year,
                'month': tab.month,
                'cells': []
            }
            table = tab.table
            days = self.project.days_in_month(tab.year, tab.month)
            for emp_idx, emp in enumerate(self.project.employees):
                for day in range(1, days + 1):
                    code_item = table.item(emp_idx, day - 1)
                    if code_item and code_item.text():
                        code = code_item.text()
                        option = self.cell_options.get((tab.index, emp_idx, day), "normal")
                        month_data['cells'].append({
                            'emp_idx': emp_idx,
                            'emp_fio': emp.fio,
                            'emp_tab': emp.tab_num,
                            'day': day,
                            'code': code,
                            'option': option
                        })
            data['months'].append(month_data)

        try:
            with open(fname, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            QMessageBox.information(self, "Сохранение", "План успешно сохранён.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить план:\n{e}")

    def load_plan(self):
        fname, _ = QFileDialog.getOpenFileName(self, "Загрузить план", "", "Plan Files (*.plan *.json);;All Files (*)")
        if not fname:
            return

        try:
            with open(fname, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось прочитать файл:\n{e}")
            return

        if data['year'] != self.year_combo.currentData() or data['quarter'] != self.quarter_combo.currentIndex():
            reply = QMessageBox.question(self, "Несоответствие периода",
                                         "Год или квартал в файле не совпадают с текущими. Всё равно загрузить?",
                                         QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply != QMessageBox.StandardButton.Yes:
                return

        self.cell_options.clear()
        for tab in self.month_tabs:
            table = tab.table
            for row in range(table.rowCount()):
                for col in range(table.columnCount()):
                    table.setItem(row, col, QTableWidgetItem(""))

        emp_map = {}
        for idx, emp in enumerate(self.project.employees):
            key = (emp.fio, emp.tab_num)
            emp_map[key] = idx

        for month_data in data['months']:
            tab = None
            for t in self.month_tabs:
                if t.year == month_data['year'] and t.month == month_data['month']:
                    tab = t
                    break
            if tab is None:
                continue
            table = tab.table
            for cell in month_data['cells']:
                emp_idx = cell.get('emp_idx')
                if emp_idx is not None and 0 <= emp_idx < len(self.project.employees):
                    emp = self.project.employees[emp_idx]
                    if emp.fio != cell['emp_fio'] or emp.tab_num != cell['emp_tab']:
                        key = (cell['emp_fio'], cell['emp_tab'])
                        if key in emp_map:
                            emp_idx = emp_map[key]
                        else:
                            continue
                else:
                    key = (cell['emp_fio'], cell['emp_tab'])
                    if key in emp_map:
                        emp_idx = emp_map[key]
                    else:
                        continue
                day = cell['day']
                if 1 <= day <= self.project.days_in_month(tab.year, tab.month):
                    table.setItem(emp_idx, day - 1, QTableWidgetItem(cell['code']))
                    self.cell_options[(tab.index, emp_idx, day)] = cell['option']
        QMessageBox.information(self, "Загрузка", "План загружен.")

    # ---------- Сохранение/загрузка в проект ----------
    def save_to_project(self):
        year, quarter = self.year_combo.currentData(), self.quarter_combo.currentIndex()
        plan_data = {
            'year': year,
            'quarter': quarter,
            'months': []
        }

        for tab in self.month_tabs:
            month_data = {
                'year': tab.year,
                'month': tab.month,
                'cells': []
            }
            table = tab.table
            days = self.project.days_in_month(tab.year, tab.month)
            for emp_idx, emp in enumerate(self.project.employees):
                for day in range(1, days + 1):
                    code_item = table.item(emp_idx, day - 1)
                    if code_item and code_item.text():
                        code = code_item.text()
                        option = self.cell_options.get((tab.index, emp_idx, day), "normal")
                        month_data['cells'].append({
                            'emp_idx': emp_idx,
                            'emp_fio': emp.fio,
                            'emp_tab': emp.tab_num,
                            'day': day,
                            'code': code,
                            'option': option
                        })
            plan_data['months'].append(month_data)

        self.project.plans[(year, quarter)] = plan_data
        if self.parent() and hasattr(self.parent(), 'set_modified'):
            self.parent().set_modified()
        QMessageBox.information(self, "Сохранение", "План сохранён в проекте.")

    def load_from_project(self):
        year, quarter = self.year_combo.currentData(), self.quarter_combo.currentIndex()
        plan_data = self.project.plans.get((year, quarter))
        if not plan_data:
            QMessageBox.information(self, "Загрузка", "Для данного квартала нет сохранённого плана.")
            return

        self.cell_options.clear()
        for tab in self.month_tabs:
            table = tab.table
            for row in range(table.rowCount()):
                for col in range(table.columnCount()):
                    table.setItem(row, col, QTableWidgetItem(""))

        emp_map = {}
        for idx, emp in enumerate(self.project.employees):
            key = (emp.fio, emp.tab_num)
            emp_map[key] = idx

        for month_data in plan_data['months']:
            tab = None
            for t in self.month_tabs:
                if t.year == month_data['year'] and t.month == month_data['month']:
                    tab = t
                    break
            if tab is None:
                continue
            table = tab.table
            for cell in month_data['cells']:
                emp_idx = cell.get('emp_idx')
                if emp_idx is not None and 0 <= emp_idx < len(self.project.employees):
                    emp = self.project.employees[emp_idx]
                    if emp.fio != cell['emp_fio'] or emp.tab_num != cell['emp_tab']:
                        key = (cell['emp_fio'], cell['emp_tab'])
                        if key in emp_map:
                            emp_idx = emp_map[key]
                        else:
                            continue
                else:
                    key = (cell['emp_fio'], cell['emp_tab'])
                    if key in emp_map:
                        emp_idx = emp_map[key]
                    else:
                        continue
                day = cell['day']
                if 1 <= day <= self.project.days_in_month(tab.year, tab.month):
                    table.setItem(emp_idx, day - 1, QTableWidgetItem(cell['code']))
                    self.cell_options[(tab.index, emp_idx, day)] = cell['option']
        QMessageBox.information(self, "Загрузка", "План загружен из проекта.")

    def clear_plan(self):
        reply = QMessageBox.question(self, "Очистка",
                                     "Вы уверены, что хотите очистить все данные планирования?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply != QMessageBox.StandardButton.Yes:
            return
        self.cell_options.clear()
        for tab in self.month_tabs:
            table = tab.table
            for row in range(table.rowCount()):
                for col in range(table.columnCount()):
                    table.setItem(row, col, QTableWidgetItem(""))

    # ---------- Экспорт ----------
    def export_to_excel(self):
        fname, _ = QFileDialog.getSaveFileName(self, "Сохранить отчёт в Excel", "", "Excel Files (*.xlsx)")
        if not fname:
            return
        if not fname.endswith('.xlsx'):
            fname += '.xlsx'

        try:
            import pandas as pd
            with pd.ExcelWriter(fname, engine='openpyxl') as writer:
                # ========== ФАКТ ==========
                # ---- Динамика ФОТ (факт) ----
                # Пересчитываем данные для динамики
                months = sorted(self.project.months_data.keys())
                if months:
                    data_dynamics_fact = []
                    calculator = PayrollCalculator(self.project, self.holidays)
                    for y, m in months:
                        summary, _ = calculator.calculate_for_month(y, m)
                        if summary:
                            salary = sum(row[2] for row in summary)
                            allowance = sum(row[3] for row in summary)
                            vacation = sum(row[4] for row in summary)
                            kom = sum(row[5] for row in summary)
                            sick = sum(row[6] for row in summary)
                            total = sum(row[7] for row in summary)
                            data_dynamics_fact.append([f"{m:02d}.{y}", salary, allowance, vacation, kom, sick, total])
                        else:
                            data_dynamics_fact.append([f"{m:02d}.{y}", 0, 0, 0, 0, 0, 0])
                    df_dynamics_fact = pd.DataFrame(data_dynamics_fact,
                                                    columns=["Месяц", "Оклад (руб.)", "Доплаты (руб.)",
                                                             "Отпускные (руб.)",
                                                             "Командировки (руб.)", "Больничные (руб.)",
                                                             "ИТОГО (руб.)"])
                    df_dynamics_fact.to_excel(writer, sheet_name="Факт_Динамика", index=False)

                # ---- Структура (факт) ----
                if hasattr(self.fact_structure_tab, 'table'):
                    table = self.fact_structure_tab.table
                    data_structure_fact = []
                    for row in range(table.rowCount()):
                        row_data = []
                        for col in range(table.columnCount()):
                            item = table.item(row, col)
                            row_data.append(item.text() if item else "")
                        data_structure_fact.append(row_data)
                    headers_structure = [table.horizontalHeaderItem(i).text() for i in range(table.columnCount())]
                    df_structure_fact = pd.DataFrame(data_structure_fact, columns=headers_structure)
                    df_structure_fact.to_excel(writer, sheet_name="Факт_Структура", index=False)

                # ---- По кварталам (факт) ----
                if hasattr(self.fact_quarters_tab, 'table'):
                    table = self.fact_quarters_tab.table
                    data_quarters_fact = []
                    for row in range(table.rowCount()):
                        row_data = []
                        for col in range(table.columnCount()):
                            item = table.item(row, col)
                            row_data.append(item.text() if item else "")
                        data_quarters_fact.append(row_data)
                    headers_quarters = [table.horizontalHeaderItem(i).text() for i in range(table.columnCount())]
                    df_quarters_fact = pd.DataFrame(data_quarters_fact, columns=headers_quarters)
                    df_quarters_fact.to_excel(writer, sheet_name="Факт_Кварталы", index=False)

                # ---- Квартал помесячно (факт) ----
                if hasattr(self.fact_quarter_detail_tab, 'table'):
                    table = self.fact_quarter_detail_tab.table
                    data_quarter_detail_fact = []
                    for row in range(table.rowCount()):
                        row_data = []
                        for col in range(table.columnCount()):
                            item = table.item(row, col)
                            row_data.append(item.text() if item else "")
                        data_quarter_detail_fact.append(row_data)
                    headers_detail = [table.horizontalHeaderItem(i).text() for i in range(table.columnCount())]
                    df_detail_fact = pd.DataFrame(data_quarter_detail_fact, columns=headers_detail)
                    df_detail_fact.to_excel(writer, sheet_name="Факт_КварталПомесячно", index=False)

                # ========== ПЛАН ==========
                # ---- Динамика ФОТ (план) ----
                if months:
                    data_dynamics_plan = []
                    for y, m in months:
                        plan = self.get_plan_data_for_month(y, m)
                        data_dynamics_plan.append([f"{m:02d}.{y}",
                                                   plan['salary'],
                                                   plan['allowance'],
                                                   plan['vacation'],
                                                   plan['kom'],
                                                   plan['sick'],
                                                   plan['total']])
                    df_dynamics_plan = pd.DataFrame(data_dynamics_plan,
                                                    columns=["Месяц", "Оклад (руб.)", "Доплаты (руб.)",
                                                             "Отпускные (руб.)",
                                                             "Командировки (руб.)", "Больничные (руб.)",
                                                             "ИТОГО (руб.)"])
                    df_dynamics_plan.to_excel(writer, sheet_name="План_Динамика", index=False)

                # ---- Структура (план) ----
                if hasattr(self.plan_structure_tab, 'table'):
                    table = self.plan_structure_tab.table
                    data_structure_plan = []
                    for row in range(table.rowCount()):
                        row_data = []
                        for col in range(table.columnCount()):
                            item = table.item(row, col)
                            row_data.append(item.text() if item else "")
                        data_structure_plan.append(row_data)
                    headers_structure = [table.horizontalHeaderItem(i).text() for i in range(table.columnCount())]
                    df_structure_plan = pd.DataFrame(data_structure_plan, columns=headers_structure)
                    df_structure_plan.to_excel(writer, sheet_name="План_Структура", index=False)

                # ---- По кварталам (план) ----
                if hasattr(self.plan_quarters_tab, 'table'):
                    table = self.plan_quarters_tab.table
                    data_quarters_plan = []
                    for row in range(table.rowCount()):
                        row_data = []
                        for col in range(table.columnCount()):
                            item = table.item(row, col)
                            row_data.append(item.text() if item else "")
                        data_quarters_plan.append(row_data)
                    headers_quarters = [table.horizontalHeaderItem(i).text() for i in range(table.columnCount())]
                    df_quarters_plan = pd.DataFrame(data_quarters_plan, columns=headers_quarters)
                    df_quarters_plan.to_excel(writer, sheet_name="План_Кварталы", index=False)

                # ---- Квартал помесячно (план) ----
                if hasattr(self.plan_quarter_detail_tab, 'table'):
                    table = self.plan_quarter_detail_tab.table
                    data_quarter_detail_plan = []
                    for row in range(table.rowCount()):
                        row_data = []
                        for col in range(table.columnCount()):
                            item = table.item(row, col)
                            row_data.append(item.text() if item else "")
                        data_quarter_detail_plan.append(row_data)
                    headers_detail = [table.horizontalHeaderItem(i).text() for i in range(table.columnCount())]
                    df_detail_plan = pd.DataFrame(data_quarter_detail_plan, columns=headers_detail)
                    df_detail_plan.to_excel(writer, sheet_name="План_КварталПомесячно", index=False)

                # ========== СРАВНЕНИЯ ==========
                # ---- Сравнение по месяцам ----
                data_month = []
                for row in range(self.cmp_month_table.rowCount()):
                    row_data = []
                    for col in range(self.cmp_month_table.columnCount()):
                        item = self.cmp_month_table.item(row, col)
                        row_data.append(item.text() if item else "")
                    data_month.append(row_data)
                headers_month = [self.cmp_month_table.horizontalHeaderItem(i).text() for i in
                                 range(self.cmp_month_table.columnCount())]
                df_month = pd.DataFrame(data_month, columns=headers_month)
                df_month.to_excel(writer, sheet_name="Сравнение по месяцам", index=False)

                # ---- Сравнение по кварталам ----
                data_quarter = []
                for row in range(self.cmp_quarter_table.rowCount()):
                    row_data = []
                    for col in range(self.cmp_quarter_table.columnCount()):
                        item = self.cmp_quarter_table.item(row, col)
                        row_data.append(item.text() if item else "")
                    data_quarter.append(row_data)
                headers_quarter = [self.cmp_quarter_table.horizontalHeaderItem(i).text() for i in
                                   range(self.cmp_quarter_table.columnCount())]
                df_quarter = pd.DataFrame(data_quarter, columns=headers_quarter)
                df_quarter.to_excel(writer, sheet_name="Сравнение по кварталам", index=False)

            QMessageBox.information(self, "Экспорт", f"Отчёт сохранён в {fname}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл:\n{e}")


class SummaryDialog(QDialog):
    def __init__(self, project: Project, holidays: Holidays, parent=None):
        super().__init__(parent)
        self.project = project
        self.holidays = holidays
        self.calculator = PayrollCalculator(project, holidays)
        #self.setWindowTitle("Итоговая ведомость ФОТ")
        self.resize(1200, 600)

        layout = QVBoxLayout(self)
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)

        self.summary_table = QTableWidget()
        self.summary_table.setAlternatingRowColors(True)
        self.summary_table.setColumnCount(8)
        self.summary_table.setHorizontalHeaderLabels([
            "ФИО", "Отработано дней (дн.)", "Окладная часть (руб.)", "Доплаты (всего) (руб.)",
            "Отпускные (руб.)", "Командировочные (руб.)", "Больничные (руб.)", "Итого начислено (руб.)"
        ])
        self.summary_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tabs.addTab(self.summary_table, "Итоговая ведомость")

        self.detail_table = QTableWidget()
        self.detail_table.setColumnCount(5)
        self.detail_table.setHorizontalHeaderLabels(["ФИО", "Название доплаты", "Ставка (руб.)", "Тип", "Начислено (руб.)"])
        self.detail_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tabs.addTab(self.detail_table, "Детализация доплат")

        btn_layout = QHBoxLayout()
        self.btn_refresh = QPushButton("Пересчитать")
        self.btn_export = QPushButton("Экспорт в Excel")
        self.btn_close = QPushButton("Закрыть")
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_refresh)
        btn_layout.addWidget(self.btn_export)
        btn_layout.addWidget(self.btn_close)
        layout.addLayout(btn_layout)

        self.btn_refresh.clicked.connect(self.refresh)
        self.btn_export.clicked.connect(self.export_to_excel)
        self.btn_close.clicked.connect(self.accept)

        self.calculate()

    def refresh(self):
        self.summary_table.setRowCount(0)
        self.detail_table.setRowCount(0)
        self.calculate()

    def calculate(self):
        year, month = self.project.current_month
        summary_rows, detail_rows = self.calculator.calculate_for_month(year, month)

        self.summary_table.setRowCount(len(summary_rows))
        for i, row in enumerate(summary_rows):
            for j, val in enumerate(row):
                if j == 0:
                    item = QTableWidgetItem(str(val))
                else:
                    try:
                        num_val = float(val)
                        item = QTableWidgetItem(f"{num_val:.2f}")
                    except (ValueError, TypeError):
                        item = QTableWidgetItem(str(val))
                self.summary_table.setItem(i, j, item)

        if summary_rows:
            totals = [0.0] * 8
            for row in summary_rows:
                for j in range(1, 8):
                    try:
                        totals[j] += float(row[j])
                    except (ValueError, TypeError):
                        pass
            last_row = self.summary_table.rowCount()
            self.summary_table.insertRow(last_row)
            self.summary_table.setItem(last_row, 0, QTableWidgetItem("ИТОГО"))
            for j in range(1, 8):
                self.summary_table.setItem(last_row, j, QTableWidgetItem(f"{totals[j]:.2f}"))

        self.detail_table.setRowCount(len(detail_rows))
        for i, row in enumerate(detail_rows):
            for j, val in enumerate(row):
                if j == 0:
                    item = QTableWidgetItem(str(val))
                elif j == 3:
                    item = QTableWidgetItem(str(val))
                else:
                    try:
                        num_val = float(val)
                        item = QTableWidgetItem(f"{num_val:.2f}")
                    except (ValueError, TypeError):
                        item = QTableWidgetItem(str(val))
                self.detail_table.setItem(i, j, item)

    def export_to_excel(self):
        fname, _ = QFileDialog.getSaveFileName(self, "Сохранить ведомость", "", "Excel Files (*.xlsx)")
        if not fname:
            return
        if not PANDAS_AVAILABLE:
            QMessageBox.warning(self, "Ошибка", "Библиотека pandas не установлена. Экспорт невозможен.")
            return
        try:
            with pd.ExcelWriter(fname, engine='openpyxl') as writer:
                data_summary = []
                for row in range(self.summary_table.rowCount()):
                    row_data = []
                    for col in range(self.summary_table.columnCount()):
                        item = self.summary_table.item(row, col)
                        row_data.append(item.text() if item else "")
                    data_summary.append(row_data)
                headers_summary = [self.summary_table.horizontalHeaderItem(i).text() for i in range(self.summary_table.columnCount())]
                df_summary = pd.DataFrame(data_summary, columns=headers_summary)
                df_summary.to_excel(writer, sheet_name="Итоговая ведомость", index=False)

                data_detail = []
                for row in range(self.detail_table.rowCount()):
                    row_data = []
                    for col in range(self.detail_table.columnCount()):
                        item = self.detail_table.item(row, col)
                        row_data.append(item.text() if item else "")
                    data_detail.append(row_data)
                headers_detail = [self.detail_table.horizontalHeaderItem(i).text() for i in range(self.detail_table.columnCount())]
                df_detail = pd.DataFrame(data_detail, columns=headers_detail)
                df_detail.to_excel(writer, sheet_name="Детализация доплат", index=False)

            QMessageBox.information(self, "Экспорт", "Ведомость сохранена.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка экспорта", str(e))


# ------------------------------------------------------------
# Диалог расчёта средней зарплаты
class AvgSalaryDialog(QDialog):
    def __init__(self, project: Project, holidays: Holidays, parent=None):
        super().__init__(parent)
        self.project = project
        self.holidays = holidays
        #self.setWindowTitle("Расчёт средней зарплаты")
        self.resize(800, 400)

        layout = QVBoxLayout(self)

        # Выбор месяца
        self.month_combo = QComboBox()
        months = sorted(project.months_data.keys())
        for y, m in months:
            self.month_combo.addItem(f"{m:02d}.{y}", (y, m))
        self.month_combo.currentIndexChanged.connect(self.update_current_avg)
        layout.addWidget(QLabel("Применить к месяцу:"))
        layout.addWidget(self.month_combo)

        self.table = QTableWidget(len(project.employees), 4)
        self.table.setHorizontalHeaderLabels(["ФИО", "Текущая средняя (руб.)", "Новая средняя (руб.)", "Период"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.table)

        btn_calc = QPushButton("Рассчитать за последние 12 месяцев")
        btn_calc.clicked.connect(self.calculate)
        layout.addWidget(btn_calc)

        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

        self.update_current_avg()

    def update_current_avg(self):
        """Обновляет колонку 'Текущая средняя' в соответствии с выбранным месяцем."""
        selected = self.month_combo.currentData()
        if not selected:
            return
        year, month = selected
        for i, emp in enumerate(self.project.employees):
            avg, source = emp.get_avg_salary_for_month(year, month)
            self.table.setItem(i, 0, QTableWidgetItem(emp.fio))
            self.table.setItem(i, 1, QTableWidgetItem(f"{avg:.2f}"))
            # Оставляем пустыми колонки 2 и 3
            self.table.setItem(i, 2, QTableWidgetItem(""))
            self.table.setItem(i, 3, QTableWidgetItem(f"из {source[0]}-{source[1]:02d}" if source else "нет данных"))

    def calculate(self):
        months = sorted(self.project.months_data.keys())
        if not months:
            QMessageBox.warning(self, "Ошибка", "Нет данных по месяцам.")
            return

        calculator = PayrollCalculator(self.project, self.holidays)
        for emp_idx, emp in enumerate(self.project.employees):
            total_earnings = 0.0
            total_days = 0
            count_months = 0
            for year, month in reversed(months):
                if count_months >= 12:
                    break
                summary_rows, _ = calculator.calculate_for_month(year, month)
                if emp_idx < len(summary_rows):
                    row = summary_rows[emp_idx]
                    fact_days = row[1]
                    total = row[7]
                    if fact_days > 0:
                        total_earnings += total
                        total_days += fact_days
                        count_months += 1
            if total_days > 0:
                new_avg = total_earnings / total_days
                self.table.setItem(emp_idx, 2, QTableWidgetItem(f"{new_avg:.2f}"))
                self.table.setItem(emp_idx, 3, QTableWidgetItem(f"за {count_months} мес."))
            else:
                self.table.setItem(emp_idx, 2, QTableWidgetItem("нет данных"))
                self.table.setItem(emp_idx, 3, QTableWidgetItem(""))

    def accept(self):
        selected = self.month_combo.currentData()
        if selected:
            year, month = selected
            for i in range(self.table.rowCount()):
                item = self.table.item(i, 2)
                if item and item.text():
                    try:
                        new_avg = float(item.text())
                        self.project.employees[i].avg_salary_by_month[(year, month)] = new_avg
                    except ValueError:
                        pass
        super().accept()


class AnalyticsDialog(QDialog):
    def __init__(self, project: Project, holidays: Holidays, parent=None):
        super().__init__(parent)
        self.project = project
        self.holidays = holidays
        #self.setWindowTitle("Аналитика ФОТ")
        self.resize(1100, 700)

        layout = QVBoxLayout(self)
        self.main_tabs = QTabWidget()
        layout.addWidget(self.main_tabs)

        # ---------- Вкладка "Факт" (с внутренними подвкладками) ----------
        self.fact_tab = QWidget()
        self.main_tabs.addTab(self.fact_tab, "Факт")
        self.fact_sub_tabs = QTabWidget(self.fact_tab)
        fact_layout = QVBoxLayout(self.fact_tab)
        fact_layout.addWidget(self.fact_sub_tabs)

        self.setup_fact_subtabs()

        # ---------- Вкладка "План" (с внутренними подвкладками) ----------
        self.plan_tab = QWidget()
        self.main_tabs.addTab(self.plan_tab, "План")
        self.plan_sub_tabs = QTabWidget(self.plan_tab)
        plan_layout = QVBoxLayout(self.plan_tab)
        plan_layout.addWidget(self.plan_sub_tabs)

        self.setup_plan_subtabs()

        # ---------- Вкладка "Сравнение по месяцам" ----------
        self.compare_month_tab = QWidget()
        self.main_tabs.addTab(self.compare_month_tab, "Сравнение по месяцам")
        self.setup_compare_month_tab()

        # ---------- НОВАЯ ВКЛАДКА "Сравнение по кварталам" ----------
        self.compare_quarter_tab = QWidget()
        self.main_tabs.addTab(self.compare_quarter_tab, "Сравнение по кварталам")
        self.setup_compare_quarter_tab()

        # Кнопка закрытия и экспорта
        btn_layout = QHBoxLayout()
        btn_export = QPushButton("Сохранить отчёт в Excel")
        btn_export.clicked.connect(self.export_to_excel)
        btn_close = QPushButton("Закрыть")
        btn_close.clicked.connect(self.accept)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_export)
        btn_layout.addWidget(btn_close)
        layout.addLayout(btn_layout)

    # ---------- Вспомогательный метод для создания элемента легенды ----------
    def _create_legend_item(self, text, color_hex):
        item_widget = QWidget()
        item_layout = QHBoxLayout(item_widget)
        item_layout.setContentsMargins(0, 0, 0, 0)
        item_layout.setSpacing(3)

        color_label = QLabel()
        color_label.setFixedSize(16, 16)
        color_label.setStyleSheet(f"background-color: {color_hex}; border: 1px solid black;")
        item_layout.addWidget(color_label)

        text_label = QLabel(text)
        text_label.setStyleSheet("font-size: 8pt;")
        item_layout.addWidget(text_label)

        return item_widget

    # ========== МЕТОДЫ ДЛЯ ФАКТИЧЕСКИХ ДАННЫХ ==========
    def setup_fact_subtabs(self):
        self.fact_trend_tab = QWidget()
        self.fact_sub_tabs.addTab(self.fact_trend_tab, "Динамика ФОТ")
        self.setup_trend_tab(self.fact_trend_tab, is_plan=False)

        self.fact_structure_tab = QWidget()
        self.fact_sub_tabs.addTab(self.fact_structure_tab, "Структура начислений")
        self.setup_structure_tab(self.fact_structure_tab, is_plan=False)

        self.fact_quarters_tab = QWidget()
        self.fact_sub_tabs.addTab(self.fact_quarters_tab, "По кварталам")
        self.setup_quarters_tab(self.fact_quarters_tab, is_plan=False)

        self.fact_quarter_detail_tab = QWidget()
        self.fact_sub_tabs.addTab(self.fact_quarter_detail_tab, "Квартал помесячно")
        self.setup_quarter_detail_tab(self.fact_quarter_detail_tab, is_plan=False)

    # ========== МЕТОДЫ ДЛЯ ПЛАНОВЫХ ДАННЫХ ==========
    def setup_plan_subtabs(self):
        self.plan_trend_tab = QWidget()
        self.plan_sub_tabs.addTab(self.plan_trend_tab, "Динамика ФОТ")
        self.setup_trend_tab(self.plan_trend_tab, is_plan=True)

        self.plan_structure_tab = QWidget()
        self.plan_sub_tabs.addTab(self.plan_structure_tab, "Структура начислений")
        self.setup_structure_tab(self.plan_structure_tab, is_plan=True)

        self.plan_quarters_tab = QWidget()
        self.plan_sub_tabs.addTab(self.plan_quarters_tab, "По кварталам")
        self.setup_quarters_tab(self.plan_quarters_tab, is_plan=True)

        self.plan_quarter_detail_tab = QWidget()
        self.plan_sub_tabs.addTab(self.plan_quarter_detail_tab, "Квартал помесячно")
        self.setup_quarter_detail_tab(self.plan_quarter_detail_tab, is_plan=True)

    # ========== ОБЩИЕ МЕТОДЫ ДЛЯ НАСТРОЙКИ ВКЛАДОК (С ПАРАМЕТРОМ is_plan) ==========
    def setup_trend_tab(self, parent_tab, is_plan):
        layout = QVBoxLayout(parent_tab)

        control_layout = QHBoxLayout()
        control_layout.addWidget(QLabel("Тип графика:"))
        combo_type = QComboBox()
        combo_type.addItems(["Линейный", "Столбчатый", "Накопительный"])
        control_layout.addWidget(combo_type)

        control_layout.addWidget(QLabel("Период:"))
        combo_period = QComboBox()
        combo_period.addItems(["Все месяцы", "Последние 6", "Последние 12"])
        control_layout.addWidget(combo_period)
        control_layout.addStretch()
        layout.addLayout(control_layout)

        fig = Figure(figsize=(11, 4))
        canvas = FigureCanvas(fig)
        layout.addWidget(canvas)

        legend_widget = QWidget()
        legend_layout = QHBoxLayout(legend_widget)
        legend_layout.setContentsMargins(0, 0, 0, 0)
        legend_layout.setSpacing(10)
        layout.addWidget(legend_widget)

        parent_tab.combo_type = combo_type
        parent_tab.combo_period = combo_period
        parent_tab.fig = fig
        parent_tab.canvas = canvas
        parent_tab.legend_layout = legend_layout

        def update():
            self.update_trend_chart(parent_tab, is_plan)

        combo_type.currentIndexChanged.connect(update)
        combo_period.currentIndexChanged.connect(update)

        update()

    def setup_structure_tab(self, parent_tab, is_plan):
        layout = QVBoxLayout(parent_tab)

        month_layout = QHBoxLayout()
        month_layout.addWidget(QLabel("Месяц:"))
        month_combo = QComboBox()
        month_layout.addWidget(month_combo)

        month_layout.addWidget(QLabel("Тип:"))
        type_combo = QComboBox()
        type_combo.addItems(["Круговая", "Столбчатая"])
        month_layout.addWidget(type_combo)
        month_layout.addStretch()
        layout.addLayout(month_layout)

        months = sorted(self.project.months_data.keys())
        for y, m in months:
            month_combo.addItem(f"{m:02d}.{y}", (y, m))

        fig = Figure(figsize=(11, 4))
        canvas = FigureCanvas(fig)
        layout.addWidget(canvas)

        legend_widget = QWidget()
        legend_layout = QHBoxLayout(legend_widget)
        legend_layout.setContentsMargins(0, 0, 0, 0)
        legend_layout.setSpacing(10)
        layout.addWidget(legend_widget)

        table = QTableWidget(0, 2)
        table.setHorizontalHeaderLabels(["Показатель", "Сумма (руб.)"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(table)

        parent_tab.month_combo = month_combo
        parent_tab.type_combo = type_combo
        parent_tab.fig = fig
        parent_tab.canvas = canvas
        parent_tab.legend_layout = legend_layout
        parent_tab.table = table

        def update():
            self.update_structure_chart(parent_tab, is_plan)

        month_combo.currentIndexChanged.connect(update)
        type_combo.currentIndexChanged.connect(update)

        if month_combo.count() > 0:
            update()

    def setup_quarters_tab(self, parent_tab, is_plan):
        layout = QVBoxLayout(parent_tab)

        year_layout = QHBoxLayout()
        year_layout.addWidget(QLabel("Год:"))
        year_combo = QComboBox()
        year_layout.addWidget(year_combo)
        year_layout.addStretch()
        layout.addLayout(year_layout)

        table = QTableWidget(0, 7)
        table.setHorizontalHeaderLabels(["Квартал", "Оклад (руб.)", "Доплаты (руб.)", "Отпускные (руб.)",
                                         "Командировки (руб.)", "Больничные (руб.)", "ИТОГО (руб.)"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(table)

        fig = Figure(figsize=(11, 4))
        canvas = FigureCanvas(fig)
        layout.addWidget(canvas)

        legend_widget = QWidget()
        legend_layout = QHBoxLayout(legend_widget)
        legend_layout.setContentsMargins(0, 0, 0, 0)
        legend_layout.setSpacing(10)
        layout.addWidget(legend_widget)

        parent_tab.year_combo = year_combo
        parent_tab.table = table
        parent_tab.fig = fig
        parent_tab.canvas = canvas
        parent_tab.legend_layout = legend_layout

        years = set()
        for y, m in self.project.months_data.keys():
            years.add(y)
        year_combo.addItem("Все годы", None)
        for y in sorted(years):
            year_combo.addItem(str(y), y)

        year_combo.currentIndexChanged.connect(lambda: self.update_quarters_chart(parent_tab, is_plan))

        self.update_quarters_chart(parent_tab, is_plan)

    def setup_quarter_detail_tab(self, parent_tab, is_plan):
        layout = QVBoxLayout(parent_tab)

        control_layout = QHBoxLayout()
        control_layout.addWidget(QLabel("Год:"))
        year_combo = QComboBox()
        control_layout.addWidget(year_combo)

        control_layout.addWidget(QLabel("Квартал:"))
        q_combo = QComboBox()
        q_combo.addItems(["I", "II", "III", "IV"])
        control_layout.addWidget(q_combo)

        control_layout.addWidget(QLabel("Тип:"))
        type_combo = QComboBox()
        type_combo.addItems(["Столбчатая"])
        control_layout.addWidget(type_combo)
        control_layout.addStretch()
        layout.addLayout(control_layout)

        table = QTableWidget(0, 7)
        table.setHorizontalHeaderLabels(["Месяц", "Оклад (руб.)", "Доплаты (руб.)", "Отпускные (руб.)",
                                         "Командировки (руб.)", "Больничные (руб.)", "ИТОГО (руб.)"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(table)

        fig = Figure(figsize=(11, 4))
        canvas = FigureCanvas(fig)
        layout.addWidget(canvas)

        legend_widget = QWidget()
        legend_layout = QHBoxLayout(legend_widget)
        legend_layout.setContentsMargins(0, 0, 0, 0)
        legend_layout.setSpacing(10)
        layout.addWidget(legend_widget)

        parent_tab.year_combo = year_combo
        parent_tab.q_combo = q_combo
        parent_tab.type_combo = type_combo
        parent_tab.table = table
        parent_tab.fig = fig
        parent_tab.canvas = canvas
        parent_tab.legend_layout = legend_layout

        years = set()
        for y, m in self.project.months_data.keys():
            years.add(y)
        for y in sorted(years):
            year_combo.addItem(str(y), y)

        def update():
            self.update_quarter_detail_chart(parent_tab, is_plan)

        year_combo.currentIndexChanged.connect(update)
        q_combo.currentIndexChanged.connect(update)
        type_combo.currentIndexChanged.connect(update)

        if year_combo.count() > 0:
            year_combo.setCurrentIndex(0)
            update()

    # ========== МЕТОДЫ ОБНОВЛЕНИЯ ГРАФИКОВ (С ПАРАМЕТРОМ is_plan) ==========
    def update_trend_chart(self, parent_tab, is_plan):
        fig = parent_tab.fig
        canvas = parent_tab.canvas
        legend_layout = parent_tab.legend_layout
        chart_type_idx = parent_tab.combo_type.currentIndex()
        period_idx = parent_tab.combo_period.currentIndex()

        fig.clear()
        ax = fig.add_subplot(111)

        all_months = sorted(self.project.months_data.keys())
        if not all_months:
            ax.text(0.5, 0.5, "Нет данных", ha='center', va='center', transform=ax.transAxes)
            canvas.draw()
            return

        if period_idx == 1:
            months = all_months[-6:]
        elif period_idx == 2:
            months = all_months[-12:]
        else:
            months = all_months

        labels = []
        totals = []
        salary_totals = []
        allowance_totals = []

        for y, m in months:
            if is_plan:
                plan_month = self.get_plan_data_for_month(y, m)
                month_total = plan_month.get('total', 0)
                salary_total = plan_month.get('salary', 0)
                allowance_total = plan_month.get('allowance', 0)
            else:
                calculator = PayrollCalculator(self.project, self.holidays)
                summary, _ = calculator.calculate_for_month(y, m)
                month_total = sum(row[7] for row in summary) if summary else 0
                salary_total = sum(row[2] for row in summary) if summary else 0
                allowance_total = sum(row[3] for row in summary) if summary else 0

            labels.append(f"{m:02d}.{y}")
            totals.append(month_total)
            salary_totals.append(salary_total)
            allowance_totals.append(allowance_total)

        x_pos = range(len(labels))

        if chart_type_idx == 0:  # линейный
            ax.plot(x_pos, totals, marker='o', linewidth=2, color='#2c3e50')
            ax.plot(x_pos, salary_totals, marker='s', linewidth=1.5, color='#27ae60')
            ax.plot(x_pos, allowance_totals, marker='^', linewidth=1.5, color='#e67e22')
            # подписи значений
            for i, (x, y) in enumerate(zip(x_pos, totals)):
                ax.annotate(f'{y:,.0f}', (x, y), textcoords="offset points", xytext=(0, 10), ha='center', fontsize=8)
            for i, (x, y) in enumerate(zip(x_pos, salary_totals)):
                ax.annotate(f'{y:,.0f}', (x, y), textcoords="offset points", xytext=(0, -15), ha='center', fontsize=8,
                            color='#27ae60')
            for i, (x, y) in enumerate(zip(x_pos, allowance_totals)):
                ax.annotate(f'{y:,.0f}', (x, y), textcoords="offset points", xytext=(0, -30), ha='center', fontsize=8,
                            color='#e67e22')
        elif chart_type_idx == 1:  # столбчатый
            x = np.arange(len(labels))
            width = 0.25
            bars1 = ax.bar(x - width, salary_totals, width, color='#27ae60')
            bars2 = ax.bar(x, allowance_totals, width, color='#e67e22')
            bars3 = ax.bar(x + width, totals, width, color='#2c3e50')
            for bars in [bars1, bars2, bars3]:
                for bar in bars:
                    height = bar.get_height()
                    if height > 0:
                        ax.annotate(f'{height:,.0f}', xy=(bar.get_x() + bar.get_width() / 2, height),
                                    xytext=(0, 3), textcoords="offset points", ha='center', va='bottom', fontsize=8)
            # метки оси X будут установлены ниже общим кодом
        else:  # накопительный (chart_type_idx == 2)
            bars = ax.bar(labels, totals, color='#3498db')  # используем labels как позиции
            for bar in bars:
                height = bar.get_height()
                if height > 0:
                    ax.annotate(f'{height:,.0f}', xy=(bar.get_x() + bar.get_width() / 2, height),
                                xytext=(0, 3), textcoords="offset points", ha='center', va='bottom', fontsize=8)
            # для этого типа ax.bar с labels уже создал категории, но нам нужно установить метки на оси
            # мы можем использовать числовые позиции для единообразия, но оставим как есть и потом установим метки
            # фактически, при таком вызове ax.bar(labels, totals) метки осей будут labels, но они могут быть перекрыты
            # добавим поворот позже

        # Общая настройка оси X для всех типов, кроме, возможно, накопительного (но применим и к нему)
        # Чтобы избежать дублирования, всегда используем числовые позиции для линейного и столбчатого,
        # а для накопительного оставим labels как есть, но тогда нужно применить поворот.
        # Проще всего: после построения графика задать метки через set_xticks и set_xticklabels,
        # но для накопительного, если использовали labels, позиции уже соответствуют.
        # Чтобы единообразно, для всех типов используем range(len(labels)) как позиции.
        # Для столбчатого у нас уже есть x = np.arange(len(labels)), используем его.
        # Для линейного используем x_pos = range(len(labels)).
        # Для накопительного переделаем: bars = ax.bar(range(len(labels)), totals) и потом зададим метки.
        # Это самый чистый способ.

        # Переделаем накопительный на использование числовых позиций:
        if chart_type_idx == 2:
            # заново очистим и перерисуем
            fig.clear()
            ax = fig.add_subplot(111)
            x = np.arange(len(labels))
            bars = ax.bar(x, totals, color='#3498db')
            for bar in bars:
                height = bar.get_height()
                if height > 0:
                    ax.annotate(f'{height:,.0f}', xy=(bar.get_x() + bar.get_width() / 2, height),
                                xytext=(0, 3), textcoords="offset points", ha='center', va='bottom', fontsize=8)
            ax.set_xticks(x)
            ax.set_xticklabels(labels, rotation=45, ha='right')
        else:
            # для линейного и столбчатого уже построили, теперь зададим метки
            ax.set_xticks(range(len(labels)))
            ax.set_xticklabels(labels, rotation=45, ha='right')

        ax.set_ylabel("ФОТ (руб.)")
        ax.set_title(("Плановые" if is_plan else "Фактические") + " показатели ФОТ по месяцам")
        ax.grid(True, alpha=0.3, linestyle='--')
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:,.0f}'))

        if legend_layout is not None:
            for i in reversed(range(legend_layout.count())):
                item = legend_layout.itemAt(i)
                widget = item.widget()
                if widget is not None:
                    widget.deleteLater()
            categories = ["Общий ФОТ", "Окладная часть", "Доплаты"]
            colors = ['#2c3e50', '#27ae60', '#e67e22']
            for cat, col in zip(categories, colors):
                legend_layout.addWidget(self._create_legend_item(cat, col))
            legend_layout.addStretch()

        fig.tight_layout()
        canvas.draw()


    def update_structure_chart(self, parent_tab, is_plan):
        fig = parent_tab.fig
        canvas = parent_tab.canvas
        legend_layout = parent_tab.legend_layout
        table = parent_tab.table
        selected = parent_tab.month_combo.currentData()
        chart_type = parent_tab.type_combo.currentIndex()

        if not selected:
            return
        y, m = selected

        if is_plan:
            data = self.get_plan_data_for_month(y, m)
            salary_total = data.get('salary', 0)
            allowance_total = data.get('allowance', 0)
            vacation_total = data.get('vacation', 0)
            kom_total = data.get('kom', 0)
            sick_total = data.get('sick', 0)
            special_total = data.get('special', 0)
        else:
            calculator = PayrollCalculator(self.project, self.holidays)
            summary, _ = calculator.calculate_for_month(y, m)
            if not summary:
                salary_total = allowance_total = vacation_total = kom_total = sick_total = special_total = 0
            else:
                salary_total = sum(row[2] for row in summary)
                allowance_total = sum(row[3] for row in summary)
                vacation_total = sum(row[4] for row in summary)
                kom_total = sum(row[5] for row in summary)
                sick_total = sum(row[6] for row in summary)
                total = sum(row[7] for row in summary)
                special_total = total - (salary_total + allowance_total + vacation_total + kom_total + sick_total)

        categories = ['Оклад', 'Доплаты', 'Отпускные', 'Командировки', 'Больничные', 'Спец.оплата']
        values = [salary_total, allowance_total, vacation_total, kom_total, sick_total, special_total]

        non_zero = [(cat, val) for cat, val in zip(categories, values) if val > 0]
        cats, vals = zip(*non_zero) if non_zero else ([], [])

        fig.clear()
        ax = fig.add_subplot(111)

        if legend_layout is not None:
            for i in reversed(range(legend_layout.count())):
                item = legend_layout.itemAt(i)
                widget = item.widget()
                if widget is not None:
                    widget.deleteLater()

        if chart_type == 0:  # круговая
            if not cats:
                ax.text(0.5, 0.5, "Нет данных", ha='center', va='center', transform=ax.transAxes)
            else:
                wedges, texts, autotexts = ax.pie(vals, autopct='%1.1f%%', startangle=90)
                ax.axis('equal')
                ax.set_title(f"Структура начислений за {m:02d}.{y} ({'план' if is_plan else 'факт'})")
                for cat, wedge in zip(cats, wedges):
                    rgba = wedge.get_facecolor()
                    if hasattr(rgba, '__iter__') and not isinstance(rgba, str):
                        r, g, b, a = rgba[:4]
                    else:
                        r, g, b = 0, 0, 0
                    color_hex = "#{:02x}{:02x}{:02x}".format(int(r*255), int(g*255), int(b*255))
                    legend_layout.addWidget(self._create_legend_item(cat, color_hex))
        else:  # столбчатая
            if not cats:
                ax.text(0.5, 0.5, "Нет данных", ha='center', va='center', transform=ax.transAxes)
            else:
                x = np.arange(len(cats))
                bars = ax.bar(x, vals, color=plt.cm.Set3.colors[:len(cats)])
                for bar in bars:
                    height = bar.get_height()
                    if height > 0:
                        ax.annotate(f'{height:,.0f}', xy=(bar.get_x()+bar.get_width()/2, height),
                                    xytext=(0,3), textcoords="offset points", ha='center', va='bottom', fontsize=9)
                ax.set_xticks(x)
                ax.set_xticklabels(cats, rotation=45, ha='right')
                ax.set_ylabel("Сумма (руб.)")
                ax.set_title(f"Структура начислений за {m:02d}.{y} ({'план' if is_plan else 'факт'})")
                ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:,.0f}'))
                for cat, bar in zip(cats, bars):
                    rgba = bar.get_facecolor()
                    if hasattr(rgba, '__iter__') and not isinstance(rgba, str):
                        r, g, b, a = rgba[:4]
                    else:
                        r, g, b = 0, 0, 0
                    color_hex = "#{:02x}{:02x}{:02x}".format(int(r*255), int(g*255), int(b*255))
                    legend_layout.addWidget(self._create_legend_item(cat, color_hex))

        if legend_layout is not None:
            legend_layout.addStretch()
        fig.tight_layout()
        canvas.draw()

        table.setRowCount(len(non_zero) + 1)
        total = sum(vals)
        for i, (cat, val) in enumerate(non_zero):
            table.setItem(i, 0, QTableWidgetItem(cat))
            table.setItem(i, 1, QTableWidgetItem(f"{val:,.2f}"))
        table.setItem(len(non_zero), 0, QTableWidgetItem("ИТОГО"))
        table.setItem(len(non_zero), 1, QTableWidgetItem(f"{total:,.2f}"))

    def update_quarters_chart(self, parent_tab, is_plan):
        selected_year = parent_tab.year_combo.currentData()

        quarters_data = {q: {"salary":0, "allowance":0, "vacation":0,
                             "kom":0, "sick":0, "special":0} for q in range(1,5)}

        for (y, m), data in self.project.months_data.items():
            if selected_year is not None and y != selected_year:
                continue
            q = (m - 1) // 3 + 1
            if is_plan:
                month_data = self.get_plan_data_for_month(y, m)
                salary = month_data.get('salary', 0)
                allowance = month_data.get('allowance', 0)
                vacation = month_data.get('vacation', 0)
                kom = month_data.get('kom', 0)
                sick = month_data.get('sick', 0)
                special = month_data.get('special', 0)
            else:
                calculator = PayrollCalculator(self.project, self.holidays)
                summary, _ = calculator.calculate_for_month(y, m)
                if not summary:
                    continue
                salary = sum(row[2] for row in summary)
                allowance = sum(row[3] for row in summary)
                vacation = sum(row[4] for row in summary)
                kom = sum(row[5] for row in summary)
                sick = sum(row[6] for row in summary)
                total = sum(row[7] for row in summary)
                special = total - (salary + allowance + vacation + kom + sick)

            quarters_data[q]["salary"] += salary
            quarters_data[q]["allowance"] += allowance
            quarters_data[q]["vacation"] += vacation
            quarters_data[q]["kom"] += kom
            quarters_data[q]["sick"] += sick
            quarters_data[q]["special"] += special

        labels = ["I", "II", "III", "IV"]
        parent_tab.table.setRowCount(4)
        for q in range(1,5):
            data = quarters_data[q]
            total = data["salary"]+data["allowance"]+data["vacation"]+data["kom"]+data["sick"]+data["special"]
            parent_tab.table.setItem(q-1, 0, QTableWidgetItem(labels[q-1]))
            parent_tab.table.setItem(q-1, 1, QTableWidgetItem(f"{data['salary']:,.0f}"))
            parent_tab.table.setItem(q-1, 2, QTableWidgetItem(f"{data['allowance']:,.0f}"))
            parent_tab.table.setItem(q-1, 3, QTableWidgetItem(f"{data['vacation']:,.0f}"))
            parent_tab.table.setItem(q-1, 4, QTableWidgetItem(f"{data['kom']:,.0f}"))
            parent_tab.table.setItem(q-1, 5, QTableWidgetItem(f"{data['sick']:,.0f}"))
            parent_tab.table.setItem(q-1, 6, QTableWidgetItem(f"{total:,.0f}"))

        fig = parent_tab.fig
        canvas = parent_tab.canvas
        legend_layout = parent_tab.legend_layout
        fig.clear()
        ax = fig.add_subplot(111)
        x = np.arange(4)
        width = 0.12
        categories_keys = ["salary", "allowance", "vacation", "kom", "sick", "special"]
        colors = ['#2c3e50', '#27ae60', '#f1c40f', '#e67e22', '#e74c3c', '#9b59b6']
        names = ["Оклад", "Доплаты", "Отпускные", "Командировки", "Больничные", "Спец."]

        for i, (key, col, name) in enumerate(zip(categories_keys, colors, names)):
            values = [quarters_data[q][key] for q in range(1,5)]
            bars = ax.bar(x + i*width, values, width, color=col)
            for bar in bars:
                height = bar.get_height()
                if height > 0:
                    ax.annotate(f'{height:,.0f}', xy=(bar.get_x()+bar.get_width()/2, height),
                                xytext=(0,3), textcoords="offset points", ha='center', va='bottom', fontsize=7)

        ax.set_xlabel("Квартал")
        ax.set_ylabel("Сумма (руб.)")
        ax.set_title(("Плановые" if is_plan else "Фактические") + " показатели по кварталам" +
                     (f" за {selected_year} г." if selected_year else " (все годы)"))
        ax.set_xticks(x + width*(len(categories_keys)-1)/2)
        ax.set_xticklabels(labels)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:,.0f}'))
        fig.tight_layout()
        canvas.draw()

        if legend_layout is not None:
            for i in reversed(range(legend_layout.count())):
                item = legend_layout.itemAt(i)
                widget = item.widget()
                if widget is not None:
                    widget.deleteLater()
            for cat, col in zip(names, colors):
                legend_layout.addWidget(self._create_legend_item(cat, col))
            legend_layout.addStretch()

    def update_quarter_detail_chart(self, parent_tab, is_plan):
        year = parent_tab.year_combo.currentData()
        q_index = parent_tab.q_combo.currentIndex()
        if year is None:
            return

        start_month = q_index * 3 + 1
        end_month = start_month + 2
        months = [(year, m) for m in range(start_month, end_month + 1)]

        labels = []
        salary_list = []
        allowance_list = []
        vacation_list = []
        kom_list = []
        sick_list = []
        special_list = []

        for y, m in months:
            if is_plan:
                month_data = self.get_plan_data_for_month(y, m)
                salary = month_data.get('salary', 0)
                allowance = month_data.get('allowance', 0)
                vacation = month_data.get('vacation', 0)
                kom = month_data.get('kom', 0)
                sick = month_data.get('sick', 0)
                special = month_data.get('special', 0)
            else:
                calculator = PayrollCalculator(self.project, self.holidays)
                summary, _ = calculator.calculate_for_month(y, m)
                if not summary:
                    salary = allowance = vacation = kom = sick = special = 0
                else:
                    salary = sum(row[2] for row in summary)
                    allowance = sum(row[3] for row in summary)
                    vacation = sum(row[4] for row in summary)
                    kom = sum(row[5] for row in summary)
                    sick = sum(row[6] for row in summary)
                    total = sum(row[7] for row in summary)
                    special = total - (salary + allowance + vacation + kom + sick)

            salary_list.append(salary)
            allowance_list.append(allowance)
            vacation_list.append(vacation)
            kom_list.append(kom)
            sick_list.append(sick)
            special_list.append(special)
            labels.append(f"{m:02d}")

        table = parent_tab.table
        table.setRowCount(3)
        for i in range(3):
            table.setItem(i, 0, QTableWidgetItem(labels[i]))
            table.setItem(i, 1, QTableWidgetItem(f"{salary_list[i]:,.0f}"))
            table.setItem(i, 2, QTableWidgetItem(f"{allowance_list[i]:,.0f}"))
            table.setItem(i, 3, QTableWidgetItem(f"{vacation_list[i]:,.0f}"))
            table.setItem(i, 4, QTableWidgetItem(f"{kom_list[i]:,.0f}"))
            table.setItem(i, 5, QTableWidgetItem(f"{sick_list[i]:,.0f}"))
            total = salary_list[i] + allowance_list[i] + vacation_list[i] + kom_list[i] + sick_list[i] + special_list[i]
            table.setItem(i, 6, QTableWidgetItem(f"{total:,.0f}"))

        fig = parent_tab.fig
        canvas = parent_tab.canvas
        legend_layout = parent_tab.legend_layout
        fig.clear()
        ax = fig.add_subplot(111)
        x = np.arange(len(labels))
        width = 0.12
        data_lists = [salary_list, allowance_list, vacation_list, kom_list, sick_list, special_list]
        colors = ['#2c3e50', '#27ae60', '#f1c40f', '#e67e22', '#e74c3c', '#9b59b6']
        names = ["Оклад", "Доплаты", "Отпускные", "Командировки", "Больничные", "Спец."]

        for i, (vals, col) in enumerate(zip(data_lists, colors)):
            bars = ax.bar(x + i*width, vals, width, color=col)
            for bar in bars:
                height = bar.get_height()
                if height > 0:
                    ax.annotate(f'{height:,.0f}', xy=(bar.get_x()+bar.get_width()/2, height),
                                xytext=(0,3), textcoords="offset points", ha='center', va='bottom', fontsize=7)

        ax.set_xlabel("Месяц")
        ax.set_ylabel("Сумма (руб.)")
        ax.set_title(f"Помесячная структура {['I','II','III','IV'][q_index]} квартала {year} г. ({'план' if is_plan else 'факт'})")
        ax.set_xticks(x + width*(len(data_lists)-1)/2)
        ax.set_xticklabels(labels)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:,.0f}'))
        fig.tight_layout()
        canvas.draw()

        if legend_layout is not None:
            for i in reversed(range(legend_layout.count())):
                item = legend_layout.itemAt(i)
                widget = item.widget()
                if widget is not None:
                    widget.deleteLater()
            for name, col in zip(names, colors):
                legend_layout.addWidget(self._create_legend_item(name, col))
            legend_layout.addStretch()

    # ========== МЕТОДЫ ДЛЯ ПОЛУЧЕНИЯ ПЛАНОВЫХ ДАННЫХ ==========
    def get_plan_data_for_month(self, year: int, month: int) -> Dict[str, float]:
        quarter = (month - 1) // 3 + 1
        plan_key = (year, quarter)
        plan_data = self.project.plans.get(plan_key)
        if not plan_data:
            return {'salary':0, 'allowance':0, 'vacation':0, 'kom':0, 'sick':0, 'special':0, 'total':0}

        cells = []
        for mdata in plan_data.get('months', []):
            if mdata['year'] == year and mdata['month'] == month:
                cells = mdata.get('cells', [])
                break

        if not cells:
            return {'salary':0, 'allowance':0, 'vacation':0, 'kom':0, 'sick':0, 'special':0, 'total':0}

        data = {}
        for cell in cells:
            emp_idx = cell.get('emp_idx')
            day = cell.get('day')
            code = cell.get('code', '')
            option = cell.get('option', 'normal')
            if emp_idx is not None and 0 <= emp_idx < len(self.project.employees):
                data[(emp_idx, day)] = (code, '', option)

        norm = self.project.get_norm(year, month)
        calculator = PayrollCalculator(self.project, self.holidays)
        summary, _ = calculator.calculate_for_month_with_data(year, month, data, norm)

        if not summary:
            return {'salary':0, 'allowance':0, 'vacation':0, 'kom':0, 'sick':0, 'special':0, 'total':0}

        salary = sum(row[2] for row in summary)
        allowance = sum(row[3] for row in summary)
        vacation = sum(row[4] for row in summary)
        kom = sum(row[5] for row in summary)
        sick = sum(row[6] for row in summary)
        total = sum(row[7] for row in summary)
        special = total - (salary + allowance + vacation + kom + sick)

        return {
            'salary': salary,
            'allowance': allowance,
            'vacation': vacation,
            'kom': kom,
            'sick': sick,
            'special': special,
            'total': total
        }

    # ========== ВКЛАДКА СРАВНЕНИЯ ПО МЕСЯЦАМ ==========
    def setup_compare_month_tab(self):
        layout = QVBoxLayout(self.compare_month_tab)

        control_layout = QHBoxLayout()
        control_layout.addWidget(QLabel("Год:"))
        self.cmp_month_year_combo = QComboBox()
        years = set()
        for y, m in self.project.months_data.keys():
            years.add(y)
        for y in sorted(years):
            self.cmp_month_year_combo.addItem(str(y), y)
        control_layout.addWidget(self.cmp_month_year_combo)

        control_layout.addWidget(QLabel("Показатель:"))
        self.cmp_month_metric_combo = QComboBox()
        self.cmp_month_metric_combo.addItems(["Общий ФОТ", "Оклад", "Доплаты", "Отпускные", "Командировки", "Больничные", "Спец.оплата"])
        control_layout.addWidget(self.cmp_month_metric_combo)

        self.cmp_month_update_btn = QPushButton("Обновить")
        self.cmp_month_update_btn.clicked.connect(self.update_compare_month_chart)
        control_layout.addWidget(self.cmp_month_update_btn)
        control_layout.addStretch()
        layout.addLayout(control_layout)

        self.cmp_month_table = QTableWidget(12, 5)
        self.cmp_month_table.setHorizontalHeaderLabels(["Месяц", "План (руб.)", "Факт (руб.)", "Отклонение (руб.)", "% выполнения"])
        self.cmp_month_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.cmp_month_table)

        self.cmp_month_fig = Figure(figsize=(11, 4))
        self.cmp_month_canvas = FigureCanvas(self.cmp_month_fig)
        layout.addWidget(self.cmp_month_canvas)

        export_btn = QPushButton("Экспорт в Excel")
        export_btn.clicked.connect(self.export_compare_month)
        layout.addWidget(export_btn, alignment=Qt.AlignmentFlag.AlignRight)

        self.update_compare_month_chart()

    def update_compare_month_chart(self):
        year = self.cmp_month_year_combo.currentData()
        if year is None:
            return

        metric_idx = self.cmp_month_metric_combo.currentIndex()
        metric_keys = ['total', 'salary', 'allowance', 'vacation', 'kom', 'sick', 'special']
        metric_key = metric_keys[metric_idx]

        months = list(range(1, 13))
        plan_vals = []
        fact_vals = []
        labels = []

        for m in months:
            plan = self.get_plan_data_for_month(year, m)
            fact = self.get_fact_data_for_month(year, m)

            plan_val = plan.get(metric_key, 0)
            fact_val = fact.get(metric_key, 0)
            plan_vals.append(plan_val)
            fact_vals.append(fact_val)
            labels.append(f"{m:02d}")

        self.cmp_month_table.setRowCount(12)
        for i in range(12):
            self.cmp_month_table.setItem(i, 0, QTableWidgetItem(labels[i]))
            self.cmp_month_table.setItem(i, 1, QTableWidgetItem(f"{plan_vals[i]:,.2f}"))
            self.cmp_month_table.setItem(i, 2, QTableWidgetItem(f"{fact_vals[i]:,.2f}"))
            diff = fact_vals[i] - plan_vals[i]
            self.cmp_month_table.setItem(i, 3, QTableWidgetItem(f"{diff:+,.2f}"))
            percent = (fact_vals[i] / plan_vals[i] * 100) if plan_vals[i] != 0 else 0.0
            self.cmp_month_table.setItem(i, 4, QTableWidgetItem(f"{percent:.1f}%"))

        self.cmp_month_fig.clear()
        ax = self.cmp_month_fig.add_subplot(111)

        x = np.arange(len(labels))
        width = 0.35
        bars1 = ax.bar(x - width/2, plan_vals, width, label='План', color='steelblue')
        bars2 = ax.bar(x + width/2, fact_vals, width, label='Факт', color='darkorange')

        ax.set_xlabel('Месяц')
        ax.set_ylabel('Сумма (руб.)')
        metric_name = self.cmp_month_metric_combo.currentText()
        ax.set_title(f'Сравнение план/факт по месяцам ({metric_name}) за {year} г.')
        ax.set_xticks(x)
        ax.set_xticklabels(labels)
        ax.legend()
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:,.0f}'))

        for bars in [bars1, bars2]:
            for bar in bars:
                height = bar.get_height()
                if height > 0:
                    ax.annotate(f'{height:,.0f}',
                                xy=(bar.get_x() + bar.get_width()/2, height),
                                xytext=(0, 3), textcoords="offset points",
                                ha='center', va='bottom', fontsize=7)

        self.cmp_month_fig.tight_layout()
        self.cmp_month_canvas.draw()

    def get_fact_data_for_month(self, year: int, month: int) -> Dict[str, float]:
        calculator = PayrollCalculator(self.project, self.holidays)
        summary, _ = calculator.calculate_for_month(year, month)
        if not summary:
            return {'salary':0, 'allowance':0, 'vacation':0, 'kom':0, 'sick':0, 'special':0, 'total':0}
        salary = sum(row[2] for row in summary)
        allowance = sum(row[3] for row in summary)
        vacation = sum(row[4] for row in summary)
        kom = sum(row[5] for row in summary)
        sick = sum(row[6] for row in summary)
        total = sum(row[7] for row in summary)
        special = total - (salary + allowance + vacation + kom + sick)
        return {
            'salary': salary,
            'allowance': allowance,
            'vacation': vacation,
            'kom': kom,
            'sick': sick,
            'special': special,
            'total': total
        }

    def export_compare_month(self):
        fname, _ = QFileDialog.getSaveFileName(self, "Сохранить сравнение по месяцам", "", "Excel Files (*.xlsx)")
        if not fname:
            return
        if not fname.endswith('.xlsx'):
            fname += '.xlsx'
        try:
            import pandas as pd
            data = []
            for row in range(self.cmp_month_table.rowCount()):
                row_data = []
                for col in range(self.cmp_month_table.columnCount()):
                    item = self.cmp_month_table.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)
            headers = [self.cmp_month_table.horizontalHeaderItem(i).text() for i in range(self.cmp_month_table.columnCount())]
            df = pd.DataFrame(data, columns=headers)
            df.to_excel(fname, index=False)
            QMessageBox.information(self, "Экспорт", "Данные сохранены.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл:\n{e}")

    # ========== НОВАЯ ВКЛАДКА СРАВНЕНИЯ ПО КВАРТАЛАМ ==========
    def setup_compare_quarter_tab(self):
        layout = QVBoxLayout(self.compare_quarter_tab)

        control_layout = QHBoxLayout()
        control_layout.addWidget(QLabel("Год:"))
        self.cmp_quarter_year_combo = QComboBox()
        years = set()
        for y, m in self.project.months_data.keys():
            years.add(y)
        for y in sorted(years):
            self.cmp_quarter_year_combo.addItem(str(y), y)
        control_layout.addWidget(self.cmp_quarter_year_combo)

        control_layout.addWidget(QLabel("Показатель:"))
        self.cmp_quarter_metric_combo = QComboBox()
        self.cmp_quarter_metric_combo.addItems(["Общий ФОТ", "Оклад", "Доплаты", "Отпускные", "Командировки", "Больничные", "Спец.оплата"])
        control_layout.addWidget(self.cmp_quarter_metric_combo)

        self.cmp_quarter_update_btn = QPushButton("Обновить")
        self.cmp_quarter_update_btn.clicked.connect(self.update_compare_quarter_chart)
        control_layout.addWidget(self.cmp_quarter_update_btn)
        control_layout.addStretch()
        layout.addLayout(control_layout)

        self.cmp_quarter_table = QTableWidget(4, 5)
        self.cmp_quarter_table.setHorizontalHeaderLabels(["Квартал", "План (руб.)", "Факт (руб.)", "Отклонение (руб.)", "% выполнения"])
        self.cmp_quarter_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.cmp_quarter_table)

        self.cmp_quarter_fig = Figure(figsize=(11, 4))
        self.cmp_quarter_canvas = FigureCanvas(self.cmp_quarter_fig)
        layout.addWidget(self.cmp_quarter_canvas)

        export_btn = QPushButton("Экспорт в Excel")
        export_btn.clicked.connect(self.export_compare_quarter)
        layout.addWidget(export_btn, alignment=Qt.AlignmentFlag.AlignRight)

        self.update_compare_quarter_chart()

    def update_compare_quarter_chart(self):
        year = self.cmp_quarter_year_combo.currentData()
        if year is None:
            return

        metric_idx = self.cmp_quarter_metric_combo.currentIndex()
        metric_keys = ['total', 'salary', 'allowance', 'vacation', 'kom', 'sick', 'special']
        metric_key = metric_keys[metric_idx]

        quarters = [1, 2, 3, 4]
        plan_vals = []
        fact_vals = []
        labels = ["I", "II", "III", "IV"]

        for q in quarters:
            plan_sum = 0
            fact_sum = 0
            for m in range(q*3 - 2, q*3 + 1):
                plan = self.get_plan_data_for_month(year, m)
                fact = self.get_fact_data_for_month(year, m)
                plan_sum += plan.get(metric_key, 0)
                fact_sum += fact.get(metric_key, 0)
            plan_vals.append(plan_sum)
            fact_vals.append(fact_sum)

        self.cmp_quarter_table.setRowCount(4)
        for i in range(4):
            self.cmp_quarter_table.setItem(i, 0, QTableWidgetItem(labels[i]))
            self.cmp_quarter_table.setItem(i, 1, QTableWidgetItem(f"{plan_vals[i]:,.2f}"))
            self.cmp_quarter_table.setItem(i, 2, QTableWidgetItem(f"{fact_vals[i]:,.2f}"))
            diff = fact_vals[i] - plan_vals[i]
            self.cmp_quarter_table.setItem(i, 3, QTableWidgetItem(f"{diff:+,.2f}"))
            percent = (fact_vals[i] / plan_vals[i] * 100) if plan_vals[i] != 0 else 0.0
            self.cmp_quarter_table.setItem(i, 4, QTableWidgetItem(f"{percent:.1f}%"))

        self.cmp_quarter_fig.clear()
        ax = self.cmp_quarter_fig.add_subplot(111)

        x = np.arange(len(labels))
        width = 0.35
        bars1 = ax.bar(x - width/2, plan_vals, width, label='План', color='steelblue')
        bars2 = ax.bar(x + width/2, fact_vals, width, label='Факт', color='darkorange')

        ax.set_xlabel('Квартал')
        ax.set_ylabel('Сумма (руб.)')
        metric_name = self.cmp_quarter_metric_combo.currentText()
        ax.set_title(f'Сравнение план/факт по кварталам ({metric_name}) за {year} г.')
        ax.set_xticks(x)
        ax.set_xticklabels(labels)
        ax.legend()
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:,.0f}'))

        for bars in [bars1, bars2]:
            for bar in bars:
                height = bar.get_height()
                if height > 0:
                    ax.annotate(f'{height:,.0f}',
                                xy=(bar.get_x() + bar.get_width()/2, height),
                                xytext=(0, 3), textcoords="offset points",
                                ha='center', va='bottom', fontsize=7)

        self.cmp_quarter_fig.tight_layout()
        self.cmp_quarter_canvas.draw()

    def export_compare_quarter(self):
        fname, _ = QFileDialog.getSaveFileName(self, "Сохранить сравнение по кварталам", "", "Excel Files (*.xlsx)")
        if not fname:
            return
        if not fname.endswith('.xlsx'):
            fname += '.xlsx'
        try:
            import pandas as pd
            data = []
            for row in range(self.cmp_quarter_table.rowCount()):
                row_data = []
                for col in range(self.cmp_quarter_table.columnCount()):
                    item = self.cmp_quarter_table.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)
            headers = [self.cmp_quarter_table.horizontalHeaderItem(i).text() for i in range(self.cmp_quarter_table.columnCount())]
            df = pd.DataFrame(data, columns=headers)
            df.to_excel(fname, index=False)
            QMessageBox.information(self, "Экспорт", "Данные сохранены.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл:\n{e}")

    # ========== ОБЩИЙ ЭКСПОРТ В EXCEL (расширенный) ==========
    def export_to_excel(self):
        # Здесь можно объединить экспорт всех вкладок, но для простоты экспортируем только сравнения
        # Можно вызвать оба метода экспорта или создать единый отчёт.
        # В текущей реализации вызываем экспорт сравнения по месяцам и кварталам в разные файлы.
        # Но для удобства можно сохранить все данные в один файл с несколькими листами.
        # Реализуем единый экспорт в один файл с несколькими листами.
        fname, _ = QFileDialog.getSaveFileName(self, "Сохранить отчёт в Excel", "", "Excel Files (*.xlsx)")
        if not fname:
            return
        if not fname.endswith('.xlsx'):
            fname += '.xlsx'

        try:
            import pandas as pd
            with pd.ExcelWriter(fname, engine='openpyxl') as writer:
                # Лист: Сравнение по месяцам
                data_month = []
                for row in range(self.cmp_month_table.rowCount()):
                    row_data = []
                    for col in range(self.cmp_month_table.columnCount()):
                        item = self.cmp_month_table.item(row, col)
                        row_data.append(item.text() if item else "")
                    data_month.append(row_data)
                headers_month = [self.cmp_month_table.horizontalHeaderItem(i).text() for i in range(self.cmp_month_table.columnCount())]
                df_month = pd.DataFrame(data_month, columns=headers_month)
                df_month.to_excel(writer, sheet_name="Сравнение по месяцам", index=False)

                # Лист: Сравнение по кварталам
                data_quarter = []
                for row in range(self.cmp_quarter_table.rowCount()):
                    row_data = []
                    for col in range(self.cmp_quarter_table.columnCount()):
                        item = self.cmp_quarter_table.item(row, col)
                        row_data.append(item.text() if item else "")
                    data_quarter.append(row_data)
                headers_quarter = [self.cmp_quarter_table.horizontalHeaderItem(i).text() for i in range(self.cmp_quarter_table.columnCount())]
                df_quarter = pd.DataFrame(data_quarter, columns=headers_quarter)
                df_quarter.to_excel(writer, sheet_name="Сравнение по кварталам", index=False)

                # Можно добавить и другие вкладки, но для краткости оставим так.

            QMessageBox.information(self, "Экспорт", f"Отчёт сохранён в {fname}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл:\n{e}")


class FOTWidget(QWidget):
    def __init__(self, parent=None):  # добавлен parent
        super().__init__(parent)       # передача parent в базовый класс
        # ... остальной код конструктора

        self.project = Project()
        self.holidays = Holidays()
        self.current_file = None
        self.modified = False
        self.allowance_history: Set[str] = set()
        self.history_file = "allowance_history.json"
        self.employee_db = EmployeeDatabase()
        self.vacation_db = VacationDatabase()
        self.action_log = ActionLog()
        self.load_allowance_history()

        ##("Табель учёта рабочего времени и расчёт ФОТ")
        self.resize(1300, 700)



        layout = QVBoxLayout(self)

        # Панель выбора месяца (теперь два спиннера)
        month_selector_layout = QHBoxLayout()
        month_selector_layout.addWidget(QLabel("Год:"))
        self.year_spin = QSpinBox()
        self.year_spin.setRange(2000, 2100)
        self.year_spin.setValue(self.project.current_month[0])
        self.year_spin.valueChanged.connect(self.on_month_changed)
        month_selector_layout.addWidget(self.year_spin)

        month_selector_layout.addWidget(QLabel("Месяц:"))
        self.month_spin = QSpinBox()
        self.month_spin.setRange(1, 12)
        self.month_spin.setValue(self.project.current_month[1])
        self.month_spin.valueChanged.connect(self.on_month_changed)
        month_selector_layout.addWidget(self.month_spin)

        self.btn_add_month = QPushButton("Добавить месяц")
        self.btn_add_month.clicked.connect(self.add_month)
        month_selector_layout.addWidget(self.btn_add_month)

        month_selector_layout.addStretch()

        month_selector_layout.addWidget(QLabel("Норма дней:"))
        self.norm_spin = QSpinBox()
        self.norm_spin.setRange(1, 31)
        month_selector_layout.addWidget(self.norm_spin)

        self.btn_save_norm = QPushButton("Сохранить норму")
        self.btn_save_norm.clicked.connect(self.save_norm)
        month_selector_layout.addWidget(self.btn_save_norm)

        layout.addLayout(month_selector_layout)

        # Информационная метка
        info_layout = QHBoxLayout()
        self.info_label = QLabel()
        info_layout.addWidget(self.info_label)
        info_layout.addStretch()
        layout.addLayout(info_layout)

        # Таблица
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setShowGrid(True)
        self.table.setGridStyle(Qt.PenStyle.SolidLine)
        self.table.setStyleSheet("background-color: #e6e6e6; gridline-color: black;")
        self.table.horizontalHeader().setStyleSheet(
            "QHeaderView::section { background-color: #e6e6e6; color: black; font-weight: bold; border: 1px solid black; }"
        )
        self.table.verticalHeader().setStyleSheet(
            "QHeaderView::section { background-color: #e6e6e6; color: black; font-weight: bold; border: 1px solid black; }"
        )

        self.table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked)
        self.table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectItems)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.cellDoubleClicked.connect(self.on_cell_double_clicked)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        self.table.itemSelectionChanged.connect(self.update_selection_summary)
        layout.addWidget(self.table)


        self.status_label = QLabel("Готов")
        self.status_label.setStyleSheet("padding: 2px;")
        layout.addWidget(self.status_label)

        self.set_current_month_to_today()


        QTimer.singleShot(0, self.load_last_project)

    # --- Методы для работы с месяцем ---

    def on_month_changed(self):
        """Обрабатывает изменение года или месяца в спиннерах."""
        year = self.year_spin.value()
        month = self.month_spin.value()
        self.project.current_month = (year, month)
        self.update_table_structure()
        self.update_info_label()
        self.norm_spin.setValue(self.project.get_norm())

    def set_current_month_to_today(self):
        """Устанавливает текущий месяц как сегодняшний день, создаёт его при необходимости."""
        today = datetime.date.today()
        year, month = today.year, today.month
        if (year, month) not in self.project.months_data:
            self.project.months_data[(year, month)] = {}
            self.project.norm_by_month[(year, month)] = self.project.days_in_month(year, month)
        self.project.current_month = (year, month)
        # Устанавливаем значения спиннеров
        self.year_spin.setValue(year)
        self.month_spin.setValue(month)
        self.update_table_structure()
        self.update_info_label()
        self.norm_spin.setValue(self.project.get_norm())

    def add_month(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Добавить месяц")
        layout = QFormLayout(dlg)
        year_spin = QSpinBox()
        year_spin.setRange(2000, 2100)
        year_spin.setValue(self.project.current_month[0])
        month_spin = QSpinBox()
        month_spin.setRange(1, 12)
        month_spin.setValue(self.project.current_month[1])
        layout.addRow("Год:", year_spin)
        layout.addRow("Месяц:", month_spin)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dlg.accept)
        buttons.rejected.connect(dlg.reject)
        layout.addRow(buttons)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            year = year_spin.value()
            month = month_spin.value()
            if (year, month) not in self.project.months_data:
                self.project.months_data[(year, month)] = {}
                self.project.norm_by_month[(year, month)] = self.project.days_in_month(year, month)
                self.project.current_month = (year, month)
                self.year_spin.setValue(year)
                self.month_spin.setValue(month)
                self.update_table_structure()
                self.update_info_label()
                self.norm_spin.setValue(self.project.get_norm())
                self.action_log.add("MONTH_ADD", f"Добавлен месяц {month:02d}.{year}")
                self.set_modified()
            else:
                QMessageBox.information(self, "Добавление", "Такой месяц уже существует.")

    # Добавьте в класс MainWindow следующие методы:

    def show_quarter_planning(self):
        """Открывает диалог планирования на квартал."""
        dlg = QuarterPlanningDialog(self.project, self.holidays, self)
        dlg.exec()
    # Добавьте этот метод в класс MainWindow
    def update_sick_period_markers(self):
        """Обновляет звёздочки в полях часов для дней, начинающих новый период больничного."""
        data = self.project.get_current_data()
        emp_count = len(self.project.employees)
        days = self.project.days_in_month()
        for emp_idx in range(emp_count):
            prev_was_sick = False
            for day in range(1, days + 1):
                code, hours, option = data.get((emp_idx, day), ("", "", ""))
                is_sick = (code == 'Б' and option != 'child_care')
                is_new_period = is_sick and (not prev_was_sick or option == 'new_period')
                hours_item = self.table.item(emp_idx * 2 + 1, day - 1)
                if hours_item:
                    if is_new_period and hours_item.text() == "":
                        hours_item.setText("*")
                    else:
                        if hours_item.text() == "*" and not is_new_period:
                            hours_item.setText("")
                prev_was_sick = is_sick

    # --- Флаг изменений и заголовок ---
    def set_modified(self, value=True):
        if self.modified != value:
            self.modified = value



    # --- Загрузка последнего проекта ---
    def load_last_project(self):
        last_file = "last_project.txt"
        if os.path.exists(last_file):
            try:
                with open(last_file, "r", encoding="utf-8") as f:
                    fname = f.read().strip()
                if fname and os.path.exists(fname):
                    self.open_project_from_file(fname)
                    self.action_log.add("PROJECT_OPEN", f"Автозагрузка последнего проекта: {fname}")
            except Exception as e:
                print(f"Не удалось загрузить последний проект: {e}")



    def open_project_from_file(self, fname):
        try:
            with open(fname, "r", encoding="utf-8") as f:
                data = json.load(f)
            self.project = Project.from_dict(data)
            self.current_file = fname
            self.set_current_month_to_today()  # переключает на текущий месяц и обновляет интерфейс
            self.modified = False

            self.status_label.setText(f"Загружен {fname}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить проект:\n{e}")

    def set_current_month_to_today(self):
        today = datetime.date.today()
        year, month = today.year, today.month
        if (year, month) not in self.project.months_data:
            self.project.months_data[(year, month)] = {}
            self.project.norm_by_month[(year, month)] = self.project.days_in_month(year, month)
        self.project.current_month = (year, month)
        self.update_month_combo()
        self.update_table_structure()
        self.update_info_label()
        self.norm_spin.setValue(self.project.get_norm())

    # --- Сохранение нормы ---
    def save_norm(self):
        self.project.set_norm(self.norm_spin.value())
        self.status_label.setText(f"Норма сохранена: {self.norm_spin.value()} дней")
        self.set_modified()

    # --- Очистка графика отпусков ---
    def clear_all_vacations(self):
        reply = QMessageBox.question(self, "Очистка графика отпусков",
                                     "Вы уверены, что хотите удалить ВСЕ записи из графика отпусков?\n"
                                     "Это действие нельзя отменить.",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.project.vacations.clear()
            QMessageBox.information(self, "Очистка", "График отпусков полностью очищен.")
            self.set_modified()

    def clear_vacations_by_year(self):
        year, ok = QInputDialog.getInt(self, "Очистка графика отпусков", "Введите год:",
                                       value=self.project.current_month[0], min=2000, max=2100)
        if not ok:
            return
        to_remove = []
        for vac in self.project.vacations:
            if vac.start_date and vac.start_date.year == year:
                to_remove.append(vac)
            elif vac.end_date and vac.end_date.year == year:
                to_remove.append(vac)
            elif vac.start_date and vac.end_date and vac.start_date.year <= year <= vac.end_date.year:
                to_remove.append(vac)
        if not to_remove:
            QMessageBox.information(self, "Очистка", f"Записей за {year} год не найдено.")
            return
        reply = QMessageBox.question(self, "Подтверждение",
                                     f"Найдено {len(to_remove)} записей за {year} год.\n"
                                     "Удалить их?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            for vac in to_remove:
                self.project.vacations.remove(vac)
            QMessageBox.information(self, "Очистка", f"Удалено {len(to_remove)} записей за {year} год.")
            self.set_modified()

    def show_bonus_dialog(self):
        dlg = BonusDialog(self.project, self)
        dlg.exec()

    # --- История доплат ---
    def load_allowance_history(self):
        if os.path.exists(self.history_file):
            try:
                with open(self.history_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                self.allowance_history = set(data)
            except:
                self.allowance_history = set()

    def save_last_project(self, fname):
        try:
            with open("last_project.txt", "w", encoding="utf-8") as f:
                f.write(fname)
        except Exception as e:
            print(f"Не удалось сохранить последний проект: {e}")

    def save_allowance_history(self):
        try:
            with open(self.history_file, "w", encoding="utf-8") as f:
                json.dump(list(self.allowance_history), f, ensure_ascii=False, indent=2)
        except:
            pass

    # --- Меню ---
    def populate_menu(self, menu_bar):


        file_menu = menu_bar.addMenu("Файл")
        new_action = QAction("Новый проект", self)
        new_action.triggered.connect(self.new_project)
        file_menu.addAction(new_action)
        open_action = QAction("Открыть проект", self)
        open_action.triggered.connect(self.open_project)
        file_menu.addAction(open_action)
        save_action = QAction("Сохранить проект", self)
        save_action.triggered.connect(self.save_project)
        file_menu.addAction(save_action)
        save_as_action = QAction("Сохранить проект как", self)
        save_as_action.triggered.connect(self.save_project_as)
        file_menu.addAction(save_as_action)
        file_menu.addSeparator()


        edit_menu = menu_bar.addMenu("Правка")
        settings_action = QAction("Настройки проекта", self)
        settings_action.triggered.connect(self.show_settings)
        edit_menu.addAction(settings_action)

        code_mapping_action = QAction("Настройка кодов табеля", self)
        code_mapping_action.triggered.connect(self.show_code_mapping)
        edit_menu.addAction(code_mapping_action)

        db_menu = menu_bar.addMenu("Справочники")
        employees_action = QAction("Сотрудники", self)
        employees_action.triggered.connect(self.show_employees)
        db_menu.addAction(employees_action)
        edit_vac_action = QAction("Графики отпусков", self)
        edit_vac_action.triggered.connect(self.show_vacations)
        db_menu.addAction(edit_vac_action)



        analit_menu = menu_bar.addMenu("Аналитика")
        analytics_btn = QAction("Аналитика ФОТ", self)
        analytics_btn.triggered.connect(self.show_analytics)
        analit_menu.addAction(analytics_btn)
        quarter_plan_action = QAction("Планирование на квартал и аналиитка", self)
        quarter_plan_action.triggered.connect(self.show_quarter_planning)
        analit_menu.addAction(quarter_plan_action)

        action_menu = menu_bar.addMenu("Действия")
        generate_action = QAction("Сгенерировать табели", self)
        generate_action.triggered.connect(self.generate_tabels)
        action_menu.addAction(generate_action)
        calc_fot_action = QAction("Рассчитать ФОТ", self)
        calc_fot_action.triggered.connect(self.calc_fot)
        action_menu.addAction(calc_fot_action)

        bonus_btn = QAction("Расчёт премии ППП", self)
        bonus_btn.triggered.connect(self.show_bonus_dialog)
        action_menu.addAction(bonus_btn)
        calc_avg_action = QAction("Рассчитать среднюю зарплату", self)
        calc_avg_action.triggered.connect(self.calc_avg_salary)
        action_menu.addAction(calc_avg_action)


        log_menu = menu_bar.addMenu("Журнал")
        show_log_action = QAction("Просмотр действий", self)
        show_log_action.triggered.connect(self.show_log)
        log_menu.addAction(show_log_action)
        clear_log_action = QAction("Очистить журнал", self)
        clear_log_action.triggered.connect(self.clear_log)
        log_menu.addAction(clear_log_action)

        help_menu = menu_bar.addMenu("Справка")
        about_action = QAction("О программе", self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)
        instruction_action = QAction("Инструкция", self)
        instruction_action.triggered.connect(self.show_instruction)
        help_menu.addAction(instruction_action)



    
    def show_log(self):
        dlg = LogDialog(self.action_log, self)
        dlg.exec()

    def clear_log(self):
        reply = QMessageBox.question(self, "Очистка журнала", "Вы уверены, что хотите очистить журнал действий?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.action_log.clear()
            QMessageBox.information(self, "Журнал", "Журнал очищен.")

    # def create_toolbar(self):
    #     toolbar = QToolBar("Основные действия")
    #     self.addToolBar(toolbar)
    #     generate_action = QAction("Сгенерировать табели", self)
    #     generate_action.triggered.connect(self.generate_tabels)
    #     toolbar.addAction(generate_action)
    #     calc_fot_action = QAction("Рассчитать ФОТ", self)
    #     calc_fot_action.triggered.connect(self.calc_fot)
    #     toolbar.addAction(calc_fot_action)
    #     analytics_btn = QAction("Аналитика ФОТ", self)
    #     analytics_btn.triggered.connect(self.show_analytics)
    #     toolbar.addAction(analytics_btn)
    #     bonus_btn = QAction("Расчёт премии ППП...", self)
    #     bonus_btn.triggered.connect(self.show_bonus_dialog)
    #     toolbar.addAction(bonus_btn)
    #     quarter_plan_action = QAction("Планирование на квартал...", self)
    #     quarter_plan_action.triggered.connect(self.show_quarter_planning)
    #     toolbar.addAction(quarter_plan_action)

    # --- Обновление информации ---
    def update_selection_summary(self):
        selected = self.table.selectedIndexes()
        if not selected:
            self.status_label.setText("Готов")
            return
        total_cells = len(selected)
        numeric_values = []
        for idx in selected:
            item = self.table.item(idx.row(), idx.column())
            if item and item.text():
                try:
                    val = float(item.text())
                    numeric_values.append(val)
                except ValueError:
                    pass
        count_numeric = len(numeric_values)
        if count_numeric > 0:
            total_sum = sum(numeric_values)
            avg = total_sum / count_numeric
            msg = f"Количество ячеек: {total_cells} | Сумма: {total_sum:.2f} | Среднее: {avg:.2f}"
        else:
            msg = f"Количество ячеек: {total_cells}"
        self.status_label.setText(msg)

    def update_info_label(self):
        year, month = self.project.current_month
        norm = self.project.get_norm()
        text = f"{self.project.organization} / {self.project.department} — {month:02d}.{year}  (норма: {norm} дней)"
        self.info_label.setText(text)

    def update_month_combo(self):
        if not hasattr(self, 'month_combo') or self.month_combo is None:
            return
        self.month_combo.clear()
        months = sorted(self.project.months_data.keys())
        for year, month in months:
            self.month_combo.addItem(f"{month:02d}.{year}", (year, month))
        if self.project.current_month in months:
            index = months.index(self.project.current_month)
            self.month_combo.setCurrentIndex(index)

    def on_norm_changed(self, value):
        self.project.set_norm(value)

    def on_month_changed(self, value=None):
        year = self.year_spin.value()
        month = self.month_spin.value()
        self.project.current_month = (year, month)
        self.update_table_structure()
        self.update_info_label()
        self.norm_spin.setValue(self.project.get_norm())

    def add_month(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Добавить месяц")
        layout = QFormLayout(dlg)
        year_spin = QSpinBox()
        year_spin.setRange(2000, 2100)
        year_spin.setValue(self.project.current_month[0])
        month_spin = QSpinBox()
        month_spin.setRange(1, 12)
        month_spin.setValue(self.project.current_month[1])
        layout.addRow("Год:", year_spin)
        layout.addRow("Месяц:", month_spin)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dlg.accept)
        buttons.rejected.connect(dlg.reject)
        layout.addRow(buttons)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            year = year_spin.value()
            month = month_spin.value()
            if (year, month) not in self.project.months_data:
                self.project.months_data[(year, month)] = {}
                self.project.norm_by_month[(year, month)] = self.project.days_in_month(year, month)
                self.project.current_month = (year, month)
                self.update_month_combo()
                self.update_table_structure()
                self.update_info_label()
                self.norm_spin.setValue(self.project.get_norm())
                self.action_log.add("MONTH_ADD", f"Добавлен месяц {month:02d}.{year}")
                self.set_modified()
            else:
                QMessageBox.information(self, "Добавление", "Такой месяц уже существует.")

    def apply_weekends(self):
        self.project.apply_weekends_and_holidays(self.holidays)
        self.load_data_from_project()

    def update_table_structure(self):
        days = self.project.days_in_month()
        emp_count = len(self.project.employees)

        self.table.setRowCount(emp_count * 2)
        self.table.setColumnCount(days)

        headers = [str(i + 1) for i in range(days)]
        self.table.setHorizontalHeaderLabels(headers)

        for i, emp in enumerate(self.project.employees):
            item_code = QTableWidgetItem(emp.fio)
            item_code.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            self.table.setVerticalHeaderItem(i * 2, item_code)

            item_hours = QTableWidgetItem("часы")
            item_hours.setFlags(Qt.ItemFlag.ItemIsEnabled)
            self.table.setVerticalHeaderItem(i * 2 + 1, item_hours)

        self.load_data_from_project()
        self.apply_weekends()
        self.project.apply_vacation_schedule(self.holidays)
        self.load_data_from_project()

    # Обновлённый метод load_data_from_project
    def load_data_from_project(self):
        data = self.project.get_current_data()
        emp_count = len(self.project.employees)
        days = self.project.days_in_month()
        self.table.setStyleSheet("background-color: #e6e6e6;")

        # Заполняем ячейки кодами и часами (без звёздочек)
        for emp_idx in range(emp_count):
            for day in range(1, days + 1):
                code, hours, option = data.get((emp_idx, day), ("", "", ""))
                bg_color = self.get_color_for_code(code)

                item_code = QTableWidgetItem(code)
                font = QFont()
                font.setBold(True)
                item_code.setFont(font)
                item_code.setForeground(QBrush(QColor(0, 0, 0)))
                item_code.setBackground(bg_color)
                item_code.setFlags(
                    Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)  # добавлен ItemIsSelectable
                self.table.setItem(emp_idx * 2, day - 1, item_code)

                item_hours = QTableWidgetItem(hours)
                item_hours.setFont(font)
                item_hours.setForeground(QBrush(QColor(0, 0, 0)))
                item_hours.setBackground(bg_color)
                self.table.setItem(emp_idx * 2 + 1, day - 1, item_hours)

        # Проставляем звёздочки для периодов
        self.update_sick_period_markers()

    def save_data_to_project(self):
        data = {}
        emp_count = len(self.project.employees)
        days = self.project.days_in_month()
        for emp_idx in range(emp_count):
            for day in range(1, days+1):
                item_code = self.table.item(emp_idx*2, day-1)
                item_hours = self.table.item(emp_idx*2+1, day-1)
                code = item_code.text() if item_code else ""
                hours = item_hours.text() if item_hours else ""
                old_data = self.project.get_current_data().get((emp_idx, day), ("", "", "double"))
                _, _, old_option = old_data
                option = old_option if not code and not hours else "double"
                if code or hours:
                    data[(emp_idx, day)] = (code, hours, option)
        self.project.set_current_data(data)

    def get_color_for_code(self, code):
        colors = {
            'Ф': QColor(200, 255, 200),
            'В': QColor(220, 220, 220),
            'О': QColor(200, 220, 255),
            'ОД': QColor(200, 220, 255),
            'ОУ': QColor(200, 220, 255),
            'ОР': QColor(230, 200, 255),
            'К': QColor(255, 255, 200),
            'Б': QColor(255, 200, 200),
            'А': QColor(255, 220, 180),
            'НН': QColor(255, 220, 180),
            'ДО': QColor(255, 220, 180),
            'С': QColor(200, 230, 255),
            'РП': QColor(255, 200, 220),
            'КРВ': QColor(255, 255, 130),
        }
        return colors.get(code, QColor(230, 230, 230))

    def on_cell_double_clicked(self, row, col):
        print("Двойной клик, строка", row, "столбец", col)
        if row % 2 != 0:
            return
        emp_idx = row // 2
        day = col + 1
        data = self.project.get_current_data()
        code, hours, option = data.get((emp_idx, day), ("", "", "double"))
        dlg = CellEditDialog(self.project.code_mapping, code, hours, option, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            new_code, new_hours, new_option = dlg.get_values()
            self.action_log.add("CELL_EDITED", f"Ячейка ({row},{col}) изменена на {new_code}")
            if new_code not in HOURS_CODES:
                new_hours = ""
            self.table.setItem(row, col, QTableWidgetItem(new_code))
            self.table.setItem(row + 1, col, QTableWidgetItem(new_hours))
            bg_color = self.get_color_for_code(new_code)
            self.table.item(row, col).setBackground(bg_color)
            self.table.item(row + 1, col).setBackground(bg_color)
            for r in (row, row + 1):
                item = self.table.item(r, col)
                if item:
                    font = QFont()
                    font.setBold(True)
                    item.setFont(font)
                    item.setForeground(QBrush(QColor(0, 0, 0)))
            data[(emp_idx, day)] = (new_code, new_hours, new_option)
            self.project.set_current_data(data)
            self.update_sick_period_markers()  # обновить звёздочки
            self.set_modified()

    def new_project(self):
        if self.maybe_save():
            self.project = Project()
            self.current_file = None
            self.set_current_month_to_today()  # всё обновит
            self.status_label.setText("Новый проект создан")
            self.action_log.add("PROJECT_NEW", "Создан новый проект")

    def open_project(self):
        if self.maybe_save():
            fname, _ = QFileDialog.getOpenFileName(self, "Открыть проект", "",
                                                   "Tabel Project (*.tabelproj);;All Files (*)")
            if fname:
                try:
                    with open(fname, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    self.project = Project.from_dict(data)
                    self.current_file = fname
                    self.set_current_month_to_today()
                    self.modified = False

                    self.status_label.setText(f"Загружен {fname}")
                    self.action_log.add("PROJECT_OPEN", f"Открыт проект: {fname}")
                except Exception as e:
                    QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить проект:\n{e}")


    def load_project_from_file(self, fname):
        # используется только для автозагрузки, но оставим для совместимости
        return self.open_project_from_file(fname)

    def save_project(self):
        if self.current_file:
            self.save_project_to_file(self.current_file)
        else:
            self.save_project_as()

    def save_project_to_file(self, fname):
        try:
            #self.save_data_to_project()
            with open(fname, "w", encoding="utf-8") as f:
                json.dump(self.project.to_dict(), f, ensure_ascii=False, indent=2)
            self.status_label.setText(f"Сохранено в {fname}")
            self.action_log.add("PROJECT_SAVE", f"Сохранён проект: {fname}")
            self.save_last_project(fname)
            self.set_modified(False)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить проект:\n{e}")

    def save_project_as(self):
        fname, _ = QFileDialog.getSaveFileName(self, "Сохранить проект", "", "Tabel Project (*.tabelproj)")
        if fname:
            if not fname.endswith(".tabelproj"):
                fname += ".tabelproj"
            self.save_project_to_file(fname)
            self.current_file = fname
            self.action_log.add("PROJECT_SAVE_AS", f"Сохранён проект как: {fname}")
            self.save_last_project(fname)

    def maybe_save(self):
        if not self.modified:
            return True
        reply = QMessageBox.question(self, "Сохранение", "Сохранить изменения в проекте?",
                                      QMessageBox.StandardButton.Yes |
                                      QMessageBox.StandardButton.No |
                                      QMessageBox.StandardButton.Cancel)
        if reply == QMessageBox.StandardButton.Yes:
            self.save_project()
            return True
        elif reply == QMessageBox.StandardButton.No:
            return True
        else:
            return False

    def show_settings(self):
        dlg = SettingsDialog(self.project, self.holidays, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.update_table_structure()
            self.update_info_label()
            self.set_modified()

    def show_employees(self):
        dlg = EmployeeDialog(self.project.employees, self.allowance_history, self.employee_db, self.project, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.update_table_structure()
            self.save_allowance_history()
            self.action_log.add("EMPLOYEES_EDITED", "Изменён список сотрудников")
            self.set_modified()

    def show_vacations(self):
        dlg = VacationDialog(self.project, self.holidays, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.project.apply_vacation_schedule(self.holidays)
            self.load_data_from_project()
            self.action_log.add("VACATIONS_EDITED", "Изменён график отпусков")
            self.set_modified()

    def show_code_mapping(self):
        dlg = CodeSettingsDialog(self.project.code_mapping, self)
        dlg.exec()

    def mass_fill(self):
        selected = self.table.selectedIndexes()
        if not selected:
            QMessageBox.information(self, "Редактирование", "Сначала выделите ячейки для заполнения.")
            return
        code_indexes = [idx for idx in selected if idx.row() % 2 == 0]
        if not code_indexes:
            QMessageBox.information(self, "Редактирование", "Выделите ячейки в строках кодов (верхние строки).")
            return
        dlg = MassFillDialog(self.project.code_mapping, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            new_code, new_hours, new_option = dlg.get_values()
            self.action_log.add("MASS_FILL", f"Массовое редактирование: {len(code_indexes)} ячеек кодом {new_code}")
            if not new_code:
                return
            if new_code not in HOURS_CODES:
                new_hours = ""
            data = self.project.get_current_data()
            bg_color = self.get_color_for_code(new_code)
            for idx in code_indexes:
                row, col = idx.row(), idx.column()
                self.table.setItem(row, col, QTableWidgetItem(new_code))
                hours_row = row + 1
                if hours_row < self.table.rowCount():
                    self.table.setItem(hours_row, col, QTableWidgetItem(new_hours))
                self.table.item(row, col).setBackground(bg_color)
                if hours_row < self.table.rowCount():
                    self.table.item(hours_row, col).setBackground(bg_color)
                for r in (row, row + 1):
                    item = self.table.item(r, col)
                    if item:
                        font = QFont()
                        font.setBold(True)
                        item.setFont(font)
                        item.setForeground(QBrush(QColor(0, 0, 0)))
                emp_idx = row // 2
                day = col + 1
                data[(emp_idx, day)] = (new_code, new_hours, new_option)
            self.project.set_current_data(data)
            self.update_sick_period_markers()  # обновить звёздочки
            self.status_label.setText(f"Заполнено {len(code_indexes)} ячеек кодом {new_code}")
            self.set_modified()

    def show_context_menu(self, pos):
        print("Контекстное меню вызвано")
        menu = QMenu()
        fill_action = menu.addAction("Редактирование...")
        action = menu.exec(self.table.viewport().mapToGlobal(pos))
        if action == fill_action:
            self.mass_fill()

    def generate_tabels(self):
        #self.save_data_to_project()
        template_first = "template_first.xlsx"
        template_second = "template_second.xlsx"
        if not os.path.exists(template_first) or not os.path.exists(template_second):
            QMessageBox.critical(self, "Ошибка", "Файлы шаблонов не найдены в папке программы.")
            return
        out_dir = QFileDialog.getExistingDirectory(self, "Выберите папку для сохранения табелей")
        if not out_dir:
            return
        try:
            year, month = self.project.current_month
            days = self.project.days_in_month()
            self.generate_half(template_first, out_dir, 1, min(15, days), f"{month:02d}.{year}_первая_половина", target_col=55)
            if days > 15:
                self.generate_half(template_second, out_dir, 16, days, f"{month:02d}.{year}_вторая_половина", target_col=93)
            QMessageBox.information(self, "Готово", f"Табели для {month:02d}.{year} успешно сгенерированы в папке:\n{out_dir}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка генерации", str(e))

    def generate_half(self, template_path, out_dir, start_day, end_day, suffix, target_col):
        wb = openpyxl.load_workbook(template_path)
        if "TDSheet" not in wb.sheetnames or "TDSheet (2)" not in wb.sheetnames:
            raise Exception("Шаблон не содержит необходимых листов (TDSheet, TDSheet (2))")
        sheet_header = wb["TDSheet"]
        sheet_data = wb["TDSheet (2)"]
        self.fill_header(sheet_header, start_day, end_day)
        self.fill_employees(sheet_data, start_day, end_day, target_col)
        out_fname = f"Табель_{suffix}.xlsx"
        out_path = os.path.join(out_dir, out_fname)
        wb.save(out_path)

    def fill_header(self, sheet, start_day, end_day):
        try:
            year, month = self.project.current_month
            sheet["J9"] = self.project.organization
            sheet["J10"] = self.project.department
            sheet["J11"] = "первичный"
            last_date = datetime.date(year, month, end_day)
            sheet["CR12"] = last_date.isoformat()
            period_text = f"за период с {start_day:02d} по {end_day:02d} {self.month_name(month)} {year} г."
            sheet["AL8"] = period_text
            sheet["BZ9"] = self.project.okpo
        except Exception as e:
            print(f"Предупреждение при заполнении шапки: {e}")

    def month_name(self, month):
        months = [
            "января", "февраля", "марта", "апреля", "мая", "июня",
            "июля", "августа", "сентября", "октября", "ноября", "декабря"
        ]
        return months[month - 1]

    def fill_employees(self, sheet, start_day, end_day, target_col):
        start_row = 10
        emp_count = len(self.project.employees)
        data = self.project.get_current_data()
        day_to_col = {}
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=3, column=col)
            if cell.value is not None and str(cell.value).isdigit():
                day = int(cell.value)
                day_to_col[day] = col
        cols_in_range = [day_to_col[day] for day in range(start_day, end_day + 1) if day in day_to_col]
        if cols_in_range:
            min_col = min(cols_in_range)
            max_col = max(cols_in_range)
        else:
            min_col = max_col = None
        for emp_idx, emp in enumerate(self.project.employees):
            row_code = start_row + emp_idx * 2
            row_hours = row_code + 1
            try:
                sheet.cell(row=row_code, column=1, value=emp.fio)
                sheet.cell(row=row_code, column=11, value=emp.tab_num)
                sheet.cell(row=row_code, column=17, value=emp.position)
            except Exception as e:
                print(f"Предупреждение при записи данных сотрудника {emp.fio}: {e}")
            for day in range(start_day, end_day + 1):
                if day not in day_to_col:
                    continue
                col = day_to_col[day]
                code, hours, _ = data.get((emp_idx, day), ("", "", ""))
                try:
                    sheet.cell(row=row_code, column=col, value=code)
                    sheet.cell(row=row_hours, column=col, value=hours)
                except Exception as e:
                    print(f"Предупреждение при записи дня {day} для {emp.fio}: {e}")
            if min_col and max_col:
                formula = f"=COUNTA({get_column_letter(min_col)}{row_hours}:{get_column_letter(max_col)}{row_hours})"
                try:
                    sheet.cell(row=row_code, column=target_col, value=formula)
                except Exception as e:
                    print(f"Предупреждение при записи формулы для {emp.fio}: {e}")

    def calc_fot(self):
        #self.save_data_to_project()
        dlg = SummaryDialog(self.project, self.holidays, self)
        dlg.exec()
        self.action_log.add("FOT_CALCULATED", "Выполнен расчёт ФОТ")

    def calc_avg_salary(self):
        dlg = AvgSalaryDialog(self.project, self.holidays, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.update_table_structure()

    def show_analytics(self):
        dlg = AnalyticsDialog(self.project, self.holidays, self)
        dlg.exec()

    def save_current_to_db(self):
        for emp in self.project.employees:
            self.employee_db.add_employee(emp)
        self.employee_db.save()
        QMessageBox.information(self, "База сотрудников", f"{len(self.project.employees)} сотрудников сохранено в базу.")

    def load_from_db(self):
        if not self.employee_db.employees:
            QMessageBox.information(self, "База сотрудников", "База сотрудников пуста.")
            return
        updated = 0
        added = 0
        existing_by_tab = {e.tab_num: e for e in self.project.employees if e.tab_num}
        for db_emp in self.employee_db.employees:
            if db_emp.tab_num and db_emp.tab_num in existing_by_tab:
                idx = self.project.employees.index(existing_by_tab[db_emp.tab_num])
                self.project.employees[idx] = db_emp
                updated += 1
            else:
                self.project.employees.append(db_emp)
                added += 1
        self.update_table_structure()
        QMessageBox.information(self, "База сотрудников", f"Обновлено {updated} сотрудников, добавлено {added}.")
        self.set_modified()

    def clear_db(self):
        reply = QMessageBox.question(self, "Очистка базы", "Вы уверены, что хотите очистить базу сотрудников?",
                                      QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.employee_db.clear()
            self.employee_db.save()
            QMessageBox.information(self, "База сотрудников", "База очищена.")
            self.set_modified()

    def save_vacations_to_file(self):
        fname, _ = QFileDialog.getSaveFileName(self, "Сохранить график отпусков", "", "JSON Files (*.json)")
        if fname:
            if not fname.endswith(".json"):
                fname += ".json"
            try:
                data = [v.to_dict() for v in self.project.vacations]
                with open(fname, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                QMessageBox.information(self, "Сохранение", "График отпусков сохранён.")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))

    def load_vacations_from_file(self):
        fname, _ = QFileDialog.getOpenFileName(self, "Загрузить график отпусков", "", "JSON Files (*.json)")
        if fname:
            try:
                with open(fname, "r", encoding="utf-8") as f:
                    data = json.load(f)
                msg_box = QMessageBox(self)
                msg_box.setWindowTitle("Загрузка")
                msg_box.setText("Заменить текущий график или добавить записи?")
                btn_replace = msg_box.addButton("Заменить", QMessageBox.ButtonRole.YesRole)
                btn_add = msg_box.addButton("Добавить", QMessageBox.ButtonRole.NoRole)
                btn_cancel = msg_box.addButton("Отмена", QMessageBox.ButtonRole.RejectRole)
                msg_box.exec()
                clicked = msg_box.clickedButton()
                if clicked == btn_cancel:
                    return
                new_vacations = [VacationRecord.from_dict(d) for d in data]
                if clicked == btn_replace:
                    self.project.vacations = new_vacations
                else:
                    self.project.vacations.extend(new_vacations)
                self.project.apply_vacation_schedule(self.holidays)
                self.load_data_from_project()
                QMessageBox.information(self, "Загрузка", f"Загружено {len(new_vacations)} записей.")
                self.set_modified()
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))

    def save_vacations_to_db(self):
        if not self.project.vacations:
            QMessageBox.information(self, "Сохранение", "Нет записей для сохранения.")
            return
        reply = QMessageBox.question(self, "Сохранение в базу",
                                     "Заменить текущие записи в базе или добавить?",
                                     "Заменить", "Добавить", "Отмена")
        if reply == 2:
            return
        if reply == 0:  # Заменить
            self.vacation_db.vacations = self.project.vacations.copy()
        else:  # Добавить
            for v in self.project.vacations:
                self.vacation_db.add_vacation(v)
        self.vacation_db.save()
        QMessageBox.information(self, "Сохранение", f"Сохранено {len(self.project.vacations)} записей в базу.")

    def load_vacations_from_db(self):
        if not self.vacation_db.vacations:
            QMessageBox.information(self, "Загрузка", "База отпусков пуста.")
            return
        reply = QMessageBox.question(self, "Загрузка из базы",
                                     "Заменить текущий график или добавить записи?",
                                     "Заменить", "Добавить", "Отмена")
        if reply == 2:
            return
        if reply == 0:
            self.project.vacations = self.vacation_db.vacations.copy()
        else:
            self.project.vacations.extend(self.vacation_db.vacations)
        self.project.apply_vacation_schedule(self.holidays)
        self.load_data_from_project()
        QMessageBox.information(self, "Загрузка", f"Загружено {len(self.vacation_db.vacations)} записей.")
        self.set_modified()

    def show_about(self):
        QMessageBox.about(self, "О программе",
                          "Экономика и ФОТ\nВерсия 8.1\n\n"
                          "Программа для ведения табеля учёта рабочего времени, планирования и расчёта заработной платы.\n"
                          "Алгоритм расчёта:\n"
                          "- Оклад: оклад / норма * факт. дни (код 'Ф')\n"
                          "- Отпускные: средняя * дни отпуска (кроме 'ОР'), исключая праздники\n"
                          "- Командировки: средняя * дни командировок, исключая праздники\n"
                          "- Больничные: средняя * (первые 3 дня каждого периода) с учётом стажа\n"
                          "- Доплаты: фиксированные или пропорционально отработанному времени\n"
                          "- Специальные коды (С, РП, КРВ): двойная оплата или отгул (настраивается для каждой ячейки)\n\n"
                          "Возможности:\n"
                          "- Работа с несколькими месяцами в одном проекте, норма дней сохраняется для каждого месяца\n"
                          "- Планирование на квартал с сохранением планов в проекте и возможностью загрузки\n"
                          "- График отпусков с автоматической простановкой в табель (с учётом праздников)\n"
                          "- Автоматический расчёт средней зарплаты по истории начислений\n"
                          "- Детальная аналитика ФОТ с разделением на фактические и плановые показатели:\n"
                          "  * Динамика по месяцам (линейный, столбчатый, накопительный графики)\n"
                          "  * Структура начислений за выбранный месяц (круговая или столбчатая диаграмма)\n"
                          "  * Поквартальная сводка (таблица и групповые столбцы)\n"
                          "  * Детализация по месяцам внутри квартала\n"
                          "- Сравнение плана и факта по месяцам и кварталам с таблицами отклонений и графиками\n"
                          "- Расчёт премии ППП с гибкими настройками и экспортом в Excel\n"
                          "- Экспорт всех отчётов и аналитики в Excel с полной детализацией (несколько листов)\n"
                          "- Учёт переносов праздничных дней (производственный календарь)\n\n"
                          "Разработано по заказу ФБУ «Пермский ЦСМ».")

    def show_instruction(self):
        instruction = (
            "ИНСТРУКЦИЯ ПО ИСПОЛЬЗОВАНИЮ\n\n"
            "1. Запустите программу.\n"
            "2. Создайте новый проект (Файл → Новый проект) или откройте существующий.\n"
            "3. В меню «Правка» → «Настройки проекта» укажите учреждение, подразделение и другие данные.\n"
            "4. В меню «Правка» → «Сотрудники» добавьте сотрудников, указав оклад, стаж, среднюю зарплату и список доплат.\n"
            "   При вводе названий доплат программа запоминает их и предлагает в дальнейшем (автодополнение).\n"
            "   В этом же окне можно сохранить сотрудников в базу или загрузить из базы с обновлением.\n"
            "5. В меню «График отпусков» создайте записи об отпусках сотрудников.\n"
            "   График можно сохранять и загружать в формате JSON, а также в базу программы.\n"
            "   При указании дат программа автоматически учитывает праздничные дни и продлевает отпуск.\n"
            "6. В меню «Правка» → «Настройка кодов табеля» задайте соответствие кодов категориям и настройте специальные коды.\n"
            "7. На панели инструментов выберите месяц, с которым хотите работать. Кнопка «Добавить месяц» позволяет добавить новый месяц.\n"
            "   Рядом с выбором месяца находится поле «Норма дней» – его можно редактировать для каждого месяца отдельно (кнопка «Сохранить норму»).\n"
            "8. В основной таблице для каждого дня введите код (двойной клик по ячейке открывает диалог с выбором опции для специальных кодов).\n"
            "   Выходные и праздничные дни проставляются автоматически (код «В»), но только если код не был задан вручную.\n"
            "   Отпуска из графика проставляются автоматически при открытии месяца.\n"
            "9. Для редактирования нескольких ячеек выделите их и выберите «Действия» → «Редактирование» (Ctrl+M).\n"
            "10. При выделении ячеек внизу окна отображается количество, сумма и среднее выделенных чисел.\n"
            "11. После заполнения табеля выберите «Действия» → «Рассчитать ФОТ» для просмотра итоговой ведомости за текущий месяц.\n"
            "12. Для автоматического расчёта средней зарплаты по истории используйте «Действия» → «Рассчитать среднюю зарплату».\n"
            "13. Планирование на квартал: выберите «Действия» → «Планирование на квартал». В открывшемся окне можно:\n"
            "    - Заполнить табель на три месяца вперёд (вручную или автоматически по выходным/отпускам)\n"
            "    - Сохранить план в проекте или загрузить ранее сохранённый\n"
            "    - Рассчитать плановые показатели и увидеть итоги по месяцам и за квартал\n"
            "    - Экспортировать план в Excel\n"
            "14. Для анализа ФОТ выберите «Действия» → «Аналитика ФОТ». Откроется окно с тремя основными разделами:\n"
            "    - **Факт** – аналитика по фактическим данным (аналогично предыдущим версиям).\n"
            "    - **План** – те же отчёты, но построенные на основе сохранённых планов.\n"
            "    - **Сравнение по месяцам** – таблица и график план/факт для выбранного года и показателя.\n"
            "    - **Сравнение по кварталам** – аналогичное сравнение, но сгруппированное по кварталам.\n"
            "    В каждом разделе доступны динамика, структура, поквартальные итоги и детализация.\n"
            "15. Для расчёта премии ППП выберите «Действия» → «Расчёт премии ППП». В диалоге можно задать фонд премии, процент, исключить руководителей и водителей, округление.\n"
            "16. Для экспорта ведомостей, отчётов и сравнений используйте кнопки «Экспорт в Excel» в соответствующих диалогах. В аналитике также доступна общая кнопка «Сохранить отчёт в Excel», которая формирует файл с несколькими листами, включая всю статистику.\n"
            "17. Для формирования официальных табелей (форма 0504421) выберите «Действия» → «Сгенерировать табели».\n"
            "    Необходимые шаблоны (template_first.xlsx, template_second.xlsx) должны лежать в папке с программой.\n"
            "18. Для очистки данных за год используйте «Действия» → «Очистить табель за год...». Для очистки графика отпусков – пункты в меню «График отпусков».\n\n"
            "Все данные сохраняются в проекте (формат .tabelproj). При выходе программа предложит сохранить изменения."
        )
        dlg = QDialog(self)
        dlg.setWindowTitle("Инструкция")
        dlg.resize(900, 600)
        layout = QVBoxLayout(dlg)
        text_edit = QTextEdit()
        text_edit.setReadOnly(True)
        text_edit.setPlainText(instruction)
        layout.addWidget(text_edit)
        btn_close = QPushButton("Закрыть")
        btn_close.clicked.connect(dlg.accept)
        layout.addWidget(btn_close, alignment=Qt.AlignmentFlag.AlignCenter)
        dlg.exec()



#if __name__ == "__main__":
 #   app = QApplication(sys.argv)
  #  app.setApplicationName("TabelAutoFOT")
   # app.setOrganizationName("MyOrg")

    # Устанавливаем иконку приложения
    #icon_path = "icon.ico"  # или "icon.png"
    #if os.path.exists(icon_path):
     #   app.setWindowIcon(QIcon(icon_path))

    #window = FOTMainWindow
    # Если хотите, можно также установить иконку для окна отдельно (обычно наследуется от приложения)
   # if os.path.exists(icon_path):
    #    window.setWindowIcon(QIcon(icon_path))

  #  window.show()
   # sys.exit(app.exec())