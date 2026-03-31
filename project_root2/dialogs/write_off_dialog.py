# dialogs/write_off_dialog.py

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QMessageBox, QFileDialog, QLineEdit
)
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os
from database.repair_db import get_all_materials, get_material_by_id

class WriteOffDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Акт списания материалов")
        self.resize(800, 500)
        self.materials_data = []  # список выбранных материалов

        layout = QVBoxLayout(self)

        # Поиск
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Поиск по наименованию...")
        self.search_edit.textChanged.connect(self.filter_table)
        layout.addWidget(self.search_edit)

        # Таблица материалов
        self.table = QTableWidget(0, 6)
        self.table.setHorizontalHeaderLabels(["Выбрать", "ID", "Наименование", "Инв. №", "Ед.", "Остаток"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.hideColumn(1)
        layout.addWidget(self.table)

        # Кнопки
        btn_layout = QHBoxLayout()
        self.btn_select_all = QPushButton("Выбрать всех")
        self.btn_clear_all = QPushButton("Снять выделение")
        self.btn_export = QPushButton("Сформировать акт")
        self.btn_cancel = QPushButton("Отмена")
        btn_layout.addWidget(self.btn_select_all)
        btn_layout.addWidget(self.btn_clear_all)
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_export)
        btn_layout.addWidget(self.btn_cancel)
        layout.addLayout(btn_layout)

        self.btn_select_all.clicked.connect(self.select_all)
        self.btn_clear_all.clicked.connect(self.clear_all)
        self.btn_export.clicked.connect(self.export_act)
        self.btn_cancel.clicked.connect(self.reject)

        self.load_materials()

    def load_materials(self):
        materials = get_all_materials()
        self.table.setRowCount(len(materials))
        for i, m in enumerate(materials):
            chk = QTableWidgetItem()
            chk.setCheckState(Qt.Unchecked)
            self.table.setItem(i, 0, chk)
            self.table.setItem(i, 1, QTableWidgetItem(str(m.id)))
            self.table.setItem(i, 2, QTableWidgetItem(m.name))
            self.table.setItem(i, 3, QTableWidgetItem(m.inventory_number))
            self.table.setItem(i, 4, QTableWidgetItem(m.unit))
            self.table.setItem(i, 5, QTableWidgetItem(f"{m.stock:.2f}"))

    def filter_table(self):
        text = self.search_edit.text().strip().lower()
        for row in range(self.table.rowCount()):
            name = self.table.item(row, 2).text().lower()
            visible = text in name
            self.table.setRowHidden(row, not visible)

    def select_all(self):
        for row in range(self.table.rowCount()):
            if not self.table.isRowHidden(row):
                self.table.item(row, 0).setCheckState(Qt.Checked)

    def clear_all(self):
        for row in range(self.table.rowCount()):
            self.table.item(row, 0).setCheckState(Qt.Unchecked)

    def export_act(self):
        # Собираем выбранные материалы и запрашиваем количество
        selected = []
        for row in range(self.table.rowCount()):
            if not self.table.isRowHidden(row) and self.table.item(row, 0).checkState() == Qt.Checked:
                mat_id = int(self.table.item(row, 1).text())
                mat = get_material_by_id(mat_id)
                if mat:
                    # Запрашиваем количество (не более остатка)
                    from PySide6.QtWidgets import QInputDialog
                    qty, ok = QInputDialog.getDouble(self, "Количество",
                                                     f"Сколько списать материала '{mat.name}'?",
                                                     1.0,  # значение по умолчанию
                                                     0.01,  # минимум
                                                     mat.stock,  # максимум
                                                     3)  # количество знаков после запятой
                    if ok and qty > 0:
                        selected.append((mat, qty))
        if not selected:
            QMessageBox.warning(self, "Внимание", "Не выбрано ни одного материала для списания")
            return

        # Формируем имя файла
        fname, _ = QFileDialog.getSaveFileName(self, "Сохранить акт списания",
                                               "Акт_списания.xlsx",
                                               "Excel Files (*.xlsx)")
        if not fname:
            return
        if not fname.endswith('.xlsx'):
            fname += '.xlsx'

        # Создаём Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Акт списания"

        # Заголовок
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = "Прошу списать следующие ТМЦ, использованные в"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')

        # Шапка
        headers = ['№ п/п', 'Наименование', 'Инвентарный номер', 'Кол-во', 'Причина списания']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

        # Данные
        for i, (mat, qty) in enumerate(selected, 1):
            ws.cell(row=3 + i, column=1, value=i).alignment = Alignment(horizontal='center')
            ws.cell(row=3 + i, column=2, value=mat.name)
            ws.cell(row=3 + i, column=3, value=mat.inventory_number).alignment = Alignment(horizontal='center')
            ws.cell(row=3 + i, column=4, value=f"{qty} {mat.unit}").alignment = Alignment(horizontal='center')
            ws.cell(row=3 + i, column=5, value="Израсходовано").alignment = Alignment(horizontal='center')

        # Автоширина
        for col in range(1, 6):
            ws.column_dimensions[chr(64+col)].width = max(15, len(headers[col-1]) + 2)

        wb.save(fname)
        QMessageBox.information(self, "Экспорт", f"Акт списания сохранён:\n{fname}")
        self.accept()