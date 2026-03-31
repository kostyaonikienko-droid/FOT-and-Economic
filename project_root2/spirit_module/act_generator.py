import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict
from .db_manager import get_setting, get_spirit_norm
from .grouping import GROUPS

THIN_BORDER = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

def format_cell(cell, font_size=10, bold=False, alignment='center', wrap_text=False):
    cell.font = Font(size=font_size, bold=bold)
    cell.alignment = Alignment(horizontal=alignment, vertical='center', wrap_text=wrap_text)
    cell.border = THIN_BORDER

def generate_act(works_by_group: Dict[str, List[Dict]], period_start: str, period_end: str, output_path: str):
    wb = openpyxl.Workbook()
    ws_act = wb.active
    ws_act.title = "Акт"

    # Шапка
    ws_act.merge_cells('G1:I1')
    ws_act['G1'] = "Утверждаю"
    format_cell(ws_act['G1'], bold=True, alignment='right')

    ws_act.merge_cells('E2:F2')
    ws_act['E2'] = "И.о. директора  ФБУ \"Пермский ЦСМ\""
    format_cell(ws_act['E2'], alignment='right')
    ws_act.merge_cells('G3:I3')
    ws_act['G3'] = "В.А.Трусов"
    format_cell(ws_act['G3'], bold=True, alignment='right')

    act_number = get_setting("act_number") or ""
    act_date = get_setting("act_date") or ""
    ws_act['A4'] = f"АКТ № {act_number}"
    format_cell(ws_act['A4'], bold=True, alignment='left')
    ws_act['G4'] = act_date
    format_cell(ws_act['G4'], alignment='right')

    ws_act.merge_cells('A5:I5')
    ws_act['A5'] = "Настоящий акт составлен комиссией в составе:"
    format_cell(ws_act['A5'], alignment='left')

    chairman = get_setting("commission_chairman") or ""
    ws_act['A6'] = f"Председатель - {chairman}"
    format_cell(ws_act['A6'], alignment='left')

    members = get_setting("commission_members") or ""
    ws_act.merge_cells('A7:I7')
    ws_act['A7'] = f"Члены комиссии: {members}"
    format_cell(ws_act['A7'], alignment='left', wrap_text=True)

    ws_act.merge_cells('A9:I9')
    ws_act['A9'] = "в том, что проведена проверка расхода антисептического раствора в отделе физико-химических измерений."
    format_cell(ws_act['A9'], alignment='left')

    antiseptic_info = get_setting("antiseptic_solution_ticket") or ""
    inv_number = get_setting("antiseptic_inv_number") or ""
    ws_act.merge_cells('A10:G10')
    ws_act['A10'] = f"Раствор антисептический получен по тр. {antiseptic_info}"
    format_cell(ws_act['A10'], alignment='left')
    ws_act['H10'] = "Инв. №"
    ws_act['I10'] = inv_number
    format_cell(ws_act['H10'], bold=True)
    format_cell(ws_act['I10'])

    ws_act['A11'] = f"с {period_start} по {period_end}"
    format_cell(ws_act['A11'], alignment='left')

    current_row = 13
    for group_full_name, works in works_by_group.items():
        if not works:
            continue
        ws_act.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
        cell = ws_act.cell(row=current_row, column=1)
        cell.value = group_full_name
        format_cell(cell, font_size=10, bold=True, alignment='left')
        current_row += 1

        headers_row = current_row
        for col_group in range(3):
            start_col = col_group * 3 + 1
            ws_act.cell(row=headers_row, column=start_col, value="Дата")
            ws_act.cell(row=headers_row, column=start_col+1, value="№ счета")
            ws_act.cell(row=headers_row, column=start_col+2, value="к-во")
            for c in range(start_col, start_col+3):
                format_cell(ws_act.cell(row=headers_row, column=c), bold=True)
        current_row += 1

        columns_data = [[], [], []]
        for i, w in enumerate(works):
            col_idx = i % 3
            columns_data[col_idx].append((
                w['date_work'],
                w['invoice_num'],
                w.get('quantity', 1)
            ))

        max_rows = max(len(col) for col in columns_data)
        for row_offset in range(max_rows):
            for col_idx in range(3):
                if row_offset < len(columns_data[col_idx]):
                    date_val, inv_val, qty_val = columns_data[col_idx][row_offset]
                    base_col = col_idx * 3 + 1
                    cell = ws_act.cell(row=current_row + row_offset, column=base_col)
                    cell.value = date_val
                    format_cell(cell, alignment='center')
                    cell = ws_act.cell(row=current_row + row_offset, column=base_col+1)
                    cell.value = inv_val
                    format_cell(cell, alignment='center')
                    cell = ws_act.cell(row=current_row + row_offset, column=base_col+2)
                    cell.value = qty_val
                    format_cell(cell, alignment='center')
        current_row += max_rows + 1

        total_quantity = sum(w.get('quantity', 1) for w in works)
        group_key = None
        for key, full in GROUPS.items():
            if full == group_full_name:
                group_key = key
                break
        norm = get_spirit_norm(group_key) if group_key else 0
        spisano = total_quantity * norm
        total_line = f"Итого: {total_quantity} шт. расход спирта - {norm:.3f} л/шт Списано (литров) — {spisano:.2f}"
        ws_act.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
        cell = ws_act.cell(row=current_row, column=1)
        cell.value = total_line
        format_cell(cell, bold=True, alignment='left')
        current_row += 2

    total_spirit = 0
    for group_full_name, works in works_by_group.items():
        if not works:
            continue
        group_key = None
        for key, full in GROUPS.items():
            if full == group_full_name:
                group_key = key
                break
        if group_key:
            norm = get_spirit_norm(group_key)
            total_quantity = sum(w.get('quantity', 1) for w in works)
            total_spirit += total_quantity * norm
    ws_act.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    cell = ws_act.cell(row=current_row, column=1)
    cell.value = f"Всего списано спирта — {total_spirit:.2f} литров"
    format_cell(cell, bold=True, alignment='left')
    current_row += 2

    ws_act.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
    ws_act.cell(row=current_row, column=1).value = "Вышеуказанный расход спирта в количестве"
    format_cell(ws_act.cell(row=current_row, column=1), alignment='left')
    current_row += 1
    ws_act.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    ws_act.cell(row=current_row, column=1).value = f"{total_spirit:.2f}"
    format_cell(ws_act.cell(row=current_row, column=1), bold=True, alignment='center')
    ws_act.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=9)
    ws_act.cell(row=current_row, column=4).value = "литра соответствует"
    format_cell(ws_act.cell(row=current_row, column=4), alignment='left')
    current_row += 1
    ws_act.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
    ws_act.cell(row=current_row, column=1).value = "\"Нормам расхода этилового спирта на поверку средств измерений и техническое обслуживание эталонов и поверочного оборудования \", утвержденным директором ФБУ \"Пермский ЦСМ\""
    format_cell(ws_act.cell(row=current_row, column=1), alignment='left', wrap_text=True)
    current_row += 1
    ws_act.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
    ws_act.cell(row=current_row, column=1).value = "от 18.07.2012 г. и подлежит списанию с материально-ответственного лица."
    format_cell(ws_act.cell(row=current_row, column=1), alignment='left')
    current_row += 2

    ws_act['A' + str(current_row)] = "Председатель комиссии:"
    format_cell(ws_act['A' + str(current_row)], bold=True)
    ws_act['C' + str(current_row)] = get_setting("commission_chairman") or ""
    format_cell(ws_act['C' + str(current_row)], alignment='left')
    current_row += 1
    ws_act['A' + str(current_row)] = "Члены комиссии:"
    format_cell(ws_act['A' + str(current_row)], bold=True)
    ws_act['C' + str(current_row)] = get_setting("commission_members") or ""
    format_cell(ws_act['C' + str(current_row)], alignment='left', wrap_text=True)
    current_row += 2
    ws_act['A' + str(current_row)] = "Материально-ответственное лицо:"
    format_cell(ws_act['A' + str(current_row)], bold=True)
    ws_act['C' + str(current_row)] = get_setting("responsible_person") or ""
    format_cell(ws_act['C' + str(current_row)], alignment='left')
    current_row += 1
    ws_act['A' + str(current_row)] = "Заявления проверены:"
    format_cell(ws_act['A' + str(current_row)], bold=True)
    ws_act['C' + str(current_row)] = get_setting("economist") or ""
    format_cell(ws_act['C' + str(current_row)], alignment='left')
    current_row += 1
    ws_act['A' + str(current_row)] = "Дт 210981     Кт 210536                сумма_____________"
    format_cell(ws_act['A' + str(current_row)], bold=True)

    # Журнал
    ws_journal = wb.create_sheet("Журнал")
    headers = ["Дата вып", "Дата сч", "№ сч", "Сумма", "Работа", "Кол-во", "Исполнитель"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_journal.cell(row=1, column=col_idx)
        cell.value = header
        format_cell(cell, bold=True)

    row_idx = 2
    all_works = []
    for works_list in works_by_group.values():
        all_works.extend(works_list)
    for w in all_works:
        ws_journal.cell(row=row_idx, column=1, value=w['date_work'])
        ws_journal.cell(row=row_idx, column=2, value=w['date_invoice'])
        ws_journal.cell(row=row_idx, column=3, value=w['invoice_num'])
        ws_journal.cell(row=row_idx, column=4, value=w['amount'])
        ws_journal.cell(row=row_idx, column=5, value=w['work_name'])
        ws_journal.cell(row=row_idx, column=6, value=w.get('quantity', 1))
        ws_journal.cell(row=row_idx, column=7, value=w.get('tab_number', ''))
        row_idx += 1

    for col in ws_journal.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_journal.column_dimensions[col_letter].width = adjusted_width

    wb.save(output_path)