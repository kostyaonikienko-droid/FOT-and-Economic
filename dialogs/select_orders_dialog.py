# dialogs/select_orders_dialog.py

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QMessageBox, QFileDialog, QLineEdit, QDateEdit, QLabel, QComboBox
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from database.repair_db import get_all_orders, get_material_by_id

class SelectOrdersDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Выбор заказов для акта списания")
        self.resize(900, 600)
        layout = QVBoxLayout(self)

        # Фильтры
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Год:"))
        self.year_combo = QComboBox()
        self.year_combo.addItem("Все", None)
        # добавим годы (заполним позже)
        filter_layout.addWidget(self.year_combo)

        filter_layout.addWidget(QLabel("Поиск:"))
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Номер заказа или заказчик...")
        filter_layout.addWidget(self.search_edit)

        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # Таблица заказов
        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(["Выбрать", "Номер счёта", "Дата", "Заказчик", "ID"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.hideColumn(4)  # скрываем ID
        layout.addWidget(self.table)

        # Кнопки
        btn_layout = QHBoxLayout()
        self.btn_select_all = QPushButton("Выбрать все")
        self.btn_clear_all = QPushButton("Снять выделение")
        self.btn_export = QPushButton("Сформировать акт")
        self.btn_cancel = QPushButton("Отмена")
        btn_layout.addWidget(self.btn_select_all)
        btn_layout.addWidget(self.btn_clear_all)
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_export)
        btn_layout.addWidget(self.btn_cancel)
        layout.addLayout(btn_layout)

        self.btn_select_all.clicked.connect(self.select_all)
        self.btn_clear_all.clicked.connect(self.clear_all)
        self.btn_export.clicked.connect(self.export_act)
        self.btn_cancel.clicked.connect(self.reject)

        self.load_orders()

    def load_orders(self):
        all_orders = get_all_orders()
        # Заполняем годы
        years = set()
        for o in all_orders:
            years.add(o.date.year)
        for y in sorted(years, reverse=True):
            self.year_combo.addItem(str(y), y)

        self.filter_orders()

    def filter_orders(self):
        all_orders = get_all_orders()
        year = self.year_combo.currentData()
        search_text = self.search_edit.text().strip().lower()

        filtered = []
        for o in all_orders:
            if year and o.date.year != year:
                continue
            if search_text and search_text not in o.order_number.lower() and search_text not in o.customer_name.lower():
                continue
            filtered.append(o)

        self.table.setRowCount(len(filtered))
        for i, o in enumerate(filtered):
            chk = QTableWidgetItem()
            chk.setCheckState(Qt.Unchecked)
            self.table.setItem(i, 0, chk)
            self.table.setItem(i, 1, QTableWidgetItem(o.order_number))
            self.table.setItem(i, 2, QTableWidgetItem(o.date.strftime("%d.%m.%Y")))
            self.table.setItem(i, 3, QTableWidgetItem(o.customer_name))
            self.table.setItem(i, 4, QTableWidgetItem(str(o.id)))

        # Сохраним список заказов для быстрого доступа
        self.current_orders = filtered

    def select_all(self):
        for row in range(self.table.rowCount()):
            self.table.item(row, 0).setCheckState(Qt.Checked)

    def clear_all(self):
        for row in range(self.table.rowCount()):
            self.table.item(row, 0).setCheckState(Qt.Unchecked)

    def export_act(self):
        # Получаем выбранные заказы
        selected_ids = []
        for row in range(self.table.rowCount()):
            if self.table.item(row, 0).checkState() == Qt.Checked:
                order_id = int(self.table.item(row, 4).text())
                selected_ids.append(order_id)

        if not selected_ids:
            QMessageBox.warning(self, "Внимание", "Выберите хотя бы один заказ")
            return

        # Собираем материалы из выбранных заказов
        all_orders = get_all_orders()
        selected_orders = [o for o in all_orders if o.id in selected_ids]

        materials_data = []
        for order in selected_orders:
            for work in order.work_items:
                for om in work.materials:
                    mat = get_material_by_id(om.material_id)
                    if mat:
                        materials_data.append({
                            'name': mat.name,
                            'inventory': mat.inventory_number,
                            'unit': mat.unit,
                            'quantity': om.quantity,
                            'reason': f"Израсходовано на {work.description} по счёту №{order.order_number}"
                        })

        if not materials_data:
            QMessageBox.information(self, "Информация", "В выбранных заказах нет материалов")
            return

        # Экспорт в Excel
        fname, _ = QFileDialog.getSaveFileName(self, "Сохранить акт списания",
                                               "Акт_списания.xlsx",
                                               "Excel Files (*.xlsx)")
        if not fname:
            return
        if not fname.endswith('.xlsx'):
            fname += '.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = "Акт списания"

        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = "Прошу списать следующие ТМЦ, использованные в"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')

        headers = ['№ п/п', 'Наименование', 'Инвентарный номер', 'Кол-во', 'Причина списания']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

        total_qty = {}
        for mat in materials_data:
            key = (mat['name'], mat['inventory'], mat['unit'], mat['reason'])
            if key not in total_qty:
                total_qty[key] = 0.0
            total_qty[key] += mat['quantity']

        row = 4
        for i, ((name, inv, unit, reason), qty) in enumerate(total_qty.items(), 1):
            ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=inv).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=4, value=f"{qty:.3f} {unit}").alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=reason).alignment = Alignment(horizontal='center')
            row += 1

        for col in range(1, 6):
            ws.column_dimensions[chr(64+col)].width = max(15, len(headers[col-1]) + 2)

        wb.save(fname)
        QMessageBox.information(self, "Экспорт", f"Акт списания сохранён:\n{fname}")
        self.accept()